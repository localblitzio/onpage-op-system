from __future__ import annotations

import base64
import hashlib
import json
import mimetypes
import os
import re
import secrets
import shutil
import socket
import sqlite3
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.request
from datetime import datetime, timedelta
from html import escape as html_escape
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
ARCHIVE_DIR = DATA_DIR / "archive"
DB_PATH = DATA_DIR / "cora_runs.sqlite3"
QUEUE_PAUSE_PATH = DATA_DIR / "queue_paused.flag"
BRIDGE_SETTINGS_PATH = DATA_DIR / "cloud_bridge.json"
ACTIVITY_LOG_PATH = DATA_DIR / "dashboard_activity.jsonl"
STATIC_DIR = APP_DIR / "static"
DEFAULT_REPORT_DIR = Path.home()
DEFAULT_PORT = int(os.environ.get("CORA_DASHBOARD_PORT", "9191"))
CORA_API_BASE = os.environ.get("CORA_API_BASE", "http://127.0.0.1:9090")
CORA_DEPLOY_DIR = APP_DIR.parent / "SEO Correlation Tool 2026"
CORA_DEPLOY_BAT = CORA_DEPLOY_DIR / "SEO Correlation Tool 2026.bat"
JOB_POLL_SECONDS = 10
JOB_TIMEOUT_SECONDS = 60 * 60 * 2
CORA_FREEZE_SECONDS = int(os.environ.get("CORA_FREEZE_SECONDS", str(60 * 10)))
DEFAULT_JOB_MAX_RETRIES = int(os.environ.get("CORA_JOB_MAX_RETRIES", "1"))
CLOUDFLARE_SYNC_URL = os.environ.get("CLOUDFLARE_SYNC_URL", "").rstrip("/")
CLOUDFLARE_SYNC_TOKEN = os.environ.get("CLOUDFLARE_SYNC_TOKEN", "")
CLOUDFLARE_SYNC_BATCH_SIZE = max(1, min(int(os.environ.get("CLOUDFLARE_SYNC_BATCH_SIZE", "250")), 1000))
CLOUDFLARE_SYNC_WORKBOOK_ROWS = os.environ.get("CLOUDFLARE_SYNC_WORKBOOK_ROWS", "").strip().lower() in {"1", "true", "yes"}
CORA_JOB_LOCK = threading.Lock()
BRIDGE_WORKER_STOP = threading.Event()
BRIDGE_WORKER_LOCK = threading.Lock()
BRIDGE_WORKER_THREAD: threading.Thread | None = None
ENTITY_LSI_WORKERS: dict[int, threading.Thread] = {}
ENTITY_LSI_WORKER_LOCK = threading.Lock()

AI_PROVIDERS: dict[str, dict[str, Any]] = {
    "openai": {
        "name": "OpenAI",
        "placeholder": "sk-...",
        "base_url": "https://api.openai.com",
        "default_model": "gpt-5.5",
        "models": ["gpt-5.5", "gpt-5.4", "gpt-5.4-mini", "gpt-5.4-nano"],
    },
    "anthropic": {
        "name": "Anthropic",
        "placeholder": "sk-ant-...",
        "base_url": "https://api.anthropic.com",
        "default_model": "claude-opus-4-8",
        "models": ["claude-opus-4-8", "claude-sonnet-4-6", "claude-haiku-4-5-20251001"],
    },
    "google": {
        "name": "Google",
        "placeholder": "AIza...",
        "base_url": "https://generativelanguage.googleapis.com",
        "default_model": "gemini-3.5-flash",
        "models": ["gemini-3.5-flash", "gemini-3.1-pro-preview", "gemini-3-flash-preview", "gemini-3.1-flash-lite", "gemini-flash-latest"],
    },
    "xai": {
        "name": "xAI / Grok",
        "placeholder": "xai-...",
        "base_url": "https://api.x.ai",
        "default_model": "grok-4.3",
        "models": ["grok-4.3", "grok-4.3-latest", "grok-latest", "grok-build-0.1", "grok-code-fast"],
    },
    "perplexity": {
        "name": "Perplexity",
        "placeholder": "pplx-...",
        "base_url": "https://api.perplexity.ai",
        "default_model": "perplexity/sonar",
        "models": [
            "perplexity/sonar",
            "openai/gpt-5.4",
            "anthropic/claude-sonnet-4-6",
            "xai/grok-4.3",
            "xai/grok-4.20-reasoning",
            "xai/grok-4.20-non-reasoning",
            "xai/grok-4.20-multi-agent",
        ],
    },
    "dataforseo": {
        "name": "DataForSEO",
        "placeholder": "API password",
        "login_placeholder": "api-login@example.com",
        "base_url": "https://api.dataforseo.com",
        "default_model": "",
        "models": [],
        "auth_type": "basic",
        "test_path": "/v3/appendix/user_data",
    },
}


def queue_paused() -> bool:
    return queue_state().get("paused", False)


def queue_state() -> dict[str, Any]:
    if not QUEUE_PAUSE_PATH.exists():
        return {
            "paused": False,
            "auto_resume": False,
            "stop_after_current": False,
            "paused_at": None,
            "updated_at": None,
            "reason": "",
        }
    try:
        data = json.loads(QUEUE_PAUSE_PATH.read_text(encoding="utf-8"))
        return {
            "paused": bool(data.get("paused", True)),
            "auto_resume": bool(data.get("auto_resume", False)),
            "stop_after_current": bool(data.get("stop_after_current", False)),
            "paused_at": data.get("paused_at"),
            "updated_at": data.get("updated_at"),
            "reason": data.get("reason", ""),
        }
    except Exception:
        return {
            "paused": True,
            "auto_resume": False,
            "stop_after_current": False,
            "paused_at": None,
            "updated_at": None,
            "reason": "Unreadable queue state file",
        }


def set_queue_paused(
    paused: bool,
    auto_resume: bool = False,
    stop_after_current: bool = False,
    reason: str = "",
) -> None:
    ensure_dirs()
    now = datetime.now().isoformat(timespec="seconds")
    if paused:
        log_activity(
            "queue",
            "Queue set to stop after current run" if stop_after_current else ("Queue paused with auto resume" if auto_resume else "Queue paused"),
            "warn" if stop_after_current else "info",
            stop_after_current=bool(stop_after_current),
            auto_resume=bool(auto_resume),
            reason=clean_text(reason),
        )
        QUEUE_PAUSE_PATH.write_text(
            json.dumps(
                {
                    "paused": True,
                    "auto_resume": bool(auto_resume),
                    "stop_after_current": bool(stop_after_current),
                    "paused_at": now,
                    "updated_at": now,
                    "reason": clean_text(reason),
                }
            ),
            encoding="utf-8",
        )
    else:
        log_activity("queue", "Queue resumed", "info")
        if QUEUE_PAUSE_PATH.exists():
            QUEUE_PAUSE_PATH.unlink()


def bridge_settings() -> dict[str, Any]:
    defaults = {
        "enabled": False,
        "allow_cora": False,
        "allow_paid_tools": False,
        "poll_interval": 30,
        "bridge_id": "local-dashboard",
        "updated_at": None,
        "last_poll_at": None,
        "last_result": None,
        "last_error": None,
    }
    if not BRIDGE_SETTINGS_PATH.exists():
        return defaults
    try:
        data = json.loads(BRIDGE_SETTINGS_PATH.read_text(encoding="utf-8"))
        merged = {**defaults, **data}
        merged["poll_interval"] = max(10, min(int(merged.get("poll_interval") or 30), 3600))
        return merged
    except Exception:
        return {**defaults, "last_error": "Unreadable bridge settings file"}


def set_bridge_settings(
    enabled: bool | None = None,
    allow_cora: bool | None = None,
    allow_paid_tools: bool | None = None,
    poll_interval: int | None = None,
) -> dict[str, Any]:
    ensure_dirs()
    settings = bridge_settings()
    if enabled is not None:
        settings["enabled"] = bool(enabled)
    if allow_cora is not None:
        settings["allow_cora"] = bool(allow_cora)
    if allow_paid_tools is not None:
        settings["allow_paid_tools"] = bool(allow_paid_tools)
    if poll_interval is not None:
        settings["poll_interval"] = max(10, min(int(poll_interval), 3600))
    settings["updated_at"] = datetime.now().isoformat(timespec="seconds")
    BRIDGE_SETTINGS_PATH.write_text(json.dumps(settings, indent=2), encoding="utf-8")
    return settings


def save_bridge_runtime_state(result: dict[str, Any] | None = None, error: str | None = None) -> dict[str, Any]:
    settings = bridge_settings()
    settings["last_poll_at"] = datetime.now().isoformat(timespec="seconds")
    settings["last_result"] = result
    settings["last_error"] = error
    settings["updated_at"] = settings["last_poll_at"]
    ensure_dirs()
    BRIDGE_SETTINGS_PATH.write_text(json.dumps(settings, indent=2, default=str), encoding="utf-8")
    return settings


def ensure_dirs() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    ARCHIVE_DIR.mkdir(exist_ok=True)
    STATIC_DIR.mkdir(exist_ok=True)


class ClosingConnection(sqlite3.Connection):
    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> bool:
        result = super().__exit__(exc_type, exc_value, traceback)
        self.close()
        return result


def connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, factory=ClosingConnection)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con


def query_profile_id(query: dict[str, list[str]]) -> int | None:
    raw = (query.get("profile_id") or [""])[0].strip()
    if not raw:
        return None
    try:
        value = int(raw)
    except ValueError:
        return None
    return value if value > 0 else None


def ensure_column(con: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    columns = {row["name"] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db() -> None:
    ensure_dirs()
    with connect() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                target_url TEXT,
                target_domain TEXT,
                search_engine TEXT,
                country TEXT,
                language TEXT,
                cora_version TEXT,
                report_date TEXT,
                imported_at TEXT NOT NULL,
                source_path TEXT NOT NULL,
                archive_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                sha256 TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'imported',
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS serp_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                rank REAL,
                avg_rank REAL,
                weighted_rank REAL,
                data_rank REAL,
                host TEXT,
                title TEXT,
                url TEXT,
                summary TEXT
            );

            CREATE TABLE IF NOT EXISTS recommendations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                section TEXT,
                factor_id TEXT,
                factor TEXT,
                current_value TEXT,
                goal TEXT,
                percent REAL,
                recommendation TEXT,
                shared_correlation TEXT,
                page_one_max TEXT,
                page_one_avg TEXT
            );

            CREATE TABLE IF NOT EXISTS lsi_keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                keyword TEXT NOT NULL,
                spearman REAL,
                pearson REAL,
                best_of_both REAL,
                pages REAL,
                max_value REAL,
                avg_value REAL,
                total REAL,
                tracked_value REAL,
                deficit REAL
            );

            CREATE TABLE IF NOT EXISTS sheet_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                row_json TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS workbook_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                column_count INTEGER NOT NULL,
                row_json TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS managed_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                target_url TEXT NOT NULL,
                target_domain TEXT,
                cora_profile TEXT,
                cora_script TEXT NOT NULL,
                status TEXT NOT NULL,
                status_message TEXT,
                cora_running INTEGER NOT NULL DEFAULT 0,
                cora_action TEXT,
                progress REAL,
                started_at TEXT NOT NULL,
                last_activity_at TEXT,
                retry_count INTEGER NOT NULL DEFAULT 0,
                max_retries INTEGER NOT NULL DEFAULT 1,
                next_retry_at TEXT,
                stall_detected_at TEXT,
                completed_at TEXT,
                report_path TEXT,
                imported_run_id INTEGER REFERENCES runs(id) ON DELETE SET NULL,
                error TEXT
            );

            CREATE TABLE IF NOT EXISTS profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                client TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                profile_id INTEGER REFERENCES profiles(id) ON DELETE SET NULL,
                name TEXT NOT NULL,
                client TEXT,
                notes TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                domain TEXT NOT NULL,
                name TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL REFERENCES sites(id) ON DELETE CASCADE,
                url TEXT NOT NULL,
                title TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,
                page_id INTEGER REFERENCES pages(id) ON DELETE SET NULL,
                keyword TEXT NOT NULL,
                intent TEXT,
                priority TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS api_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                label TEXT NOT NULL,
                key_value TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS content_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,
                page_id INTEGER REFERENCES pages(id) ON DELETE SET NULL,
                keyword_id INTEGER REFERENCES keywords(id) ON DELETE SET NULL,
                title TEXT NOT NULL,
                content_type TEXT,
                intent TEXT,
                priority TEXT,
                status TEXT NOT NULL DEFAULT 'planned',
                due_date TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS entity_lsi_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                batch_id INTEGER REFERENCES entity_lsi_batches(id) ON DELETE CASCADE,
                seed_keyword TEXT NOT NULL,
                depth INTEGER NOT NULL DEFAULT 3,
                provider TEXT NOT NULL,
                provider_key TEXT,
                api_key_id INTEGER REFERENCES api_keys(id) ON DELETE SET NULL,
                model TEXT,
                main_url TEXT,
                prompt_version TEXT NOT NULL,
                prompt TEXT,
                raw_response TEXT,
                result_json TEXT,
                status TEXT NOT NULL DEFAULT 'complete',
                error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS entity_lsi_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                seed_keyword TEXT NOT NULL,
                depth INTEGER NOT NULL DEFAULT 3,
                status TEXT NOT NULL DEFAULT 'running',
                target_count INTEGER NOT NULL DEFAULT 0,
                complete_count INTEGER NOT NULL DEFAULT 0,
                failed_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS entity_sets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                source_batch_id INTEGER REFERENCES entity_lsi_batches(id) ON DELETE SET NULL,
                name TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS entity_set_terms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                set_id INTEGER NOT NULL REFERENCES entity_sets(id) ON DELETE CASCADE,
                term TEXT NOT NULL,
                normalized TEXT NOT NULL,
                type TEXT NOT NULL,
                source_count INTEGER NOT NULL DEFAULT 0,
                sources_json TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(set_id, type, normalized)
            );

            CREATE TABLE IF NOT EXISTS share_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                token TEXT NOT NULL UNIQUE,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                level TEXT NOT NULL DEFAULT 'medium',
                title TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                revoked_at TEXT
            );

            CREATE TABLE IF NOT EXISTS ranking_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
                target TEXT NOT NULL,
                location_code INTEGER NOT NULL DEFAULT 2840,
                language_code TEXT NOT NULL DEFAULT 'en',
                limit_value INTEGER NOT NULL DEFAULT 1000,
                include_subdomains INTEGER NOT NULL DEFAULT 0,
                overview_json TEXT,
                errors_json TEXT,
                source TEXT NOT NULL DEFAULT 'DataForSEO Labs',
                freshness TEXT NOT NULL DEFAULT 'weekly',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ranking_snapshot_keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_id INTEGER NOT NULL REFERENCES ranking_snapshots(id) ON DELETE CASCADE,
                keyword TEXT NOT NULL,
                ranking_url TEXT,
                position REAL,
                previous_position REAL,
                search_volume REAL,
                cpc REAL,
                competition REAL,
                competition_level TEXT,
                keyword_difficulty REAL,
                estimated_traffic REAL,
                traffic_cost REAL,
                serp_features_json TEXT,
                ai_overview_present INTEGER NOT NULL DEFAULT 0,
                ai_overview_reference INTEGER NOT NULL DEFAULT 0,
                intent TEXT,
                last_updated TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ranking_snapshot_pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_id INTEGER NOT NULL REFERENCES ranking_snapshots(id) ON DELETE CASCADE,
                url TEXT NOT NULL,
                organic_keywords REAL,
                organic_traffic REAL,
                organic_traffic_cost REAL,
                top1 REAL,
                top3 REAL,
                top10 REAL,
                top20 REAL,
                top100 REAL,
                paid_keywords REAL,
                paid_traffic REAL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ranking_optimization_targets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_id INTEGER NOT NULL REFERENCES ranking_snapshots(id) ON DELETE CASCADE,
                project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
                url TEXT NOT NULL,
                keyword TEXT,
                best_position REAL,
                ranking_keywords REAL,
                opportunity_count REAL,
                total_search_volume REAL,
                estimated_traffic REAL,
                page_organic_traffic REAL,
                page_organic_keywords REAL,
                top10 REAL,
                priority_type TEXT,
                opportunity_score REAL,
                recommended_action TEXT,
                top_keywords_json TEXT,
                status TEXT NOT NULL DEFAULT 'new',
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(snapshot_id, url)
            );

            CREATE TABLE IF NOT EXISTS cloudflare_sync_state (
                table_name TEXT PRIMARY KEY,
                last_success_at TEXT,
                last_attempt_at TEXT,
                last_row_count INTEGER NOT NULL DEFAULT 0,
                last_error TEXT
            );

            CREATE TABLE IF NOT EXISTS cloud_report_artifacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                share_report_id INTEGER NOT NULL REFERENCES share_reports(id) ON DELETE CASCADE,
                run_id INTEGER REFERENCES runs(id) ON DELETE SET NULL,
                token TEXT NOT NULL,
                artifact_type TEXT NOT NULL,
                file_name TEXT NOT NULL,
                content_type TEXT NOT NULL,
                file_size INTEGER NOT NULL DEFAULT 0,
                sha256 TEXT NOT NULL,
                r2_key TEXT NOT NULL,
                cloud_url TEXT,
                status TEXT NOT NULL DEFAULT 'pending',
                error TEXT,
                uploaded_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(share_report_id, artifact_type)
            );

            CREATE INDEX IF NOT EXISTS idx_runs_keyword ON runs(keyword);
            CREATE INDEX IF NOT EXISTS idx_runs_target_domain ON runs(target_domain);
            CREATE INDEX IF NOT EXISTS idx_serp_run_rank ON serp_results(run_id, rank);
            CREATE INDEX IF NOT EXISTS idx_recommendations_run ON recommendations(run_id);
            CREATE INDEX IF NOT EXISTS idx_lsi_run ON lsi_keywords(run_id);
            CREATE INDEX IF NOT EXISTS idx_workbook_rows_run_sheet ON workbook_rows(run_id, sheet, row_index);
            CREATE INDEX IF NOT EXISTS idx_managed_jobs_status ON managed_jobs(status);
            CREATE INDEX IF NOT EXISTS idx_sites_project ON sites(project_id);
            CREATE INDEX IF NOT EXISTS idx_pages_site ON pages(site_id);
            CREATE INDEX IF NOT EXISTS idx_keywords_project ON keywords(project_id);
            CREATE INDEX IF NOT EXISTS idx_api_keys_provider ON api_keys(provider);
            CREATE INDEX IF NOT EXISTS idx_content_plans_project ON content_plans(project_id);
            CREATE INDEX IF NOT EXISTS idx_content_plans_status ON content_plans(status);
            CREATE INDEX IF NOT EXISTS idx_entity_lsi_runs_project ON entity_lsi_runs(project_id);
            CREATE INDEX IF NOT EXISTS idx_entity_lsi_batches_project ON entity_lsi_batches(project_id);
            CREATE INDEX IF NOT EXISTS idx_entity_sets_project ON entity_sets(project_id);
            CREATE INDEX IF NOT EXISTS idx_entity_set_terms_set ON entity_set_terms(set_id);
            CREATE INDEX IF NOT EXISTS idx_share_reports_token ON share_reports(token);
            CREATE INDEX IF NOT EXISTS idx_share_reports_run ON share_reports(run_id);
            CREATE INDEX IF NOT EXISTS idx_ranking_snapshots_project ON ranking_snapshots(project_id);
            CREATE INDEX IF NOT EXISTS idx_ranking_snapshots_cache ON ranking_snapshots(target, location_code, language_code, include_subdomains, created_at);
            CREATE INDEX IF NOT EXISTS idx_ranking_snapshot_keywords_snapshot ON ranking_snapshot_keywords(snapshot_id);
            CREATE INDEX IF NOT EXISTS idx_ranking_snapshot_pages_snapshot ON ranking_snapshot_pages(snapshot_id);
            CREATE INDEX IF NOT EXISTS idx_ranking_optimization_targets_project ON ranking_optimization_targets(project_id, status);
            CREATE INDEX IF NOT EXISTS idx_ranking_optimization_targets_snapshot ON ranking_optimization_targets(snapshot_id);
            CREATE INDEX IF NOT EXISTS idx_cloud_report_artifacts_report ON cloud_report_artifacts(share_report_id);
            CREATE INDEX IF NOT EXISTS idx_cloud_report_artifacts_status ON cloud_report_artifacts(status, uploaded_at);
            """
        )
        ensure_column(con, "runs", "project_id", "INTEGER")
        ensure_column(con, "runs", "site_id", "INTEGER")
        ensure_column(con, "runs", "page_id", "INTEGER")
        ensure_column(con, "runs", "keyword_id", "INTEGER")
        ensure_column(con, "projects", "profile_id", "INTEGER")
        ensure_column(con, "managed_jobs", "cora_profile", "TEXT")
        ensure_column(con, "managed_jobs", "project_id", "INTEGER")
        ensure_column(con, "managed_jobs", "keyword_id", "INTEGER")
        ensure_column(con, "managed_jobs", "tool", "TEXT")
        ensure_column(con, "managed_jobs", "last_activity_at", "TEXT")
        ensure_column(con, "managed_jobs", "retry_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(con, "managed_jobs", "max_retries", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(con, "managed_jobs", "next_retry_at", "TEXT")
        ensure_column(con, "managed_jobs", "stall_detected_at", "TEXT")
        ensure_column(con, "api_keys", "base_url", "TEXT")
        ensure_column(con, "api_keys", "default_model", "TEXT")
        ensure_column(con, "api_keys", "status", "TEXT NOT NULL DEFAULT 'untested'")
        ensure_column(con, "api_keys", "last_tested_at", "TEXT")
        ensure_column(con, "api_keys", "last_error", "TEXT")
        ensure_column(con, "entity_lsi_runs", "batch_id", "INTEGER")
        ensure_column(con, "share_reports", "ranking_snapshot_id", "INTEGER")
        ensure_column(con, "share_reports", "entity_set_id", "INTEGER")
        ensure_column(con, "share_reports", "optimization_target_ids_json", "TEXT")
        con.execute(
            """
            UPDATE ranking_optimization_targets
            SET project_id = (
                SELECT rs.project_id
                FROM ranking_snapshots rs
                WHERE rs.id = ranking_optimization_targets.snapshot_id
            )
            WHERE EXISTS (
                SELECT 1
                FROM ranking_snapshots rs
                WHERE rs.id = ranking_optimization_targets.snapshot_id
                  AND COALESCE(rs.project_id, 0) != COALESCE(ranking_optimization_targets.project_id, 0)
            )
            """
        )
        con.execute("CREATE INDEX IF NOT EXISTS idx_projects_profile ON projects(profile_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_managed_jobs_project ON managed_jobs(project_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_managed_jobs_keyword ON managed_jobs(keyword_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_entity_lsi_runs_batch ON entity_lsi_runs(batch_id)")


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    if row is None:
        return None
    return {k: row[k] for k in row.keys()}


CLOUDFLARE_SYNC_TABLES = [
    "profiles",
    "projects",
    "sites",
    "pages",
    "keywords",
    "runs",
    "serp_results",
    "recommendations",
    "lsi_keywords",
    "sheet_rows",
    "managed_jobs",
    "content_plans",
    "entity_lsi_batches",
    "entity_lsi_runs",
    "entity_sets",
    "entity_set_terms",
    "share_reports",
    "ranking_snapshots",
    "ranking_snapshot_keywords",
    "ranking_snapshot_pages",
    "ranking_optimization_targets",
]


def cloudflare_sync_tables() -> list[str]:
    tables = list(CLOUDFLARE_SYNC_TABLES)
    if CLOUDFLARE_SYNC_WORKBOOK_ROWS:
        tables.append("workbook_rows")
    return tables


def cloudflare_sync_configured() -> bool:
    return bool(CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN)


def cloudflare_sync_state() -> dict[str, Any]:
    with connect() as con:
        rows = con.execute(
            """
            SELECT *
            FROM cloudflare_sync_state
            ORDER BY table_name
            """
        ).fetchall()
        counts = {}
        for table in cloudflare_sync_tables():
            try:
                counts[table] = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            except sqlite3.Error:
                counts[table] = 0
        artifacts = cloudflare_artifact_state(con)
    return {
        "configured": cloudflare_sync_configured(),
        "sync_url": CLOUDFLARE_SYNC_URL,
        "has_token": bool(CLOUDFLARE_SYNC_TOKEN),
        "batch_size": CLOUDFLARE_SYNC_BATCH_SIZE,
        "includes_workbook_rows": CLOUDFLARE_SYNC_WORKBOOK_ROWS,
        "counts": counts,
        "state": [row_to_dict(row) for row in rows],
        "artifacts": artifacts,
        "bridge": bridge_status(),
    }


def cloudflare_artifact_state(con: sqlite3.Connection | None = None) -> dict[str, Any]:
    close = False
    if con is None:
        con = connect()
        close = True
    try:
        counts = [
            row_to_dict(r) or {}
            for r in con.execute(
                """
                SELECT artifact_type, status, COUNT(*) AS count, COALESCE(SUM(file_size), 0) AS total_bytes, MAX(uploaded_at) AS last_uploaded_at
                FROM cloud_report_artifacts
                GROUP BY artifact_type, status
                ORDER BY artifact_type, status
                """
            ).fetchall()
        ]
        summary = row_to_dict(
            con.execute(
                """
                SELECT COUNT(*) AS total_files, COALESCE(SUM(file_size), 0) AS total_bytes, MAX(uploaded_at) AS last_uploaded_at
                FROM cloud_report_artifacts
                WHERE status = 'synced'
                """
            ).fetchone()
        ) or {}
        return {
            "counts": counts,
            "total_files": int(summary.get("total_files") or 0),
            "total_bytes": int(summary.get("total_bytes") or 0),
            "last_uploaded_at": summary.get("last_uploaded_at"),
        }
    finally:
        if close:
            con.close()


def cloudflare_table_rows(con: sqlite3.Connection, table: str, limit: int, offset: int) -> list[dict[str, Any]]:
    if table not in cloudflare_sync_tables():
        raise ValueError(f"Unsupported Cloudflare sync table: {table}")
    rows = con.execute(f"SELECT * FROM {table} ORDER BY id LIMIT ? OFFSET ?", (limit, offset)).fetchall()
    return [row_to_dict(row) or {} for row in rows]


def local_table_columns(con: sqlite3.Connection, table: str) -> list[str]:
    return [str(row["name"]) for row in con.execute(f"PRAGMA table_info({table})").fetchall()]


def upsert_local_rows(con: sqlite3.Connection, table: str, rows: list[dict[str, Any]]) -> int:
    if table not in cloudflare_sync_tables():
        raise ValueError(f"Unsupported Cloudflare sync table: {table}")
    if not rows:
        return 0
    prepared_rows = [prepare_cloudflare_row_for_local(table, row) for row in rows]
    columns = [column for column in local_table_columns(con, table) if column in prepared_rows[0]]
    if "id" not in columns:
        return 0
    placeholders = ", ".join("?" for _ in columns)
    updates = ", ".join(f"{column}=excluded.{column}" for column in columns if column != "id")
    sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders}) ON CONFLICT(id) DO UPDATE SET {updates}"
    count = 0
    for row in prepared_rows:
        con.execute(sql, [row.get(column) for column in columns])
        count += 1
    return count


def prepare_cloudflare_row_for_local(table: str, row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    if table == "entity_lsi_batches":
        if "completed_count" in item and "complete_count" not in item:
            item["complete_count"] = item.get("completed_count")
        item["complete_count"] = int(item.get("complete_count") or 0)
        item["failed_count"] = int(item.get("failed_count") or 0)
    elif table == "entity_lsi_runs":
        if item.get("provider") and not item.get("provider_key"):
            item["provider_key"] = normalize_ai_provider(item.get("provider")) or clean_text(item.get("provider"))
        if item.get("completed_at") and not item.get("updated_at"):
            item["updated_at"] = item.get("completed_at")
        if not item.get("updated_at"):
            item["updated_at"] = item.get("created_at") or datetime.now().isoformat(timespec="seconds")
        item["prompt_version"] = item.get("prompt_version") or "cloud-v1"
        item["prompt"] = item.get("prompt") or ""
        item["main_url"] = item.get("main_url") or ""
        item["status"] = item.get("status") or "complete"
        if not item.get("result_json"):
            def parse_json_field(name: str) -> list[Any]:
                try:
                    value = json.loads(item.get(name) or "[]")
                    return value if isinstance(value, list) else []
                except json.JSONDecodeError:
                    return []

            result = {
                "summary": item.get("summary") or "",
                "entities": parse_json_field("entities_json"),
                "lsi_keywords": parse_json_field("lsi_keywords_json"),
                "related_keywords": parse_json_field("related_keywords_json"),
                "questions": parse_json_field("questions_json"),
                "topics": parse_json_field("topics_json"),
                "warnings": [],
            }
            item["result_json"] = json.dumps(result)
    return item


def pull_cloudflare_sync(tables: list[str] | None = None, limit: int = 5000) -> dict[str, Any]:
    selected = [table for table in (tables or ["profiles", "projects", "sites", "keywords", "content_plans", "ranking_snapshots", "ranking_snapshot_keywords", "ranking_snapshot_pages", "ranking_optimization_targets", "entity_sets", "entity_set_terms", "share_reports"]) if table]
    table_query = quote(",".join(selected))
    data = cloudflare_get_json(f"/api/sync/export?tables={table_query}&limit={int(limit)}")
    exported = data.get("tables") if isinstance(data.get("tables"), list) else []
    results = []
    with connect() as con:
        for item in exported:
            table = str(item.get("table") or "")
            rows = item.get("rows") if isinstance(item.get("rows"), list) else []
            count = upsert_local_rows(con, table, rows)
            results.append({"table": table, "rows": count})
        con.commit()
    return {"ok": True, "direction": "cloud_to_local", "tables": results, "total_rows": sum(int(item["rows"]) for item in results)}


def post_cloudflare_sync_batch(payload: dict[str, Any]) -> dict[str, Any]:
    if not cloudflare_sync_configured():
        raise ValueError("Cloudflare sync is not configured. Set CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN.")
    body = json.dumps(payload, default=str).encode("utf-8")
    request = urllib.request.Request(
        f"{CLOUDFLARE_SYNC_URL}/api/sync/push",
        data=body,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {CLOUDFLARE_SYNC_TOKEN}",
            "User-Agent": "OnPageOptimizationSystemDashboard/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ValueError(f"Cloudflare sync failed HTTP {exc.code}: {detail}") from exc


def cloudflare_request_json(path: str, payload: dict[str, Any], timeout: int = 120) -> dict[str, Any]:
    if not cloudflare_sync_configured():
        raise ValueError("Cloudflare sync is not configured. Set CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN.")
    body = json.dumps(payload, default=str).encode("utf-8")
    request = urllib.request.Request(
        f"{CLOUDFLARE_SYNC_URL}{path}",
        data=body,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {CLOUDFLARE_SYNC_TOKEN}",
            "User-Agent": "OnPageOptimizationSystemDashboard/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ValueError(f"Cloudflare request failed HTTP {exc.code}: {detail}") from exc


def cloudflare_get_json(path: str, timeout: int = 120) -> dict[str, Any]:
    if not cloudflare_sync_configured():
        raise ValueError("Cloudflare sync is not configured. Set CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN.")
    request = urllib.request.Request(
        f"{CLOUDFLARE_SYNC_URL}{path}",
        headers={
            "Accept": "application/json",
            "Authorization": f"Bearer {CLOUDFLARE_SYNC_TOKEN}",
            "User-Agent": "OnPageOptimizationSystemDashboard/1.0",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ValueError(f"Cloudflare request failed HTTP {exc.code}: {detail}") from exc


def update_cloudflare_command(command_id: int, status: str, result: dict[str, Any] | None = None, error: str | None = None) -> dict[str, Any]:
    return cloudflare_request_json(
        f"/api/commands/{command_id}",
        {"status": status, "result": result or {}, "error": error},
    )


def find_existing_project(name: str, site_domain: str | None = None) -> dict[str, Any] | None:
    domain = domain_from_url(site_domain or "") if site_domain else None
    with connect() as con:
        row = con.execute("SELECT * FROM projects WHERE lower(name) = lower(?) ORDER BY id LIMIT 1", (name.strip(),)).fetchone()
        if row:
            return row_to_dict(row)
        if domain:
            row = con.execute(
                """
                SELECT p.*
                FROM projects p
                JOIN sites s ON s.project_id = p.id
                WHERE lower(s.domain) = lower(?)
                ORDER BY p.id
                LIMIT 1
                """,
                (domain,),
            ).fetchone()
            return row_to_dict(row)
    return None


def find_existing_keyword(project_id: int, keyword: str) -> dict[str, Any] | None:
    with connect() as con:
        return row_to_dict(
            con.execute(
                "SELECT * FROM keywords WHERE project_id = ? AND lower(keyword) = lower(?) ORDER BY id LIMIT 1",
                (project_id, keyword.strip()),
            ).fetchone()
        )


def find_existing_content_plan(project_id: int, title: str) -> dict[str, Any] | None:
    with connect() as con:
        return row_to_dict(
            con.execute(
                "SELECT * FROM content_plans WHERE project_id = ? AND lower(title) = lower(?) ORDER BY id LIMIT 1",
                (project_id, title.strip()),
            ).fetchone()
        )


def find_existing_share_report(run_id: int, level: str, title: str | None = None) -> dict[str, Any] | None:
    with connect() as con:
        if clean_text(title):
            row = con.execute(
                """
                SELECT * FROM share_reports
                WHERE run_id = ? AND level = ? AND lower(COALESCE(title, '')) = lower(?)
                  AND revoked_at IS NULL
                ORDER BY id LIMIT 1
                """,
                (run_id, level, clean_text(title)),
            ).fetchone()
        else:
            row = con.execute(
                "SELECT * FROM share_reports WHERE run_id = ? AND level = ? AND revoked_at IS NULL ORDER BY id LIMIT 1",
                (run_id, level),
            ).fetchone()
        return row_to_dict(row)


def find_existing_pending_job(project_id: int | None, keyword: str, target_url: str) -> dict[str, Any] | None:
    target_domain = domain_from_url(target_url)
    with connect() as con:
        return row_to_dict(
            con.execute(
                """
                SELECT * FROM managed_jobs
                WHERE COALESCE(project_id, 0) = COALESCE(?, 0)
                  AND lower(keyword) = lower(?)
                  AND lower(target_domain) = lower(?)
                  AND status IN ('queued', 'running', 'waiting', 'submitted')
                ORDER BY id DESC LIMIT 1
                """,
                (project_id, keyword.strip(), target_domain or ""),
            ).fetchone()
        )


def apply_cloudflare_command(command: dict[str, Any]) -> dict[str, Any]:
    command_type = command.get("command_type") or ""
    payload = command.get("payload") if isinstance(command.get("payload"), dict) else {}
    settings = bridge_settings()
    if command_type == "run_cora" and not settings.get("allow_cora"):
        raise ValueError("Cloud bridge is not allowed to queue Cora runs. Enable Allow Cora in local bridge settings.")
    paid_tool_commands = {"create_ranking_snapshot", "run_entity_lsi"}
    if command_type in paid_tool_commands and not bool(payload.get("dry_run")) and not settings.get("allow_paid_tools"):
        raise ValueError("Cloud bridge is not allowed to run paid/API tools. Enable paid tools in local bridge settings.")
    changed_tables: set[str] = set()
    artifact_report_ids: list[int] = []
    result: dict[str, Any] = {"command_type": command_type}
    if command_type == "create_project":
        name = clean_text(payload.get("name")) or ""
        existing = find_existing_project(name, payload.get("site_domain"))
        project = existing or create_project(
            name,
            client=payload.get("client"),
            site_domain=payload.get("site_domain"),
            notes=payload.get("notes"),
            profile_name=payload.get("profile_name"),
        )
        result["duplicate"] = bool(existing)
        result["project"] = project
        changed_tables.update({"profiles", "projects", "sites"})
    elif command_type == "add_keyword":
        project_id = int(payload.get("project_id") or 0)
        keyword_text = clean_text(payload.get("keyword")) or ""
        existing = find_existing_keyword(project_id, keyword_text)
        keyword = existing or create_keyword(
            project_id,
            keyword_text,
            int(payload["site_id"]) if payload.get("site_id") else None,
            int(payload["page_id"]) if payload.get("page_id") else None,
            payload.get("intent"),
            payload.get("priority"),
        )
        result["duplicate"] = bool(existing)
        result["keyword"] = keyword
        changed_tables.add("keywords")
    elif command_type == "create_content_plan":
        project_id = int(payload.get("project_id") or 0)
        title = clean_text(payload.get("title")) or ""
        existing = find_existing_content_plan(project_id, title)
        plan = existing or create_content_plan(
            project_id,
            title,
            int(payload["site_id"]) if payload.get("site_id") else None,
            int(payload["page_id"]) if payload.get("page_id") else None,
            int(payload["keyword_id"]) if payload.get("keyword_id") else None,
            payload.get("content_type"),
            payload.get("intent"),
            payload.get("priority"),
            payload.get("status"),
            payload.get("due_date"),
            payload.get("notes"),
        )
        result["duplicate"] = bool(existing)
        result["content_plan"] = plan
        changed_tables.add("content_plans")
    elif command_type == "create_share_report":
        run_id = int(payload.get("run_id") or 0)
        level = payload.get("level") or "medium"
        existing = find_existing_share_report(run_id, level, payload.get("title"))
        report = existing or create_share_report(
            run_id,
            level,
            payload.get("title"),
            payload.get("notes"),
            int(payload["ranking_snapshot_id"]) if payload.get("ranking_snapshot_id") else None,
            [int(value) for value in payload.get("optimization_target_ids", []) if value] if isinstance(payload.get("optimization_target_ids"), list) else [],
            int(payload["entity_set_id"]) if payload.get("entity_set_id") else None,
        )
        result["duplicate"] = bool(existing)
        result["report"] = report
        changed_tables.add("share_reports")
        artifact_report_ids.append(int(report["id"]))
    elif command_type == "run_cora":
        project_id = int(payload.get("project_id") or 0) or None
        keyword_text = clean_text(payload.get("keyword")) or ""
        keyword_id = int(payload["keyword_id"]) if payload.get("keyword_id") else None
        if project_id and keyword_text and not keyword_id:
            keyword_row = find_existing_keyword(project_id, keyword_text) or create_keyword(project_id, keyword_text)
            keyword_id = int(keyword_row["id"])
            changed_tables.add("keywords")
        existing = find_existing_pending_job(project_id, keyword_text, clean_text(payload.get("target_url")) or "")
        job = existing or create_managed_job(
            keyword_text,
            clean_text(payload.get("target_url")) or "",
            payload.get("cora_profile"),
            project_id=project_id,
            keyword_id=keyword_id,
            tool="cora",
        )
        result["duplicate"] = bool(existing)
        result["job"] = job
        changed_tables.add("managed_jobs")
    elif command_type == "sync_cloud_data":
        tables = payload.get("tables")
        selected_tables = [str(table) for table in tables if table] if isinstance(tables, list) else None
        result["sync"] = push_cloudflare_sync(tables=selected_tables, dry_run=bool(payload.get("dry_run")))
    elif command_type == "sync_cloud_to_local":
        tables = payload.get("tables")
        selected_tables = [str(table) for table in tables if table] if isinstance(tables, list) else None
        if bool(payload.get("dry_run")):
            result["sync"] = {"ok": True, "direction": "cloud_to_local", "dry_run": True, "tables": selected_tables or ["profiles", "projects", "sites", "keywords", "content_plans", "ranking_snapshots", "ranking_snapshot_keywords", "ranking_snapshot_pages", "ranking_optimization_targets", "entity_sets", "entity_set_terms", "share_reports"]}
        else:
            result["sync"] = pull_cloudflare_sync(tables=selected_tables, limit=int(payload.get("limit") or 5000))
    elif command_type == "sync_report_artifacts":
        report_ids = payload.get("report_ids")
        selected_report_ids = [int(value) for value in report_ids if value] if isinstance(report_ids, list) else None
        result["artifacts"] = sync_cloudflare_report_artifacts(
            report_ids=selected_report_ids,
            dry_run=bool(payload.get("dry_run")),
            force=bool(payload.get("force")),
        )
    elif command_type == "create_ranking_snapshot":
        if bool(payload.get("dry_run")):
            result["dry_run"] = True
            result["snapshot_request"] = {
                "project_id": int(payload["project_id"]) if payload.get("project_id") else None,
                "target": normalize_ranking_snapshot_target(payload.get("target")),
                "location_code": int(payload.get("location_code") or 2840),
                "language_code": clean_text(payload.get("language_code")) or "en",
                "limit": max(1, min(int(payload.get("limit") or 1000), 1000)),
                "include_subdomains": bool(payload.get("include_subdomains")),
                "force_refresh": bool(payload.get("force_refresh")),
            }
        else:
            snapshot = create_or_get_ranking_snapshot(payload)
            result["snapshot"] = snapshot.get("snapshot")
            result["meta"] = snapshot.get("meta")
            changed_tables.update({"ranking_snapshots", "ranking_snapshot_keywords", "ranking_snapshot_pages"})
    elif command_type == "run_entity_lsi":
        targets = payload.get("targets") if isinstance(payload.get("targets"), list) else []
        if bool(payload.get("dry_run")):
            result["dry_run"] = True
            result["entity_lsi_request"] = {
                "project_id": int(payload.get("project_id") or 0),
                "seed_keyword": clean_text(payload.get("seed_keyword")) or "",
                "depth": clamp_entity_depth(payload.get("depth", 3)),
                "target_count": len(targets),
                "run_async": bool(payload.get("run_async", True)),
            }
        else:
            runs = create_entity_lsi_runs(
                int(payload.get("project_id") or 0),
                clean_text(payload.get("seed_keyword")) or "",
                payload.get("depth", 3),
                targets,
                bool(payload.get("run_async", True)),
            )
            result["runs"] = runs
            result["batch_id"] = runs[0].get("batch_id") if runs else None
            changed_tables.update({"entity_lsi_batches", "entity_lsi_runs"})
    else:
        raise ValueError(f"Unsupported cloud command type: {command_type}")
    if changed_tables and cloudflare_sync_configured():
        result["sync"] = push_cloudflare_sync(tables=sorted(changed_tables), dry_run=False)
    if artifact_report_ids and cloudflare_sync_configured():
        result["artifacts"] = sync_cloudflare_report_artifacts(report_ids=artifact_report_ids, dry_run=False, force=True)
    return result


def pull_cloudflare_commands(limit: int = 25) -> dict[str, Any]:
    data = cloudflare_get_json(f"/api/commands?status=pending&limit={int(limit)}")
    commands = data.get("commands") if isinstance(data.get("commands"), list) else []
    results = []
    for command in commands:
        command_id = int(command.get("id") or 0)
        if not command_id:
            continue
        try:
            update_cloudflare_command(command_id, "claimed")
            applied = apply_cloudflare_command(command)
            update_cloudflare_command(command_id, "complete", result=applied)
            results.append({"id": command_id, "status": "complete", "result": applied})
        except Exception as exc:
            try:
                update_cloudflare_command(command_id, "failed", error=str(exc))
            except Exception:
                pass
            results.append({"id": command_id, "status": "failed", "error": str(exc)})
    return {"ok": all(item["status"] == "complete" for item in results), "processed": len(results), "commands": results}


def send_bridge_heartbeat(result: dict[str, Any] | None = None, status: str = "online") -> dict[str, Any]:
    settings = bridge_settings()
    payload = {
        "bridge_id": settings.get("bridge_id") or "local-dashboard",
        "status": status,
        "version": "local-dashboard",
        "allow_cora": bool(settings.get("allow_cora")),
        "allow_paid_tools": bool(settings.get("allow_paid_tools")),
        "poll_interval": int(settings.get("poll_interval") or 30),
        "last_poll_at": settings.get("last_poll_at"),
        "last_result": result if result is not None else settings.get("last_result"),
    }
    return cloudflare_request_json("/api/bridge/heartbeat", payload)


def bridge_status() -> dict[str, Any]:
    settings = bridge_settings()
    return {
        "configured": cloudflare_sync_configured(),
        **settings,
    }


def cloud_bridge_loop() -> None:
    while not BRIDGE_WORKER_STOP.is_set():
        settings = bridge_settings()
        interval = int(settings.get("poll_interval") or 30)
        if settings.get("enabled") and cloudflare_sync_configured():
            try:
                result = pull_cloudflare_commands(limit=25)
                save_bridge_runtime_state(result=result, error=None)
                send_bridge_heartbeat(result=result)
            except Exception as exc:
                save_bridge_runtime_state(result=None, error=str(exc))
                try:
                    send_bridge_heartbeat(result={"ok": False, "error": str(exc)}, status="error")
                except Exception:
                    pass
        BRIDGE_WORKER_STOP.wait(max(10, interval))


def ensure_bridge_worker() -> None:
    global BRIDGE_WORKER_THREAD
    with BRIDGE_WORKER_LOCK:
        if BRIDGE_WORKER_THREAD and BRIDGE_WORKER_THREAD.is_alive():
            return
        BRIDGE_WORKER_STOP.clear()
        BRIDGE_WORKER_THREAD = threading.Thread(target=cloud_bridge_loop, name="cloud-bridge", daemon=True)
        BRIDGE_WORKER_THREAD.start()


def report_artifact_key(token: str, artifact_type: str, file_name: str) -> str:
    safe_token = re.sub(r"[^A-Za-z0-9_-]+", "-", token).strip("-") or secrets.token_urlsafe(8)
    suffix = Path(file_name).suffix.lower() or (".html" if artifact_type == "report_html" else ".bin")
    base = "report" if artifact_type == "report_html" else "source-cora-report"
    return f"reports/{safe_token}/{base}{suffix}"


def collect_share_report_artifacts(report_id: int, cloud_base_url: str | None = None) -> list[dict[str, Any]]:
    with connect() as con:
        report = row_to_dict(
            con.execute(
                """
                SELECT sr.*, r.archive_path, r.file_name, r.sha256 AS run_sha256
                FROM share_reports sr
                JOIN runs r ON r.id = sr.run_id
                WHERE sr.id = ? AND sr.revoked_at IS NULL
                """,
                (report_id,),
            ).fetchone()
        )
    if not report:
        raise ValueError("Shared report not found")
    token = str(report["token"])
    base_url = (cloud_base_url or CLOUDFLARE_SYNC_URL or "").rstrip("/")
    data = shared_report_by_token(token)
    html_body = render_shared_report_html(data, base_url)
    artifacts = [
        {
            "share_report_id": report_id,
            "run_id": int(report["run_id"]),
            "token": token,
            "artifact_type": "report_html",
            "file_name": f"{token}.html",
            "content_type": "text/html; charset=utf-8",
            "body": html_body,
        }
    ]
    archive_path = Path(report["archive_path"] or "")
    if archive_path.exists() and archive_path.is_file():
        artifacts.append(
            {
                "share_report_id": report_id,
                "run_id": int(report["run_id"]),
                "token": token,
                "artifact_type": "source_xlsx",
                "file_name": report.get("file_name") or archive_path.name,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "body": archive_path.read_bytes(),
            }
        )
    return artifacts


def upsert_cloud_report_artifact(artifact: dict[str, Any], status: str = "pending", error: str | None = None, cloud_url: str | None = None) -> dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    body = artifact["body"]
    sha = hashlib.sha256(body).hexdigest()
    file_name = artifact["file_name"]
    r2_key = report_artifact_key(artifact["token"], artifact["artifact_type"], file_name)
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO cloud_report_artifacts
            (share_report_id, run_id, token, artifact_type, file_name, content_type, file_size, sha256, r2_key, cloud_url, status, error, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(share_report_id, artifact_type) DO UPDATE SET
              run_id=excluded.run_id,
              token=excluded.token,
              file_name=excluded.file_name,
              content_type=excluded.content_type,
              file_size=excluded.file_size,
              sha256=excluded.sha256,
              r2_key=excluded.r2_key,
              cloud_url=COALESCE(excluded.cloud_url, cloud_report_artifacts.cloud_url),
              status=excluded.status,
              error=excluded.error,
              updated_at=excluded.updated_at
            """,
            (
                artifact["share_report_id"],
                artifact["run_id"],
                artifact["token"],
                artifact["artifact_type"],
                file_name,
                artifact["content_type"],
                len(body),
                sha,
                r2_key,
                cloud_url,
                status,
                error,
                now,
                now,
            ),
        )
        row = row_to_dict(
            con.execute(
                "SELECT * FROM cloud_report_artifacts WHERE share_report_id = ? AND artifact_type = ?",
                (artifact["share_report_id"], artifact["artifact_type"]),
            ).fetchone()
        ) or {}
    return row


def mark_cloud_report_artifact_synced(local_id: int, response: dict[str, Any]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        con.execute(
            """
            UPDATE cloud_report_artifacts
            SET status = 'synced',
                error = NULL,
                cloud_url = ?,
                uploaded_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (response.get("public_url"), now, now, local_id),
        )


def mark_cloud_report_artifact_failed(local_id: int, error: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        con.execute(
            """
            UPDATE cloud_report_artifacts
            SET status = 'failed', error = ?, updated_at = ?
            WHERE id = ?
            """,
            (error[:1000], now, local_id),
        )


def sync_cloudflare_report_artifacts(report_ids: list[int] | None = None, dry_run: bool = False, force: bool = False) -> dict[str, Any]:
    with connect() as con:
        if report_ids:
            placeholders = ",".join("?" for _ in report_ids)
            rows = con.execute(
                f"SELECT id FROM share_reports WHERE revoked_at IS NULL AND id IN ({placeholders}) ORDER BY created_at DESC, id DESC",
                [int(value) for value in report_ids],
            ).fetchall()
        else:
            rows = con.execute("SELECT id FROM share_reports WHERE revoked_at IS NULL ORDER BY created_at DESC, id DESC").fetchall()
    selected_ids = [int(row["id"]) for row in rows]
    if not dry_run and not cloudflare_sync_configured():
        raise ValueError("Cloudflare sync is not configured. Set CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN.")
    started_at = datetime.now().isoformat(timespec="seconds")
    result: dict[str, Any] = {
        "ok": True,
        "dry_run": dry_run,
        "started_at": started_at,
        "reports": len(selected_ids),
        "artifacts": [],
        "uploaded": 0,
        "skipped": 0,
        "failed": 0,
        "total_bytes": 0,
    }
    for report_id in selected_ids:
        for artifact in collect_share_report_artifacts(report_id, CLOUDFLARE_SYNC_URL):
            body = artifact["body"]
            artifact_sha = hashlib.sha256(body).hexdigest()
            existing = None
            with connect() as con:
                existing = row_to_dict(
                    con.execute(
                        "SELECT * FROM cloud_report_artifacts WHERE share_report_id = ? AND artifact_type = ?",
                        (report_id, artifact["artifact_type"]),
                    ).fetchone()
                )
            if existing and existing.get("status") == "synced" and existing.get("sha256") == artifact_sha and not force:
                result["skipped"] += 1
                result["artifacts"].append({"id": existing["id"], "type": artifact["artifact_type"], "status": "skipped", "file_name": artifact["file_name"]})
                continue
            result["total_bytes"] += len(body)
            if dry_run:
                result["artifacts"].append({"id": (existing or {}).get("id"), "type": artifact["artifact_type"], "status": "dry_run", "file_name": artifact["file_name"], "bytes": len(body)})
                continue
            record = upsert_cloud_report_artifact(artifact, "pending")
            payload = {
                "local_id": str(record["id"]),
                "share_report_id": record["share_report_id"],
                "run_id": record["run_id"],
                "token": record["token"],
                "artifact_type": record["artifact_type"],
                "file_name": record["file_name"],
                "content_type": record["content_type"],
                "file_size": record["file_size"],
                "sha256": record["sha256"],
                "r2_key": record["r2_key"],
                "content_base64": base64.b64encode(body).decode("ascii"),
            }
            try:
                response = cloudflare_request_json("/api/artifacts/upload", payload, timeout=180)
                mark_cloud_report_artifact_synced(int(record["id"]), response)
                result["uploaded"] += 1
                result["artifacts"].append({"id": record["id"], "type": artifact["artifact_type"], "status": "synced", "file_name": artifact["file_name"], "url": response.get("public_url")})
            except Exception as exc:
                mark_cloud_report_artifact_failed(int(record["id"]), str(exc))
                result["failed"] += 1
                result["ok"] = False
                result["artifacts"].append({"id": record["id"], "type": artifact["artifact_type"], "status": "failed", "file_name": artifact["file_name"], "error": str(exc)})
    result["finished_at"] = datetime.now().isoformat(timespec="seconds")
    result["state"] = cloudflare_artifact_state()
    return result


def update_cloudflare_sync_state(table: str, row_count: int, error: str | None = None) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        con.execute(
            """
            INSERT INTO cloudflare_sync_state (table_name, last_success_at, last_attempt_at, last_row_count, last_error)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(table_name) DO UPDATE SET
              last_success_at = excluded.last_success_at,
              last_attempt_at = excluded.last_attempt_at,
              last_row_count = excluded.last_row_count,
              last_error = excluded.last_error
            """,
            (table, None if error else now, now, row_count, error),
        )


def push_cloudflare_sync(tables: list[str] | None = None, dry_run: bool = False) -> dict[str, Any]:
    selected_tables = tables or cloudflare_sync_tables()
    unsupported = [table for table in selected_tables if table not in cloudflare_sync_tables()]
    if unsupported:
        raise ValueError(f"Unsupported Cloudflare sync table(s): {', '.join(unsupported)}")
    if not dry_run and not cloudflare_sync_configured():
        raise ValueError("Cloudflare sync is not configured. Set CLOUDFLARE_SYNC_URL and CLOUDFLARE_SYNC_TOKEN.")
    started_at = datetime.now().isoformat(timespec="seconds")
    results: list[dict[str, Any]] = []
    total_rows = 0
    with connect() as con:
        for table in selected_tables:
            count = int(con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
            total_rows += count
            if dry_run:
                results.append({"table": table, "rows": count, "batches": 0, "status": "dry_run"})
                continue
            batches = 0
            try:
                for offset in range(0, count, CLOUDFLARE_SYNC_BATCH_SIZE):
                    rows = cloudflare_table_rows(con, table, CLOUDFLARE_SYNC_BATCH_SIZE, offset)
                    response = post_cloudflare_sync_batch(
                        {
                            "source": "on-page-optimization-system",
                            "table": table,
                            "rows": rows,
                            "offset": offset,
                            "total": count,
                            "sent_at": datetime.now().isoformat(timespec="seconds"),
                        }
                    )
                    if response.get("ok") is False:
                        raise ValueError(response.get("error") or "Cloudflare sync batch failed")
                    batches += 1
                update_cloudflare_sync_state(table, count)
                results.append({"table": table, "rows": count, "batches": batches, "status": "synced"})
            except Exception as exc:
                update_cloudflare_sync_state(table, count, str(exc))
                results.append({"table": table, "rows": count, "batches": batches, "status": "failed", "error": str(exc)})
                raise
    return {
        "ok": True,
        "dry_run": dry_run,
        "started_at": started_at,
        "finished_at": datetime.now().isoformat(timespec="seconds"),
        "total_rows": total_rows,
        "tables": results,
        "configured": cloudflare_sync_configured(),
    }


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def domain_from_url(url: str | None) -> str | None:
    if not url:
        return None
    parsed = urlparse(url if "://" in url else "https://" + url)
    host = parsed.netloc or parsed.path.split("/")[0]
    return host.lower().removeprefix("www.") if host else None


def comparable_url(url: str | None) -> str:
    if not url:
        return ""
    parsed = urlparse(url if "://" in url else "https://" + url)
    host = (parsed.netloc or parsed.path.split("/")[0]).lower().removeprefix("www.")
    path = parsed.path or ""
    if not parsed.netloc and "/" in parsed.path:
        path = "/" + parsed.path.split("/", 1)[1]
    path = re.sub(r"/+$", "", path)
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{host}{path}{query}".lower()


def target_url_matches(run: dict[str, Any], results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    target_url = clean_text(run.get("target_url"))
    target_domain = (clean_text(run.get("target_domain")) or domain_from_url(target_url) or "").lower().removeprefix("www.")
    target_comparable = comparable_url(target_url)
    matches: list[dict[str, Any]] = []
    if not target_domain and not target_comparable:
        return matches
    for result in results:
        result_url = clean_text(result.get("url"))
        result_host = (clean_text(result.get("host")) or domain_from_url(result_url) or "").lower().removeprefix("www.")
        result_comparable = comparable_url(result_url)
        match_type = ""
        if target_comparable and result_comparable and result_comparable == target_comparable:
            match_type = "exact_url"
        elif target_domain and result_host == target_domain:
            match_type = "same_domain"
        elif target_domain and (result_host.endswith(f".{target_domain}") or target_domain.endswith(f".{result_host}")):
            match_type = "subdomain"
        elif target_domain and result_url and target_domain in result_url.lower():
            match_type = "url_contains_domain"
        if match_type:
            item = dict(result)
            item["match_type"] = match_type
            item["target_url"] = target_url
            item["target_domain"] = target_domain
            matches.append(item)
    return matches


def normalize_url(value: str | None) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text if "://" in text else f"https://{text}"


def latest_xlsx(directory: Path = DEFAULT_REPORT_DIR) -> Path | None:
    files = sorted(directory.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def normalize_slug(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def log_activity(kind: str, message: str, level: str = "info", **fields: Any) -> None:
    try:
        ensure_dirs()
        entry = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "kind": clean_text(kind) or "system",
            "level": clean_text(level) or "info",
            "message": (clean_text(message) or "")[:600],
            **{key: value for key, value in fields.items() if value is not None},
        }
        with ACTIVITY_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(json.dumps(entry, default=str) + "\n")
        if ACTIVITY_LOG_PATH.stat().st_size > 1024 * 1024:
            lines = ACTIVITY_LOG_PATH.read_text(encoding="utf-8", errors="ignore").splitlines()[-500:]
            ACTIVITY_LOG_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except Exception:
        pass


def activity_log_tail(limit: int = 120, kind: str | None = None) -> dict[str, Any]:
    limit = max(10, min(limit, 300))
    kind = clean_text(kind)
    if not ACTIVITY_LOG_PATH.exists():
        return {"entries": []}
    entries: list[dict[str, Any]] = []
    for line in ACTIVITY_LOG_PATH.read_text(encoding="utf-8", errors="ignore").splitlines()[-1000:]:
        try:
            entry = json.loads(line)
        except json.JSONDecodeError:
            continue
        if kind and entry.get("kind") != kind:
            continue
        entries.append(entry)
    return {"entries": entries[-limit:]}


def find_report_for_job(keyword: str, started_at: datetime, directory: Path = DEFAULT_REPORT_DIR) -> Path | None:
    keyword_slug = normalize_slug(keyword)
    candidates: list[Path] = []
    for path in directory.glob("*.xlsx"):
        try:
            if datetime.fromtimestamp(path.stat().st_mtime) < started_at:
                continue
        except OSError:
            continue
        if keyword_slug and keyword_slug in normalize_slug(path.stem):
            candidates.append(path)
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def wait_for_stable_file(path: Path, checks: int = 3, delay: float = 2.0) -> bool:
    last_size = -1
    stable = 0
    for _ in range(checks + 8):
        try:
            size = path.stat().st_size
        except OSError:
            stable = 0
            time.sleep(delay)
            continue
        if size > 0 and size == last_size:
            stable += 1
            if stable >= checks:
                return True
        else:
            stable = 0
            last_size = size
        time.sleep(delay)
    return False


def find_label_value(ws, label: str) -> Any:
    label_lower = label.lower()
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip().lower() == label_lower:
                for next_value in row[idx + 1 :]:
                    if next_value not in (None, ""):
                        return next_value
    return None


def workbook_metadata(wb) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        meta["keyword"] = clean_text(find_label_value(ws, "Search Terms"))
        meta["cora_version"] = clean_text(find_label_value(ws, "Version"))
        report_date = find_label_value(ws, "Date")
        meta["report_date"] = str(report_date) if report_date else None
        user_agent = clean_text(find_label_value(ws, "User Agent"))
        meta["search_engine"] = "Google" if user_agent and "google" in user_agent.lower() else user_agent
    if "Roadmap" in wb.sheetnames:
        ws = wb["Roadmap"]
        meta["country"] = clean_text(find_label_value(ws, "Google Country"))
        meta["language"] = clean_text(find_label_value(ws, "Google Language"))
    return meta


def archive_report(source: Path, sha: str) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d")
    dest_dir = ARCHIVE_DIR / stamp / sha[:12]
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / source.name
    if not dest.exists():
        shutil.copy2(source, dest)
    return dest


def parse_results(wb, run_id: int, con: sqlite3.Connection) -> None:
    if "Results" not in wb.sheetnames:
        return
    ws = wb["Results"]
    headers = [clean_text(c.value) or "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h.lower(): i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        con.execute(
            """
            INSERT INTO serp_results
            (run_id, rank, avg_rank, weighted_rank, data_rank, host, title, url, summary)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                as_float(row[idx.get("rank", -1)]) if "rank" in idx else None,
                as_float(row[idx.get("avg rank", -1)]) if "avg rank" in idx else None,
                as_float(row[idx.get("weighted rank", -1)]) if "weighted rank" in idx else None,
                as_float(row[idx.get("data", -1)]) if "data" in idx else None,
                clean_text(row[idx.get("host", -1)]) if "host" in idx else None,
                clean_text(row[idx.get("link text", -1)]) if "link text" in idx else None,
                clean_text(row[idx.get("url", -1)]) if "url" in idx else None,
                clean_text(row[idx.get("summary", -1)]) if "summary" in idx else None,
            ),
        )


def parse_tunings(wb, run_id: int, con: sqlite3.Connection) -> None:
    sheets = ["Basic Tunings", "Intermediate Tunings", "Off Page Tunings"]
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        current_section = None
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            first = clean_text(values[0]) if len(values) > 0 else None
            factor_id = clean_text(values[1]) if len(values) > 1 else None
            factor = clean_text(values[2]) if len(values) > 2 else None
            if first and not factor_id and not factor:
                current_section = first
                continue
            if not factor_id or not factor or factor_id.lower() == "factor id":
                continue
            recommendation = clean_text(values[6]) if len(values) > 6 else None
            con.execute(
                """
                INSERT INTO recommendations
                (run_id, sheet, section, factor_id, factor, current_value, goal, percent,
                 recommendation, shared_correlation, page_one_max, page_one_avg)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    sheet,
                    current_section,
                    factor_id,
                    factor,
                    clean_text(values[3]) if len(values) > 3 else None,
                    clean_text(values[4]) if len(values) > 4 else None,
                    as_float(values[5]) if len(values) > 5 else None,
                    recommendation,
                    clean_text(values[8]) if len(values) > 8 else None,
                    clean_text(values[9]) if len(values) > 9 else None,
                    clean_text(values[10]) if len(values) > 10 else None,
                ),
            )


def parse_lsi(wb, run_id: int, con: sqlite3.Connection) -> None:
    if "LSI Keywords" not in wb.sheetnames:
        return
    ws = wb["LSI Keywords"]
    header_row = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_headers = [clean_text(v) or "" for v in row]
        if row_headers and row_headers[0] == "LSI Keyword":
            header_row = i
            headers = row_headers
            break
    if not header_row:
        return
    tracked_col = 8
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        kw = clean_text(row[0] if len(row) > 0 else None)
        if not kw:
            continue
        con.execute(
            """
            INSERT INTO lsi_keywords
            (run_id, keyword, spearman, pearson, best_of_both, pages, max_value,
             avg_value, total, tracked_value, deficit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                kw,
                as_float(row[1] if len(row) > 1 else None),
                as_float(row[2] if len(row) > 2 else None),
                as_float(row[3] if len(row) > 3 else None),
                as_float(row[4] if len(row) > 4 else None),
                as_float(row[5] if len(row) > 5 else None),
                as_float(row[6] if len(row) > 6 else None),
                as_float(row[7] if len(row) > 7 else None),
                as_float(row[tracked_col] if len(row) > tracked_col else None),
                as_float(row[9] if len(row) > 9 else None),
            ),
        )


def parse_sheet_rows(wb, run_id: int, con: sqlite3.Connection) -> None:
    keep_sheets = ["Roadmap", "Overview", "Witcher View", "Keywords", "Entities", "Questions", "Sentences"]
    for sheet in keep_sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not any(v is not None for v in row):
                continue
            con.execute(
                "INSERT INTO sheet_rows (run_id, sheet, row_index, row_json) VALUES (?, ?, ?, ?)",
                (run_id, sheet, i, json.dumps([str(v) if v is not None else None for v in row])),
            )


def parse_workbook_rows(wb, run_id: int, con: sqlite3.Connection) -> None:
    con.execute("DELETE FROM workbook_rows WHERE run_id = ?", (run_id,))
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [str(v) if v is not None else None for v in row]
            if not any(v is not None and v != "" for v in values):
                continue
            con.execute(
                """
                INSERT INTO workbook_rows (run_id, sheet, row_index, column_count, row_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                (run_id, ws.title, i, len(values), json.dumps(values)),
            )


def infer_target_url(con: sqlite3.Connection, run_id: int, target_url: str | None) -> str | None:
    if target_url:
        return target_url
    row = con.execute(
        "SELECT url FROM serp_results WHERE run_id = ? AND url IS NOT NULL ORDER BY rank LIMIT 1",
        (run_id,),
    ).fetchone()
    return row["url"] if row else None


def ingest_report(path: Path, target_url: str | None = None, keyword: str | None = None, notes: str | None = None) -> dict[str, Any]:
    path = path.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx Cora reports can be imported")

    sha = sha256_file(path)
    archive_path = archive_report(path, sha)
    wb = load_workbook(path, read_only=True, data_only=True)
    meta = workbook_metadata(wb)
    keyword = keyword or meta.get("keyword") or path.stem.replace("_", " ")
    imported_at = datetime.now().isoformat(timespec="seconds")

    with connect() as con:
        existing = con.execute("SELECT * FROM runs WHERE sha256 = ?", (sha,)).fetchone()
        if existing:
            return {"run": row_to_dict(existing), "created": False}
        cur = con.execute(
            """
            INSERT INTO runs
            (keyword, target_url, target_domain, search_engine, country, language, cora_version,
             report_date, imported_at, source_path, archive_path, file_name, file_size, sha256, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                keyword,
                target_url,
                domain_from_url(target_url),
                meta.get("search_engine"),
                meta.get("country"),
                meta.get("language"),
                meta.get("cora_version"),
                meta.get("report_date"),
                imported_at,
                str(path),
                str(archive_path),
                path.name,
                path.stat().st_size,
                sha,
                notes,
            ),
        )
        run_id = int(cur.lastrowid)
        parse_results(wb, run_id, con)
        if not target_url:
            target_url = infer_target_url(con, run_id, target_url)
            con.execute(
                "UPDATE runs SET target_url = ?, target_domain = ? WHERE id = ?",
                (target_url, domain_from_url(target_url), run_id),
            )
        parse_tunings(wb, run_id, con)
        parse_lsi(wb, run_id, con)
        parse_sheet_rows(wb, run_id, con)
        parse_workbook_rows(wb, run_id, con)
        run = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        return {"run": row_to_dict(run), "created": True}


def backfill_workbook_rows(run_id: int | None = None) -> dict[str, Any]:
    init_db()
    with connect() as con:
        if run_id is None:
            runs = con.execute("SELECT id, archive_path, source_path FROM runs ORDER BY id").fetchall()
        else:
            runs = con.execute("SELECT id, archive_path, source_path FROM runs WHERE id = ?", (run_id,)).fetchall()
        updated: list[dict[str, Any]] = []
        for run in runs:
            path = Path(run["archive_path"])
            if not path.exists():
                path = Path(run["source_path"])
            if not path.exists():
                updated.append({"run_id": run["id"], "status": "missing file"})
                continue
            wb = load_workbook(path, read_only=True, data_only=True)
            parse_workbook_rows(wb, int(run["id"]), con)
            count = con.execute("SELECT COUNT(*) FROM workbook_rows WHERE run_id = ?", (run["id"],)).fetchone()[0]
            updated.append({"run_id": run["id"], "status": "backfilled", "rows": count})
    return {"updated": updated}


def json_response(handler: BaseHTTPRequestHandler, data: Any, status: int = 200) -> None:
    body = json.dumps(data, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def error_response(handler: BaseHTTPRequestHandler, message: str, status: int = 400) -> None:
    json_response(handler, {"error": message}, status)


def read_json_body(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length") or 0)
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw) if raw.strip() else {}


def query_cora(path: str) -> Any:
    try:
        with urllib.request.urlopen(CORA_API_BASE + path, timeout=4) as resp:
            raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
    except urllib.error.URLError as exc:
        return {"error": str(exc)}


def post_cora(path: str, payload: dict[str, Any]) -> Any:
    try:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            CORA_API_BASE + path,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
    except urllib.error.URLError as exc:
        return {"error": str(exc)}


def cora_log_tail(limit: int = 80) -> dict[str, Any]:
    limit = max(10, min(limit, 150))
    data = query_cora("/api/log?lines=20")
    if isinstance(data, dict) and data.get("error"):
        return data
    raw_lines = data.get("lines", []) if isinstance(data, dict) else []
    entries: list[str] = []
    for value in raw_lines:
        text = clean_text(value or "")
        if not text:
            continue
        for part in re.split(r"(?=\b\d+\s+-\s)", text):
            line = clean_text(part)
            if not re.match(r"^\d+\s+-\s", line):
                continue
            entries.append(line[:500])
    return {"lines": entries[-limit:]}


def cora_status_signature(status: dict[str, Any]) -> tuple[Any, ...]:
    progress = as_float(status.get("progress"))
    return (
        bool(status.get("running")),
        bool(status.get("searchRunning")),
        clean_text(status.get("searchTerm")) or "",
        clean_text(status.get("action")) or "",
        round(progress, 4) if progress is not None else None,
    )


def force_stop_cora(reason: str = "Stop requested") -> dict[str, Any]:
    before = query_cora("/api/status")
    response = post_cora("/api/stop", {"reason": reason})
    ok = not (isinstance(response, dict) and response.get("error"))
    return {"ok": ok, "reason": reason, "status_before": before, "response": response}


def ps_quote(value: Path | str) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def restart_cora(reason: str = "Dashboard restart requested") -> dict[str, Any]:
    stop_result = force_stop_cora(reason)
    time.sleep(2)
    kill_script = """
Get-CimInstance Win32_Process -Filter "Name = 'javaw.exe'" |
  Where-Object { $_.CommandLine -like '*cora-recompiled.jar*' -or $_.ExecutablePath -like '*SEO Correlation Tool 2026*' } |
  ForEach-Object {
    Stop-Process -Id $_.ProcessId -Force
    $_.ProcessId
  }
"""
    killed = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", kill_script],
        capture_output=True,
        text=True,
        timeout=30,
    )
    if not CORA_DEPLOY_BAT.exists():
        return {
            "ok": False,
            "error": f"Cora launcher not found: {CORA_DEPLOY_BAT}",
            "stop": stop_result,
            "killed": killed.stdout.strip().splitlines(),
            "kill_error": killed.stderr.strip(),
        }
    start_script = f"Start-Process -WindowStyle Hidden -FilePath {ps_quote(CORA_DEPLOY_BAT)} -WorkingDirectory {ps_quote(CORA_DEPLOY_DIR)}"
    started = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", start_script],
        capture_output=True,
        text=True,
        timeout=30,
    )
    ok = killed.returncode == 0 and started.returncode == 0
    return {
        "ok": ok,
        "reason": reason,
        "stop": stop_result,
        "killed": killed.stdout.strip().splitlines(),
        "kill_error": killed.stderr.strip(),
        "start_error": started.stderr.strip(),
    }


def cora_profiles() -> dict[str, Any]:
    result = query_cora("/api/profiles")
    if isinstance(result, dict):
        return result
    return {"error": "Unexpected Cora profile response"}


def create_cora_profile(name: str) -> dict[str, Any]:
    name = name.strip()
    if not name:
        raise ValueError("Cora profile name is required")
    result = post_cora("/api/profiles", {"name": name})
    if isinstance(result, dict) and result.get("error"):
        raise RuntimeError(f"Cora profile could not be created: {result['error']}")
    return result if isinstance(result, dict) else {"ok": True, "profile": name}


def sync_cora_profiles() -> list[dict[str, Any]]:
    result = cora_profiles()
    if result.get("error"):
        with connect() as con:
            rows = con.execute(
                """
                SELECT pr.*,
                       (SELECT COUNT(*) FROM projects p WHERE p.profile_id = pr.id) AS project_count
                FROM profiles pr
                ORDER BY pr.name COLLATE NOCASE
                """
            ).fetchall()
            return [row_to_dict(r) for r in rows]
    names = result.get("profiles") or []
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        for name in names:
            if not clean_text(name):
                continue
            existing = find_profile_by_name(con, str(name))
            if not existing:
                con.execute(
                    """
                    INSERT INTO profiles (name, client, notes, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (str(name), None, "Imported from Cora profiles", now, now),
                )
        rows = con.execute(
            """
            SELECT pr.*,
                   (SELECT COUNT(*) FROM projects p WHERE p.profile_id = pr.id) AS project_count
            FROM profiles pr
            ORDER BY pr.name COLLATE NOCASE
            """
        ).fetchall()
        return [row_to_dict(r) for r in rows]


def job_update(job_id: int, **fields: Any) -> None:
    if not fields:
        return
    cols = ", ".join(f"{key} = ?" for key in fields)
    values = list(fields.values()) + [job_id]
    with connect() as con:
        con.execute(f"UPDATE managed_jobs SET {cols} WHERE id = ?", values)
        row = con.execute("SELECT keyword, project_id, status FROM managed_jobs WHERE id = ?", (job_id,)).fetchone()
    log_job_activity(job_id, row, fields)


def log_job_activity(job_id: int, row: sqlite3.Row | None, fields: dict[str, Any]) -> None:
    status = clean_text(fields.get("status"))
    message = clean_text(fields.get("status_message")) or status
    if not message:
        return
    should_log = False
    if status in {"submitting", "running", "imported", "error", "timeout", "stopped"}:
        should_log = True
    elif status == "queued" and any(token in message.lower() for token in ["retry", "paused", "clearing", "resume"]):
        should_log = True
    elif fields.get("imported_run_id") or fields.get("stall_detected_at"):
        should_log = True
    if not should_log:
        return
    level = "info"
    kind = "job"
    if status in {"error", "timeout", "stopped"} or fields.get("stall_detected_at"):
        level = "error"
    elif status == "queued":
        level = "warn"
    if status == "imported" or fields.get("imported_run_id"):
        kind = "import"
    keyword = row["keyword"] if row else ""
    log_activity(
        kind,
        f"Job {job_id}: {keyword} - {message}",
        level,
        job_id=job_id,
        keyword=keyword,
        project_id=row["project_id"] if row and "project_id" in row.keys() else None,
        status=status or (row["status"] if row else None),
    )


def queue_pause_message() -> str:
    state = queue_state()
    if state.get("stop_after_current"):
        return "Queued; stop after current run is active"
    if state.get("auto_resume"):
        return "Queued; waiting for Cora to become idle"
    return "Queued; queue paused"


def schedule_job_thread(job_id: int) -> None:
    thread = threading.Thread(target=run_managed_job, args=(job_id,), daemon=True)
    thread.start()


def fail_or_retry_job(
    job_id: int,
    status: str,
    status_message: str,
    error: str,
    retryable: bool = True,
    **fields: Any,
) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    spawn_retry = False
    activity: tuple[str, str, str, dict[str, Any]] | None = None
    with connect() as con:
        row = con.execute("SELECT * FROM managed_jobs WHERE id = ?", (job_id,)).fetchone()
        if not row:
            return
        retry_count = int(row["retry_count"] or 0)
        max_retries = int(row["max_retries"] or 0)
        if retryable and retry_count < max_retries:
            next_retry = retry_count + 1
            con.execute(
                """
                UPDATE managed_jobs
                SET status = ?, status_message = ?, error = ?, retry_count = ?, next_retry_at = ?,
                    completed_at = NULL, cora_running = 0, cora_action = NULL, progress = NULL,
                    stall_detected_at = NULL, last_activity_at = ?
                WHERE id = ?
                """,
                (
                    "queued",
                    f"{status_message}; retry {next_retry} of {max_retries} queued",
                    error,
                    next_retry,
                    now,
                    now,
                    job_id,
                ),
            )
            spawn_retry = not queue_paused()
            activity = (
                "job",
                f"Job {job_id}: {row['keyword']} - {status_message}; retry {next_retry} of {max_retries} queued",
                "warn",
                {"job_id": job_id, "keyword": row["keyword"], "project_id": row["project_id"], "status": "queued"},
            )
        else:
            updates = {
                "status": status,
                "status_message": status_message,
                "error": error,
                "completed_at": now,
                "cora_running": 0,
                "last_activity_at": now,
                **fields,
            }
            cols = ", ".join(f"{key} = ?" for key in updates)
            con.execute(f"UPDATE managed_jobs SET {cols} WHERE id = ?", [*updates.values(), job_id])
            activity = (
                "job",
                f"Job {job_id}: {row['keyword']} - {status_message}",
                "error" if status in {"error", "timeout", "stopped"} else "info",
                {"job_id": job_id, "keyword": row["keyword"], "project_id": row["project_id"], "status": status},
            )
    if activity:
        log_activity(activity[0], activity[1], activity[2], **activity[3])
    if spawn_retry:
        schedule_job_thread(job_id)


def create_managed_job(
    keyword: str,
    target_url: str,
    cora_profile: str | None = None,
    project_id: int | None = None,
    keyword_id: int | None = None,
    tool: str = "cora",
) -> dict[str, Any]:
    keyword = keyword.strip()
    target_url = target_url.strip()
    cora_profile = clean_text(cora_profile)
    if not keyword:
        raise ValueError("Keyword is required")
    if not target_url:
        raise ValueError("Target URL/domain is required")
    target_domain = domain_from_url(target_url)
    if not target_domain:
        raise ValueError("Could not determine target domain")

    script = f"search {keyword}; force {target_url}; track domain {target_domain}; click get data"
    if cora_profile:
        script = f"profile {cora_profile}; " + script
    started_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO managed_jobs
            (keyword, target_url, target_domain, cora_profile, cora_script, status, status_message,
             started_at, last_activity_at, retry_count, max_retries, project_id, keyword_id, tool)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                keyword,
                target_url,
                target_domain,
                cora_profile,
                script,
                "queued",
                "Queued",
                started_at,
                started_at,
                0,
                DEFAULT_JOB_MAX_RETRIES,
                project_id,
                keyword_id,
                tool,
            ),
        )
        job_id = int(cur.lastrowid)
        job = row_to_dict(con.execute("SELECT * FROM managed_jobs WHERE id = ?", (job_id,)).fetchone())

    log_activity(
        "queue",
        f"Queued Cora job {job_id}: {keyword}",
        "info",
        job_id=job_id,
        keyword=keyword,
        project_id=project_id,
        target_url=target_url,
    )
    if queue_paused():
        job_update(job_id, status_message=queue_pause_message())
    else:
        schedule_job_thread(job_id)
    return job or {}


def wait_for_cora_idle(job_id: int, deadline: float) -> bool:
    stale_clear_sent = False
    while time.time() < deadline:
        status = query_cora("/api/status")
        if isinstance(status, dict) and status.get("error"):
            raise RuntimeError(f"Cora API is not reachable: {status['error']}")
        if isinstance(status, dict):
            action = clean_text(status.get("action"))
            progress = as_float(status.get("progress"))
            running = bool(status.get("running"))
            search_running = bool(status.get("searchRunning"))
            stale_blank_state = not action and progress in (None, 0.0, 1.0)
            if not running and not search_running:
                return True
            if stale_blank_state and not stale_clear_sent:
                job_update(job_id, status="queued", status_message="Clearing stale Cora busy state")
                force_stop_cora("Clearing stale Cora busy state before next dashboard job")
                stale_clear_sent = True
                time.sleep(JOB_POLL_SECONDS)
                continue
        job_update(job_id, status="queued", status_message="Waiting for Cora to become idle")
        time.sleep(JOB_POLL_SECONDS)
    return False


def cora_is_cleanly_idle() -> bool:
    status = query_cora("/api/status")
    if not isinstance(status, dict) or status.get("error"):
        return False
    return not bool(status.get("running")) and not bool(status.get("searchRunning"))


def queued_job_count() -> int:
    with connect() as con:
        return int(con.execute("SELECT COUNT(*) FROM managed_jobs WHERE status = 'queued'").fetchone()[0])


def active_job_count() -> int:
    with connect() as con:
        return int(
            con.execute("SELECT COUNT(*) FROM managed_jobs WHERE status IN ('submitting', 'running')").fetchone()[0]
        )


def parse_iso(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def decorate_job(row: sqlite3.Row | dict[str, Any], now: datetime | None = None) -> dict[str, Any]:
    data = row_to_dict(row) if isinstance(row, sqlite3.Row) else dict(row)
    now = now or datetime.now()
    status = str(data.get("status") or "")
    last_activity = parse_iso(data.get("last_activity_at")) or parse_iso(data.get("started_at"))
    seconds_since_activity = int((now - last_activity).total_seconds()) if last_activity else None
    retry_count = int(data.get("retry_count") or 0)
    max_retries = int(data.get("max_retries") or 0)
    stalled = status in {"running", "submitting"} and seconds_since_activity is not None and seconds_since_activity >= CORA_FREEZE_SECONDS
    data["seconds_since_activity"] = seconds_since_activity
    data["stall_after_seconds"] = CORA_FREEZE_SECONDS
    data["stalled"] = stalled
    data["retry_remaining"] = max(0, max_retries - retry_count)
    return data


def queue_summary(jobs: list[dict[str, Any]], state: dict[str, Any]) -> dict[str, Any]:
    total = len(jobs)
    complete = sum(1 for job in jobs if job.get("status") == "imported")
    queued = [job for job in jobs if job.get("status") == "queued"]
    active = [job for job in jobs if job.get("status") in {"running", "submitting"}]
    failed = [job for job in jobs if job.get("status") in {"error", "timeout", "stopped"}]
    stalled = [job for job in active if job.get("stalled")]
    if state.get("stop_after_current"):
        label = "stopping_after_current"
    elif stalled:
        label = "stalled"
    elif state.get("paused"):
        label = "paused"
    elif active:
        label = "running"
    elif queued:
        label = "queued"
    elif total and complete + len(failed) >= total:
        label = "completed"
    else:
        label = "idle"
    return {
        "total": total,
        "complete": complete,
        "queued": len(queued),
        "active": len(active),
        "failed": len(failed),
        "stalled": len(stalled),
        "label": label,
        "running_job": active[0] if active else None,
        "next_job": queued[0] if queued else None,
    }


def run_managed_job(job_id: int) -> None:
    deadline = time.time() + JOB_TIMEOUT_SECONDS
    with CORA_JOB_LOCK:
        with connect() as con:
            job = con.execute("SELECT * FROM managed_jobs WHERE id = ?", (job_id,)).fetchone()
        if not job:
            return
        if job["status"] in {"imported", "error", "timeout", "stopped"}:
            return
        if queue_paused():
            job_update(job_id, status="queued", status_message=queue_pause_message(), cora_running=0)
            return

        keyword = job["keyword"]
        target_url = job["target_url"]
        script = job["cora_script"]
        attempt_started_at = datetime.now()
        now = attempt_started_at.isoformat(timespec="seconds")
        job_update(job_id, last_activity_at=now, stall_detected_at=None)
        try:
            if not wait_for_cora_idle(job_id, deadline):
                raise TimeoutError("Timed out waiting for Cora to become idle")
        except Exception as exc:
            fail_or_retry_job(
                job_id,
                "error",
                "Cora was not ready",
                str(exc),
            )
            return

        job_update(
            job_id,
            status="submitting",
            status_message="Submitting Cora workflow",
            last_activity_at=datetime.now().isoformat(timespec="seconds"),
        )

        response = post_cora("/api/script", {"script": script})
        if isinstance(response, dict) and response.get("error"):
            fail_or_retry_job(
                job_id,
                "error",
                "Cora workflow submission failed",
                str(response["error"]),
            )
            return

        job_update(
            job_id,
            status="running",
            status_message="Cora workflow submitted",
            last_activity_at=datetime.now().isoformat(timespec="seconds"),
        )
        seen_report: Path | None = None
        last_signature: tuple[Any, ...] | None = None
        last_activity = time.time()
        last_candidate_state: tuple[str, int, float] | None = None

        while time.time() < deadline:
            status = query_cora("/api/status")
            if isinstance(status, dict) and not status.get("error"):
                signature = cora_status_signature(status)
                if signature != last_signature:
                    last_signature = signature
                    last_activity = time.time()
                    job_update(job_id, last_activity_at=datetime.now().isoformat(timespec="seconds"))
                job_update(
                    job_id,
                    cora_running=1 if status.get("running") else 0,
                    cora_action=status.get("action"),
                    progress=as_float(status.get("progress")),
                    status_message=status.get("action") or "Waiting for Cora report",
                )
                if (status.get("running") or status.get("searchRunning")) and time.time() - last_activity >= CORA_FREEZE_SECONDS:
                    minutes = max(1, round(CORA_FREEZE_SECONDS / 60))
                    stop = force_stop_cora(f"Frozen dashboard managed job {job_id}")
                    fail_or_retry_job(
                        job_id,
                        "stopped",
                        f"Cora appeared frozen for {minutes} minutes; stop sent",
                        json.dumps(stop, default=str),
                        stall_detected_at=datetime.now().isoformat(timespec="seconds"),
                    )
                    return

            candidate = find_report_for_job(keyword, attempt_started_at)
            if candidate:
                seen_report = candidate
                candidate_state = (str(candidate), candidate.stat().st_size, candidate.stat().st_mtime)
                if candidate_state != last_candidate_state:
                    last_candidate_state = candidate_state
                    last_activity = time.time()
                    job_update(job_id, last_activity_at=datetime.now().isoformat(timespec="seconds"))
                if wait_for_stable_file(candidate):
                    try:
                        result = ingest_report(candidate, target_url=target_url, keyword=keyword, notes=f"Managed job {job_id}")
                        run = result.get("run") or {}
                        if run.get("id") and job["project_id"]:
                            with connect() as con:
                                keyword_row = con.execute(
                                    "SELECT site_id, page_id FROM keywords WHERE id = ?",
                                    (job["keyword_id"],),
                                ).fetchone() if job["keyword_id"] else None
                            assign_run(
                                int(run["id"]),
                                int(job["project_id"]),
                                int(keyword_row["site_id"]) if keyword_row and keyword_row["site_id"] else None,
                                int(keyword_row["page_id"]) if keyword_row and keyword_row["page_id"] else None,
                                int(job["keyword_id"]) if job["keyword_id"] else None,
                            )
                        job_update(
                            job_id,
                            status="imported",
                            status_message="Report imported",
                            completed_at=datetime.now().isoformat(timespec="seconds"),
                            report_path=str(candidate),
                            imported_run_id=run.get("id"),
                            progress=1.0,
                            last_activity_at=datetime.now().isoformat(timespec="seconds"),
                        )
                        return
                    except Exception as exc:
                        fail_or_retry_job(
                            job_id,
                            "error",
                            "Report import failed",
                            str(exc),
                            retryable=False,
                            report_path=str(candidate),
                        )
                        return

            time.sleep(JOB_POLL_SECONDS)

    fail_or_retry_job(
        job_id,
        "timeout",
        "Timed out waiting for Cora report",
        "Timed out waiting for a matching .xlsx report",
        report_path=str(seen_report) if seen_report else None,
    )


def resume_pending_jobs() -> None:
    if queue_paused():
        return
    with connect() as con:
        rows = con.execute(
            """
            SELECT id FROM managed_jobs
            WHERE status IN ('queued', 'submitting', 'running')
            ORDER BY id
            """
        ).fetchall()
        for row in rows:
            con.execute(
                """
                UPDATE managed_jobs
                SET status = ?, status_message = ?, cora_running = 0, stall_detected_at = NULL,
                    last_activity_at = ?
                WHERE id = ?
                """,
                ("queued", "Queued for resume", datetime.now().isoformat(timespec="seconds"), row["id"]),
            )
    for row in rows:
        schedule_job_thread(int(row["id"]))


def auto_resume_loop() -> None:
    while True:
        try:
            state = queue_state()
            if (
                state.get("paused")
                and state.get("auto_resume")
                and queued_job_count() > 0
                and active_job_count() == 0
                and cora_is_cleanly_idle()
            ):
                set_queue_paused(False)
                resume_pending_jobs()
        except Exception:
            pass
        time.sleep(JOB_POLL_SECONDS)


def find_profile_by_name(con: sqlite3.Connection, name: str) -> sqlite3.Row | None:
    return con.execute(
        "SELECT * FROM profiles WHERE lower(name) = lower(?)",
        (name.strip(),),
    ).fetchone()


def create_profile(name: str, client: str | None = None, notes: str | None = None) -> dict[str, Any]:
    name = name.strip()
    if not name:
        raise ValueError("Profile name is required")
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        existing = find_profile_by_name(con, name)
        if existing:
            return row_to_dict(existing) or {}
        cur = con.execute(
            """
            INSERT INTO profiles (name, client, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (name, clean_text(client), clean_text(notes), now, now),
        )
        return row_to_dict(con.execute("SELECT * FROM profiles WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def update_profile(profile_id: int, name: str, client: str | None = None, notes: str | None = None) -> dict[str, Any]:
    name = name.strip()
    if not name:
        raise ValueError("Profile name is required")
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        existing = con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
        if not existing:
            raise ValueError("Profile not found")
        duplicate = con.execute(
            "SELECT id FROM profiles WHERE lower(name) = lower(?) AND id != ?",
            (name, profile_id),
        ).fetchone()
        if duplicate:
            raise ValueError("Another profile already uses that name")
        con.execute(
            """
            UPDATE profiles
            SET name = ?, client = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (name, clean_text(client), clean_text(notes), now, profile_id),
        )
        row = con.execute(
            """
            SELECT pr.*,
                   (SELECT COUNT(*) FROM projects p WHERE p.profile_id = pr.id) AS project_count
            FROM profiles pr
            WHERE pr.id = ?
            """,
            (profile_id,),
        ).fetchone()
    return row_to_dict(row) or {}


def get_or_create_profile(
    con: sqlite3.Connection,
    name: str,
    client: str | None = None,
    notes: str | None = None,
) -> sqlite3.Row:
    existing = find_profile_by_name(con, name)
    if existing:
        return existing
    now = datetime.now().isoformat(timespec="seconds")
    cur = con.execute(
        """
        INSERT INTO profiles (name, client, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (name.strip(), clean_text(client), clean_text(notes), now, now),
    )
    return con.execute("SELECT * FROM profiles WHERE id = ?", (cur.lastrowid,)).fetchone()


def create_project(
    name: str,
    client: str | None = None,
    site_domain: str | None = None,
    notes: str | None = None,
    profile_id: int | None = None,
    profile_name: str | None = None,
) -> dict[str, Any]:
    name = name.strip()
    if not name:
        raise ValueError("Project name is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        if profile_id:
            profile = con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
            if not profile:
                raise ValueError("Selected profile was not found")
        elif clean_text(profile_name):
            profile = get_or_create_profile(con, clean_text(profile_name) or "", client=client)
            profile_id = int(profile["id"])
        cur = con.execute(
            "INSERT INTO projects (profile_id, name, client, notes, created_at) VALUES (?, ?, ?, ?, ?)",
            (profile_id, name, clean_text(client), clean_text(notes), created_at),
        )
        project_id = int(cur.lastrowid)
        if clean_text(site_domain):
            domain = domain_from_url(site_domain) or site_domain.strip().lower()
            con.execute(
                "INSERT INTO sites (project_id, domain, name, created_at) VALUES (?, ?, ?, ?)",
                (project_id, domain, domain, created_at),
            )
        return row_to_dict(
            con.execute(
                """
                SELECT p.*, pr.name AS profile_name, pr.client AS profile_client
                FROM projects p
                LEFT JOIN profiles pr ON pr.id = p.profile_id
                WHERE p.id = ?
                """,
                (project_id,),
            ).fetchone()
        ) or {}


def attach_project_profile(
    project_id: int,
    profile_id: int | None = None,
    profile_name: str | None = None,
    detach: bool = False,
) -> dict[str, Any]:
    profile_name = clean_text(profile_name)
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Client not found")
        resolved_profile_id: int | None = None
        if not detach:
            if profile_id:
                profile = con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
                if not profile:
                    raise ValueError("Selected Cora profile was not found")
                resolved_profile_id = int(profile["id"])
            elif profile_name:
                profile = get_or_create_profile(con, profile_name, client=project["client"] or project["name"])
                resolved_profile_id = int(profile["id"])
            else:
                raise ValueError("Select an existing Cora profile or enter a new profile name")
        con.execute("UPDATE projects SET profile_id = ? WHERE id = ?", (resolved_profile_id, project_id))
        if resolved_profile_id:
            con.execute("UPDATE profiles SET updated_at = ? WHERE id = ?", (now, resolved_profile_id))
        row = con.execute(
            """
            SELECT p.*, pr.name AS profile_name, pr.client AS profile_client
            FROM projects p
            LEFT JOIN profiles pr ON pr.id = p.profile_id
            WHERE p.id = ?
            """,
            (project_id,),
        ).fetchone()
        return row_to_dict(row) or {}


def create_site(project_id: int, domain: str, name: str | None = None) -> dict[str, Any]:
    domain = domain_from_url(domain) or domain.strip().lower()
    if not domain:
        raise ValueError("Domain is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            "INSERT INTO sites (project_id, domain, name, created_at) VALUES (?, ?, ?, ?)",
            (project_id, domain, clean_text(name) or domain, created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM sites WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def create_page(site_id: int, url: str, title: str | None = None) -> dict[str, Any]:
    url = url.strip()
    if not url:
        raise ValueError("Page URL is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            "INSERT INTO pages (site_id, url, title, created_at) VALUES (?, ?, ?, ?)",
            (site_id, url, clean_text(title), created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM pages WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def create_keyword(
    project_id: int,
    keyword: str,
    site_id: int | None = None,
    page_id: int | None = None,
    intent: str | None = None,
    priority: str | None = None,
) -> dict[str, Any]:
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("Keyword is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO keywords (project_id, site_id, page_id, keyword, intent, priority, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (project_id, site_id, page_id, keyword, clean_text(intent), clean_text(priority), created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM keywords WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def find_keyword_by_text(project_id: int, keyword: str) -> dict[str, Any] | None:
    with connect() as con:
        row = con.execute(
            """
            SELECT * FROM keywords
            WHERE project_id = ? AND lower(keyword) = lower(?)
            ORDER BY id
            LIMIT 1
            """,
            (project_id, keyword.strip()),
        ).fetchone()
    return row_to_dict(row)


def queue_ranking_snapshot_cora_job(
    project_id: int,
    keyword: str,
    ranking_url: str,
    cora_profile: str | None = None,
    create_keyword_if_missing: bool = True,
) -> dict[str, Any]:
    keyword = keyword.strip()
    ranking_url = normalize_url(ranking_url) or ranking_url.strip()
    if not keyword:
        raise ValueError("Keyword is required")
    if not ranking_url:
        raise ValueError("Ranking URL is required")
    with connect() as con:
        project = con.execute(
            """
            SELECT p.*, pr.name AS profile_name
            FROM projects p
            LEFT JOIN profiles pr ON pr.id = p.profile_id
            WHERE p.id = ?
            """,
            (project_id,),
        ).fetchone()
    if not project:
        raise ValueError("Client not found")
    keyword_row = find_keyword_by_text(project_id, keyword)
    created_keyword = False
    if not keyword_row and create_keyword_if_missing:
        keyword_row = create_keyword(project_id, keyword, intent="Discovered", priority="Medium")
        created_keyword = True
    profile = clean_text(cora_profile) or clean_text(project["profile_name"])
    job = create_managed_job(
        keyword,
        ranking_url,
        profile,
        project_id=project_id,
        keyword_id=int(keyword_row["id"]) if keyword_row else None,
        tool="cora",
    )
    return {"job": job, "keyword": keyword_row, "created_keyword": created_keyword}


def keyword_target_url(keyword: sqlite3.Row, project: sqlite3.Row) -> str:
    if keyword["page_url"]:
        return normalize_url(str(keyword["page_url"])) or str(keyword["page_url"])
    if keyword["site_domain"]:
        return normalize_url(str(keyword["site_domain"])) or str(keyword["site_domain"])
    if project["main_url"]:
        return normalize_url(str(project["main_url"])) or str(project["main_url"])
    if project["main_domain"]:
        return normalize_url(str(project["main_domain"])) or str(project["main_domain"])
    raise ValueError("Client needs a Main URL or keyword page before running Cora")


def run_client_tool(project_id: int, keyword_ids: list[int], tool: str, cora_profile: str | None = None) -> dict[str, Any]:
    tool = (tool or "").strip().lower()
    if not keyword_ids:
        raise ValueError("Select at least one keyword")
    if tool != "cora":
        labels = {
            "entity-lsi": "Entity & LSI Explorer",
            "tools-2": "Tools 2",
            "aeo": "AEO Tool",
        }
        return {
            "tool": tool,
            "placeholder": True,
            "message": f"{labels.get(tool, 'Selected tool')} is a placeholder for now.",
            "keyword_ids": keyword_ids,
        }
    with connect() as con:
        project = con.execute(
            """
            SELECT p.*, pr.name AS profile_name,
                   (SELECT pg.url FROM pages pg JOIN sites s ON s.id = pg.site_id WHERE s.project_id = p.id ORDER BY pg.id LIMIT 1) AS main_url,
                   (SELECT s.domain FROM sites s WHERE s.project_id = p.id ORDER BY s.id LIMIT 1) AS main_domain
            FROM projects p
            LEFT JOIN profiles pr ON pr.id = p.profile_id
            WHERE p.id = ?
            """,
            (project_id,),
        ).fetchone()
        if not project:
            raise ValueError("Client not found")
        rows = con.execute(
            f"""
            SELECT k.*, s.domain AS site_domain, pg.url AS page_url
            FROM keywords k
            LEFT JOIN sites s ON s.id = k.site_id
            LEFT JOIN pages pg ON pg.id = k.page_id
            WHERE k.project_id = ? AND k.id IN ({','.join('?' for _ in keyword_ids)})
            ORDER BY k.id
            """,
            [project_id, *keyword_ids],
        ).fetchall()
    found_ids = {int(row["id"]) for row in rows}
    missing = [kid for kid in keyword_ids if kid not in found_ids]
    if missing:
        raise ValueError(f"Keyword not found for this client: {missing[0]}")
    profile_name = clean_text(cora_profile) or clean_text(project["profile_name"])
    jobs = [
        create_managed_job(
            str(row["keyword"]),
            keyword_target_url(row, project),
            profile_name,
            project_id=project_id,
            keyword_id=int(row["id"]),
            tool="cora",
        )
        for row in rows
    ]
    return {"tool": "cora", "jobs": jobs}


def assign_run(run_id: int, project_id: int | None, site_id: int | None, page_id: int | None, keyword_id: int | None) -> dict[str, Any]:
    with connect() as con:
        con.execute(
            "UPDATE runs SET project_id = ?, site_id = ?, page_id = ?, keyword_id = ? WHERE id = ?",
            (project_id, site_id, page_id, keyword_id, run_id),
        )
        row = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
    if not row:
        raise ValueError("Run not found")
    return row_to_dict(row) or {}


def mask_secret(value: str) -> str:
    if len(value) <= 8:
        return "*" * len(value)
    return f"{value[:4]}...{value[-4:]}"


def api_key_public(row: sqlite3.Row) -> dict[str, Any]:
    data = row_to_dict(row) or {}
    key_value = data.pop("key_value", "") or ""
    provider_key = normalize_ai_provider(data.get("provider", ""))
    if provider_key:
        data["provider_key"] = provider_key
        data["provider_name"] = AI_PROVIDERS[provider_key]["name"]
    data["key_preview"] = "login:********" if provider_key == "dataforseo" else mask_secret(key_value)
    data["key_length"] = len(key_value)
    return data


def normalize_ai_provider(provider: str) -> str:
    normalized = (provider or "").strip().lower()
    aliases = {
        "open ai": "openai",
        "chatgpt": "openai",
        "claude": "anthropic",
        "gemini": "google",
        "google ai": "google",
        "google gemini": "google",
        "grok": "xai",
        "x.ai": "xai",
        "xai": "xai",
        "xai / grok": "xai",
        "xai/grok": "xai",
        "pplx": "perplexity",
        "perplexity ai": "perplexity",
        "data for seo": "dataforseo",
        "dataforseo": "dataforseo",
        "dfs": "dataforseo",
    }
    return aliases.get(normalized, normalized if normalized in AI_PROVIDERS else "")


def ai_provider_catalog() -> list[dict[str, Any]]:
    return [
        {"key": key, **meta}
        for key, meta in AI_PROVIDERS.items()
    ]


def sanitize_provider_message(message: str, key_value: str = "") -> str:
    cleaned = str(message or "").replace("\r", " ").replace("\n", " ")
    if key_value:
        cleaned = cleaned.replace(key_value, "[api key]")
        if ":" in key_value:
            login, password = key_value.split(":", 1)
            if login:
                cleaned = cleaned.replace(login, "[api login]")
            if password:
                cleaned = cleaned.replace(password, "[api password]")
    cleaned = re.sub(r"(sk-[A-Za-z0-9_\-]{8,})", "[api key]", cleaned)
    cleaned = re.sub(r"(AIza[A-Za-z0-9_\-]{8,})", "[api key]", cleaned)
    cleaned = re.sub(r"(xai-[A-Za-z0-9_\-]{8,})", "[api key]", cleaned)
    cleaned = re.sub(r"(pplx-[A-Za-z0-9_\-]{8,})", "[api key]", cleaned)
    cleaned = re.sub(r"(Basic\s+[A-Za-z0-9+/=]{12,})", "Basic [api key]", cleaned)
    return cleaned[:500]


def api_key_value_from_payload(provider: str, payload: dict[str, Any]) -> str:
    key_value = str(payload.get("key_value") or "")
    if normalize_ai_provider(provider) != "dataforseo" or key_value.strip():
        return key_value
    login = str(payload.get("api_login") or "").strip()
    password = str(payload.get("api_password") or "").strip()
    if login and password:
        return f"{login}:{password}"
    return ""


def create_api_key(
    provider: str,
    label: str,
    key_value: str,
    notes: str | None = None,
    base_url: str | None = None,
    default_model: str | None = None,
    status: str | None = None,
    last_tested_at: str | None = None,
    last_error: str | None = None,
) -> dict[str, Any]:
    provider_key = normalize_ai_provider(provider)
    provider = AI_PROVIDERS[provider_key]["name"] if provider_key else provider.strip()
    label = label.strip()
    key_value = key_value.strip()
    if not provider:
        raise ValueError("Provider is required")
    if not label:
        raise ValueError("Label is required")
    if not key_value:
        raise ValueError("API key is required")
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO api_keys
            (provider, label, key_value, notes, base_url, default_model, status, last_tested_at, last_error, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                provider,
                label,
                key_value,
                clean_text(notes),
                clean_text(base_url),
                clean_text(default_model),
                clean_text(status) or "untested",
                clean_text(last_tested_at),
                sanitize_provider_message(last_error or "", key_value) if last_error else None,
                now,
                now,
            ),
        )
        row = con.execute("SELECT * FROM api_keys WHERE id = ?", (cur.lastrowid,)).fetchone()
    return api_key_public(row)


def provider_request(provider_key: str, key_value: str, base_url: str | None = None) -> urllib.request.Request:
    meta = AI_PROVIDERS[provider_key]
    root = (clean_text(base_url) or meta["base_url"]).rstrip("/")
    if provider_key == "google":
        url = f"{root}/v1beta/models?key={quote(key_value)}"
        return urllib.request.Request(url, headers={"Accept": "application/json"})
    if provider_key == "perplexity":
        return urllib.request.Request(
            f"{root}/v1/models",
            headers={"Accept": "application/json", "Authorization": f"Bearer {key_value}"},
        )
    if provider_key == "dataforseo":
        if ":" not in key_value:
            raise ValueError("DataForSEO requires API login and API password")
        token = base64.b64encode(key_value.encode("utf-8")).decode("ascii")
        return urllib.request.Request(
            f"{root}{meta.get('test_path', '/v3/appendix/user_data')}",
            headers={
                "Accept": "application/json",
                "Content-Type": "application/json",
                "Authorization": f"Basic {token}",
            },
        )
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {key_value}",
    }
    if provider_key == "anthropic":
        headers = {
            "Accept": "application/json",
            "x-api-key": key_value,
            "anthropic-version": "2023-06-01",
        }
    return urllib.request.Request(f"{root}/v1/models", headers=headers)


def test_ai_provider_key(provider: str, key_value: str, base_url: str | None = None, timeout: int = 12) -> dict[str, Any]:
    provider_key = normalize_ai_provider(provider)
    key_value = (key_value or "").strip()
    if not provider_key:
        raise ValueError("Choose OpenAI, Anthropic, Google, xAI / Grok, Perplexity, or DataForSEO")
    if not key_value:
        raise ValueError("API credentials are required")
    request = provider_request(provider_key, key_value, base_url)
    tested_at = datetime.now().isoformat(timespec="seconds")
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            body = response.read(2048).decode("utf-8", errors="replace")
            message = "Connection verified."
            if body:
                message = f"Connection verified. HTTP {response.status}."
            return {
                "ok": 200 <= response.status < 300,
                "provider_key": provider_key,
                "provider_name": AI_PROVIDERS[provider_key]["name"],
                "status_code": response.status,
                "status": "valid",
                "message": message,
                "tested_at": tested_at,
            }
    except urllib.error.HTTPError as exc:
        body = exc.read(2048).decode("utf-8", errors="replace")
        message = sanitize_provider_message(body or exc.reason or str(exc), key_value)
        return {
            "ok": False,
            "provider_key": provider_key,
            "provider_name": AI_PROVIDERS[provider_key]["name"],
            "status_code": exc.code,
            "status": "failed",
            "message": message or f"HTTP {exc.code}",
            "tested_at": tested_at,
        }
    except Exception as exc:
        return {
            "ok": False,
            "provider_key": provider_key,
            "provider_name": AI_PROVIDERS[provider_key]["name"],
            "status_code": None,
            "status": "failed",
            "message": sanitize_provider_message(str(exc), key_value),
            "tested_at": tested_at,
        }


def test_api_key_payload(payload: dict[str, Any]) -> dict[str, Any]:
    key_id = int(payload["key_id"]) if payload.get("key_id") else None
    if key_id:
        with connect() as con:
            row = con.execute("SELECT * FROM api_keys WHERE id = ?", (key_id,)).fetchone()
        if not row:
            raise ValueError("API key not found")
        provider = row["provider"]
        key_value = row["key_value"]
        base_url = row["base_url"] if "base_url" in row.keys() else None
    else:
        provider = payload.get("provider", "")
        key_value = api_key_value_from_payload(provider, payload)
        base_url = payload.get("base_url")
    result = test_ai_provider_key(provider, key_value, base_url)
    if key_id:
        with connect() as con:
            con.execute(
                """
                UPDATE api_keys
                SET status = ?, last_tested_at = ?, last_error = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    result["status"],
                    result["tested_at"],
                    None if result["ok"] else sanitize_provider_message(result["message"], key_value),
                    datetime.now().isoformat(timespec="seconds"),
                    key_id,
                ),
            )
    return result


def delete_api_key(key_id: int) -> None:
    with connect() as con:
        cur = con.execute("DELETE FROM api_keys WHERE id = ?", (key_id,))
    if cur.rowcount == 0:
        raise ValueError("API key not found")


RANKING_SNAPSHOT_TTL_DAYS = 7
RANKING_SNAPSHOT_FRESHNESS_NOTE = "DataForSEO Labs ranking data is updated weekly and should be treated as a ranking snapshot, not live rank tracking."


def normalize_ranking_snapshot_target(value: Any) -> str:
    text = clean_text(value) or ""
    if not text:
        raise ValueError("Target domain is required")
    parsed = urlparse(text if "://" in text else "https://" + text)
    host = (parsed.netloc or parsed.path.split("/")[0]).strip().lower()
    host = host.split("@")[-1].split(":")[0].removeprefix("www.")
    if not host or "/" in host or "\\" in host:
        raise ValueError("Enter a valid domain, such as example.com")
    if not re.match(r"^(?!-)(?:[a-z0-9-]{1,63}\.)+[a-z]{2,63}$", host):
        raise ValueError("Enter a valid domain, such as example.com")
    return host


def snapshot_number(value: Any) -> float | None:
    return as_float(value)


def deep_get(data: Any, *paths: str) -> Any:
    for path in paths:
        current = data
        found = True
        for part in path.split("."):
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                found = False
                break
        if found and current is not None:
            return current
    return None


def dataforseo_api_key() -> sqlite3.Row:
    with connect() as con:
        row = con.execute(
            """
            SELECT * FROM api_keys
            WHERE lower(provider) IN ('dataforseo', 'dataforseo', 'data for seo')
            ORDER BY CASE status WHEN 'valid' THEN 1 ELSE 2 END, updated_at DESC, id DESC
            LIMIT 1
            """
        ).fetchone()
    if not row:
        raise ValueError("Add and test DataForSEO credentials in Settings before running Ranking Snapshot.")
    if ":" not in row["key_value"]:
        raise ValueError("DataForSEO requires API login and API password.")
    return row


def dataforseo_post(path: str, payload: list[dict[str, Any]], timeout: int = 90) -> dict[str, Any]:
    key = dataforseo_api_key()
    meta = AI_PROVIDERS["dataforseo"]
    root = (clean_text(key["base_url"] if "base_url" in key.keys() else None) or meta["base_url"]).rstrip("/")
    token = base64.b64encode(str(key["key_value"]).encode("utf-8")).decode("ascii")
    request = urllib.request.Request(
        f"{root}{path}",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Basic {token}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8", errors="replace")
            data = json.loads(raw) if raw else {}
    except urllib.error.HTTPError as exc:
        body = exc.read(4096).decode("utf-8", errors="replace")
        raise ValueError(f"DataForSEO HTTP {exc.code}: {sanitize_provider_message(body or exc.reason or str(exc), key['key_value'])}") from exc
    except Exception as exc:
        raise ValueError(f"DataForSEO request failed: {sanitize_provider_message(str(exc), key['key_value'])}") from exc
    status_code = int(data.get("status_code") or 0)
    if status_code and status_code not in (20000, 20100):
        raise ValueError(f"DataForSEO error: {sanitize_provider_message(data.get('status_message') or 'Request failed', key['key_value'])}")
    return data


def ranking_snapshot_payload(
    target: str,
    location_code: int = 2840,
    language_code: str = "en",
    limit: int | None = None,
    include_subdomains: bool = False,
    order_by: list[str] | None = None,
) -> list[dict[str, Any]]:
    item: dict[str, Any] = {
        "target": target,
        "location_code": int(location_code or 2840),
        "language_code": clean_text(language_code) or "en",
    }
    if limit:
        item["limit"] = max(1, min(int(limit), 1000))
    if include_subdomains:
        item["include_subdomains"] = True
    if order_by:
        item["order_by"] = order_by
    return [item]


def dataforseo_result_items(data: dict[str, Any]) -> list[dict[str, Any]]:
    tasks = data.get("tasks") if isinstance(data, dict) else []
    if not isinstance(tasks, list) or not tasks:
        return []
    task = tasks[0] if isinstance(tasks[0], dict) else {}
    if int(task.get("status_code") or 20000) not in (20000, 20100):
        raise ValueError(clean_text(task.get("status_message")) or "DataForSEO task failed")
    result = task.get("result") or []
    if isinstance(result, dict):
        result = [result]
    if not isinstance(result, list) or not result:
        return []
    first = result[0] if isinstance(result[0], dict) else {}
    items = first.get("items")
    if isinstance(items, list):
        return [item for item in items if isinstance(item, dict)]
    return [item for item in result if isinstance(item, dict)]


def normalize_serp_features(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, str):
        return [value]
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            if isinstance(item, str):
                result.append(item)
            elif isinstance(item, dict):
                label = item.get("type") or item.get("name") or item.get("feature")
                if label:
                    result.append(str(label))
        return sorted(set(result))
    return []


def normalize_ranked_keyword_item(item: dict[str, Any]) -> dict[str, Any]:
    features = normalize_serp_features(
        deep_get(item, "keyword_data.serp_info.serp_item_types", "serp_info.serp_item_types", "serp_features", "serp_item_types")
    )
    feature_text = " ".join(features).lower()
    rank_item = deep_get(item, "ranked_serp_element.serp_item") or {}
    if not isinstance(rank_item, dict):
        rank_item = {}
    return {
        "keyword": clean_text(deep_get(item, "keyword_data.keyword", "keyword", "keyword_data.keyword_info.keyword")) or "",
        "rankingUrl": clean_text(deep_get(item, "ranked_serp_element.serp_item.url", "ranked_serp_element.url", "url", "ranking_url")),
        "position": snapshot_number(deep_get(item, "ranked_serp_element.serp_item.rank_absolute", "ranked_serp_element.serp_item.rank_group", "rank_absolute", "position")),
        "previousPosition": snapshot_number(deep_get(item, "ranked_serp_element.serp_item.previous_rank_absolute", "rank_changes.previous_rank_absolute", "previous_position")),
        "searchVolume": snapshot_number(deep_get(item, "keyword_data.keyword_info.search_volume", "keyword_info.search_volume", "search_volume")),
        "cpc": snapshot_number(deep_get(item, "keyword_data.keyword_info.cpc", "keyword_info.cpc", "cpc")),
        "competition": snapshot_number(deep_get(item, "keyword_data.keyword_info.competition", "keyword_info.competition", "competition")),
        "competitionLevel": clean_text(deep_get(item, "keyword_data.keyword_info.competition_level", "keyword_info.competition_level", "competition_level")),
        "keywordDifficulty": snapshot_number(deep_get(item, "keyword_data.keyword_properties.keyword_difficulty", "keyword_properties.keyword_difficulty", "keyword_difficulty")),
        "estimatedTraffic": snapshot_number(deep_get(item, "ranked_serp_element.serp_item.etv", "ranked_serp_element.etv", "metrics.organic.etv", "etv")),
        "trafficCost": snapshot_number(deep_get(item, "ranked_serp_element.serp_item.estimated_paid_traffic_cost", "traffic_cost", "estimated_paid_traffic_cost")),
        "serpFeatures": features,
        "aiOverviewPresent": "ai_overview" in feature_text or "ai overview" in feature_text,
        "aiOverviewReference": bool(deep_get(item, "ranked_serp_element.serp_item.ai_overview_reference", "ai_overview_reference")) or str(rank_item.get("type") or "").lower() == "ai_overview",
        "intent": clean_text(deep_get(item, "keyword_data.search_intent_info.main_intent", "search_intent_info.main_intent", "intent")),
        "lastUpdated": clean_text(deep_get(item, "keyword_data.keyword_info.last_updated_time", "last_updated_time", "last_updated")),
    }


def normalize_relevant_page_item(item: dict[str, Any]) -> dict[str, Any]:
    organic = deep_get(item, "metrics.organic") or {}
    paid = deep_get(item, "metrics.paid") or {}
    dist = deep_get(item, "metrics.organic.pos_distribution", "ranking_distribution.organic", "metrics.organic.ranking_distribution") or {}
    if not isinstance(organic, dict):
        organic = {}
    if not isinstance(paid, dict):
        paid = {}
    if not isinstance(dist, dict):
        dist = {}
    return {
        "url": clean_text(item.get("page_address") or item.get("url") or item.get("target")) or "",
        "organicKeywords": snapshot_number(organic.get("count") or item.get("organic_keywords")),
        "organicTraffic": snapshot_number(organic.get("etv") or item.get("organic_traffic")),
        "organicTrafficCost": snapshot_number(organic.get("estimated_paid_traffic_cost") or item.get("organic_traffic_cost")),
        "top1": snapshot_number(dist.get("pos_1") or dist.get("top1") or item.get("top1")),
        "top3": snapshot_number(dist.get("pos_2_3") or dist.get("top3") or item.get("top3")),
        "top10": snapshot_number(dist.get("pos_4_10") or dist.get("top10") or item.get("top10")),
        "top20": snapshot_number(dist.get("pos_11_20") or dist.get("top20") or item.get("top20")),
        "top100": snapshot_number(dist.get("pos_21_100") or dist.get("top100") or item.get("top100")),
        "paidKeywords": snapshot_number(paid.get("count") or item.get("paid_keywords")),
        "paidTraffic": snapshot_number(paid.get("etv") or item.get("paid_traffic")),
    }


def normalize_domain_overview_item(item: dict[str, Any], target: str, location_code: int, language_code: str) -> dict[str, Any]:
    organic = deep_get(item, "metrics.organic") or deep_get(item, "organic") or {}
    paid = deep_get(item, "metrics.paid") or deep_get(item, "paid") or {}
    dist = deep_get(item, "metrics.organic.pos_distribution", "ranking_distribution.organic", "organic.pos_distribution") or {}
    if not isinstance(organic, dict):
        organic = {}
    if not isinstance(paid, dict):
        paid = {}
    if not isinstance(dist, dict):
        dist = {}
    return {
        "target": target,
        "locationCode": location_code,
        "languageCode": language_code,
        "organicKeywords": snapshot_number(organic.get("count") or item.get("organic_keywords")),
        "organicTraffic": snapshot_number(organic.get("etv") or item.get("organic_traffic")),
        "organicTrafficCost": snapshot_number(organic.get("estimated_paid_traffic_cost") or item.get("organic_traffic_cost")),
        "paidKeywords": snapshot_number(paid.get("count") or item.get("paid_keywords")),
        "paidTraffic": snapshot_number(paid.get("etv") or item.get("paid_traffic")),
        "rankingDistribution": {
            "top1": snapshot_number(dist.get("pos_1") or dist.get("top1")),
            "top3": snapshot_number(dist.get("pos_2_3") or dist.get("top3")),
            "top10": snapshot_number(dist.get("pos_4_10") or dist.get("top10")),
            "top20": snapshot_number(dist.get("pos_11_20") or dist.get("top20")),
            "top100": snapshot_number(dist.get("pos_21_100") or dist.get("top100")),
        },
        "dataSource": "DataForSEO Labs",
        "dataFreshnessNote": RANKING_SNAPSHOT_FRESHNESS_NOTE,
    }


def opportunity_action(row: dict[str, Any]) -> str:
    position = as_float(row.get("position"))
    if row.get("aiOverviewPresent") and not row.get("aiOverviewReference"):
        return "Add concise answer blocks, entity-rich explanations, citations, and schema where relevant."
    if position is not None and 4 <= position <= 10:
        return "Improve on-page optimization, internal links, title/meta, and content depth to push into top 3."
    if position is not None and 11 <= position <= 20:
        return "Refresh content and strengthen topical coverage to move onto page one."
    if position is not None and 21 <= position <= 30:
        return "Consider content expansion, backlinks/internal links, or a dedicated page."
    return "Review the ranking URL and strengthen the page around the keyword intent."


def classify_ranking_opportunities(keywords: list[dict[str, Any]], high_volume_threshold: int = 1000) -> list[dict[str, Any]]:
    opportunities: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in keywords:
        keyword = clean_text(row.get("keyword")) or ""
        position = as_float(row.get("position"))
        previous = as_float(row.get("previousPosition"))
        volume = as_float(row.get("searchVolume")) or 0
        labels: list[str] = []
        if position is not None and 4 <= position <= 20:
            labels.append("Striking Distance")
        if position is not None and 4 <= position <= 10:
            labels.append("Page One Wins")
        if position is not None and 11 <= position <= 20:
            labels.append("Page Two Opportunities")
        if position is not None and 4 <= position <= 30 and volume >= high_volume_threshold:
            labels.append("High Volume Opportunities")
        if position is not None and (11 <= position <= 30 or (previous is not None and previous < position)):
            labels.append("Content Refresh Candidates")
        if row.get("aiOverviewPresent") and not row.get("aiOverviewReference"):
            labels.append("AI Overview Opportunities")
        for label in labels:
            key = (keyword.lower(), label)
            if not keyword or key in seen:
                continue
            seen.add(key)
            opportunities.append(
                {
                    "opportunityType": label,
                    "keyword": keyword,
                    "position": position,
                    "rankingUrl": row.get("rankingUrl"),
                    "searchVolume": row.get("searchVolume"),
                    "estimatedTraffic": row.get("estimatedTraffic"),
                    "cpc": row.get("cpc"),
                    "serpFeatures": row.get("serpFeatures") or [],
                    "recommendedAction": opportunity_action(row),
                }
            )
    return opportunities


def ranking_snapshot_row_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    data = row_to_dict(row)
    if not data:
        return None
    try:
        data["overview"] = json.loads(data.pop("overview_json") or "{}")
    except json.JSONDecodeError:
        data["overview"] = {}
    try:
        data["errors"] = json.loads(data.pop("errors_json") or "{}")
    except json.JSONDecodeError:
        data["errors"] = {}
    data["partial"] = bool(data["errors"])
    return data


def get_ranking_snapshot(snapshot_id: int) -> dict[str, Any]:
    with connect() as con:
        snapshot = ranking_snapshot_row_public(con.execute("SELECT * FROM ranking_snapshots WHERE id = ?", (snapshot_id,)).fetchone())
        if not snapshot:
            raise ValueError("Ranking snapshot not found")
        keyword_rows = con.execute("SELECT * FROM ranking_snapshot_keywords WHERE snapshot_id = ? ORDER BY COALESCE(search_volume, 0) DESC, COALESCE(position, 999)", (snapshot_id,)).fetchall()
        page_rows = con.execute("SELECT * FROM ranking_snapshot_pages WHERE snapshot_id = ? ORDER BY COALESCE(organic_traffic, 0) DESC, COALESCE(organic_keywords, 0) DESC", (snapshot_id,)).fetchall()
    keywords = []
    for row in keyword_rows:
        item = row_to_dict(row) or {}
        item["rankingUrl"] = item.pop("ranking_url")
        item["previousPosition"] = item.pop("previous_position")
        item["searchVolume"] = item.pop("search_volume")
        item["competitionLevel"] = item.pop("competition_level")
        item["keywordDifficulty"] = item.pop("keyword_difficulty")
        item["estimatedTraffic"] = item.pop("estimated_traffic")
        item["trafficCost"] = item.pop("traffic_cost")
        item["aiOverviewPresent"] = bool(item.pop("ai_overview_present"))
        item["aiOverviewReference"] = bool(item.pop("ai_overview_reference"))
        item["lastUpdated"] = item.pop("last_updated")
        try:
            item["serpFeatures"] = json.loads(item.pop("serp_features_json") or "[]")
        except json.JSONDecodeError:
            item["serpFeatures"] = []
        keywords.append(item)
    pages = []
    for row in page_rows:
        item = row_to_dict(row) or {}
        item["organicKeywords"] = item.pop("organic_keywords")
        item["organicTraffic"] = item.pop("organic_traffic")
        item["organicTrafficCost"] = item.pop("organic_traffic_cost")
        item["paidKeywords"] = item.pop("paid_keywords")
        item["paidTraffic"] = item.pop("paid_traffic")
        pages.append(item)
    opportunities = classify_ranking_opportunities(keywords)
    return {
        "snapshot": snapshot,
        "overview": snapshot.get("overview") or {},
        "keywords": keywords,
        "pages": pages,
        "opportunities": opportunities,
        "savedTargets": list_ranking_optimization_targets(snapshot_id=snapshot_id),
        "meta": {
            "target": snapshot["target"],
            "location_code": snapshot["location_code"],
            "language_code": snapshot["language_code"],
            "limit": snapshot["limit_value"],
            "source": snapshot["source"],
            "freshness": snapshot["freshness"],
            "generated_at": snapshot["created_at"],
            "cached": False,
            "partial": snapshot["partial"],
            "errors": snapshot["errors"],
        },
    }


RANKING_TARGET_STATUSES = {"new", "selected", "in_cora", "in_entity_explorer", "content_plan_created", "optimized", "archived"}


def ranking_target_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    data = row_to_dict(row)
    if not data:
        return None
    data["snapshotId"] = data.pop("snapshot_id", None)
    data["projectId"] = data.pop("project_id", None)
    data["bestPosition"] = data.pop("best_position", None)
    data["rankingKeywords"] = data.pop("ranking_keywords", None)
    data["opportunityCount"] = data.pop("opportunity_count", None)
    data["totalSearchVolume"] = data.pop("total_search_volume", None)
    data["estimatedTraffic"] = data.pop("estimated_traffic", None)
    data["pageOrganicTraffic"] = data.pop("page_organic_traffic", None)
    data["pageOrganicKeywords"] = data.pop("page_organic_keywords", None)
    data["priorityType"] = data.pop("priority_type", None)
    data["opportunityScore"] = data.pop("opportunity_score", None)
    data["recommendedAction"] = data.pop("recommended_action", None)
    try:
        data["topKeywords"] = json.loads(data.pop("top_keywords_json") or "[]")
    except json.JSONDecodeError:
        data["topKeywords"] = []
    return data


def list_ranking_optimization_targets(
    project_id: int | None = None,
    snapshot_id: int | None = None,
    target_ids: list[int] | None = None,
) -> list[dict[str, Any]]:
    where: list[str] = []
    params: list[Any] = []
    if project_id:
        where.append("rot.project_id = ?")
        params.append(project_id)
    if snapshot_id:
        where.append("rot.snapshot_id = ?")
        params.append(snapshot_id)
    if target_ids:
        placeholders = ",".join("?" for _ in target_ids)
        where.append(f"rot.id IN ({placeholders})")
        params.extend(target_ids)
    clause = f"WHERE {' AND '.join(where)}" if where else ""
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT rot.*, rs.target AS snapshot_target, rs.created_at AS snapshot_created_at, p.name AS project_name
            FROM ranking_optimization_targets rot
            JOIN ranking_snapshots rs ON rs.id = rot.snapshot_id
            LEFT JOIN projects p ON p.id = rot.project_id
            {clause}
            ORDER BY COALESCE(rot.opportunity_score, 0) DESC, COALESCE(rot.opportunity_count, 0) DESC, rot.updated_at DESC
            """,
            params,
        ).fetchall()
    return [ranking_target_public(row) or {} for row in rows]


def save_ranking_optimization_targets(payload: dict[str, Any]) -> dict[str, Any]:
    snapshot_id = int(payload.get("snapshot_id") or 0)
    project_id = int(payload.get("project_id") or 0) or None
    targets = payload.get("targets") if isinstance(payload.get("targets"), list) else []
    status = clean_text(payload.get("status")) or "selected"
    if status not in RANKING_TARGET_STATUSES:
        status = "selected"
    if not snapshot_id:
        raise ValueError("snapshot_id is required")
    if not targets:
        raise ValueError("Select at least one optimization target")
    now = datetime.now().isoformat(timespec="seconds")
    saved_ids: list[int] = []
    with connect() as con:
        snapshot = con.execute("SELECT project_id FROM ranking_snapshots WHERE id = ?", (snapshot_id,)).fetchone()
        if not snapshot:
            raise ValueError("Ranking snapshot not found")
        resolved_project_id = snapshot["project_id"]
        if project_id and resolved_project_id and int(project_id) != int(resolved_project_id):
            raise ValueError("Optimization targets must be saved to the same client as the Ranking Snapshot")
        for item in targets:
            url = clean_text(item.get("url") or item.get("rankingUrl"))
            if not url:
                continue
            top_keywords = item.get("topKeywords") if isinstance(item.get("topKeywords"), list) else []
            cur = con.execute(
                """
                INSERT INTO ranking_optimization_targets
                (snapshot_id, project_id, url, keyword, best_position, ranking_keywords, opportunity_count,
                 total_search_volume, estimated_traffic, page_organic_traffic, page_organic_keywords, top10,
                 priority_type, opportunity_score, recommended_action, top_keywords_json, status, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(snapshot_id, url) DO UPDATE SET
                  keyword = excluded.keyword,
                  best_position = excluded.best_position,
                  ranking_keywords = excluded.ranking_keywords,
                  opportunity_count = excluded.opportunity_count,
                  total_search_volume = excluded.total_search_volume,
                  estimated_traffic = excluded.estimated_traffic,
                  page_organic_traffic = excluded.page_organic_traffic,
                  page_organic_keywords = excluded.page_organic_keywords,
                  top10 = excluded.top10,
                  priority_type = excluded.priority_type,
                  opportunity_score = excluded.opportunity_score,
                  recommended_action = excluded.recommended_action,
                  top_keywords_json = excluded.top_keywords_json,
                  status = excluded.status,
                  notes = COALESCE(excluded.notes, ranking_optimization_targets.notes),
                  updated_at = excluded.updated_at
                """,
                (
                    snapshot_id,
                    resolved_project_id,
                    url,
                    clean_text(item.get("keyword")),
                    item.get("bestPosition"),
                    item.get("rankingKeywords"),
                    item.get("opportunityCount"),
                    item.get("totalSearchVolume"),
                    item.get("estimatedTraffic"),
                    item.get("pageOrganicTraffic"),
                    item.get("pageOrganicKeywords"),
                    item.get("top10"),
                    clean_text(item.get("priorityType")),
                    item.get("opportunityScore"),
                    clean_text(item.get("recommendedAction")),
                    json.dumps(top_keywords),
                    status,
                    clean_text(item.get("notes")),
                    now,
                    now,
                ),
            )
            row = con.execute("SELECT id FROM ranking_optimization_targets WHERE snapshot_id = ? AND url = ?", (snapshot_id, url)).fetchone()
            if row:
                saved_ids.append(int(row["id"]))
    return {"targets": list_ranking_optimization_targets(target_ids=saved_ids), "saved_ids": saved_ids}


def update_ranking_optimization_target_status(target_ids: list[int], status: str, project_id: int | None = None) -> dict[str, Any]:
    clean_status = clean_text(status) or ""
    if clean_status not in RANKING_TARGET_STATUSES:
        raise ValueError("Invalid optimization target status")
    ids: list[int] = []
    for value in target_ids:
        try:
            item_id = int(value)
        except (TypeError, ValueError):
            continue
        if item_id:
            ids.append(item_id)
    if not ids:
        raise ValueError("Select at least one saved target")
    now = datetime.now().isoformat(timespec="seconds")
    placeholders = ",".join("?" for _ in ids)
    with connect() as con:
        if project_id:
            count = con.execute(
                f"SELECT COUNT(*) FROM ranking_optimization_targets WHERE id IN ({placeholders}) AND project_id = ?",
                [*ids, project_id],
            ).fetchone()[0]
            if int(count) != len(ids):
                raise ValueError("Optimization targets must belong to the selected client")
        con.execute(
            f"UPDATE ranking_optimization_targets SET status = ?, updated_at = ? WHERE id IN ({placeholders})",
            [clean_status, now, *ids],
        )
    return {"targets": list_ranking_optimization_targets(target_ids=ids)}


def list_ranking_snapshots(project_id: int | None = None) -> list[dict[str, Any]]:
    params: list[Any] = []
    where = ""
    if project_id:
        where = "WHERE rs.project_id = ?"
        params.append(project_id)
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT rs.*,
                   p.name AS project_name,
                   (SELECT COUNT(*) FROM ranking_snapshot_keywords k WHERE k.snapshot_id = rs.id) AS keyword_count,
                   (SELECT COUNT(*) FROM ranking_snapshot_pages pg WHERE pg.snapshot_id = rs.id) AS page_count
            FROM ranking_snapshots rs
            LEFT JOIN projects p ON p.id = rs.project_id
            {where}
            ORDER BY rs.created_at DESC, rs.id DESC
            LIMIT 100
            """,
            params,
        ).fetchall()
    return [ranking_snapshot_row_public(row) or {} for row in rows]


def keyword_compare_key(row: dict[str, Any]) -> str:
    return str(row.get("keyword") or "").strip().lower()


def page_compare_key(row: dict[str, Any]) -> str:
    return comparable_url(row.get("url") or "")


def compare_ranking_snapshots(base_id: int, compare_id: int) -> dict[str, Any]:
    if base_id == compare_id:
        raise ValueError("Choose two different snapshots to compare")
    base = get_ranking_snapshot(base_id)
    compare = get_ranking_snapshot(compare_id)
    base_snapshot = base["snapshot"]
    compare_snapshot = compare["snapshot"]
    if base_snapshot.get("project_id") and compare_snapshot.get("project_id") and base_snapshot.get("project_id") != compare_snapshot.get("project_id"):
        raise ValueError("Snapshots must belong to the same client")

    base_keywords = {keyword_compare_key(row): row for row in base.get("keywords") or [] if keyword_compare_key(row)}
    compare_keywords = {keyword_compare_key(row): row for row in compare.get("keywords") or [] if keyword_compare_key(row)}
    keyword_rows: list[dict[str, Any]] = []
    new_keywords: list[dict[str, Any]] = []
    lost_keywords: list[dict[str, Any]] = []
    improved: list[dict[str, Any]] = []
    declined: list[dict[str, Any]] = []
    for key in sorted(set(base_keywords) | set(compare_keywords)):
        before = base_keywords.get(key)
        after = compare_keywords.get(key)
        if before and not after:
            row = {
                "keyword": before.get("keyword"),
                "rankingUrl": before.get("rankingUrl"),
                "basePosition": before.get("position"),
                "comparePosition": None,
                "positionDelta": None,
                "searchVolume": before.get("searchVolume"),
                "estimatedTrafficDelta": -(as_float(before.get("estimatedTraffic")) or 0),
                "status": "lost",
            }
            lost_keywords.append(row)
            keyword_rows.append(row)
            continue
        if after and not before:
            row = {
                "keyword": after.get("keyword"),
                "rankingUrl": after.get("rankingUrl"),
                "basePosition": None,
                "comparePosition": after.get("position"),
                "positionDelta": None,
                "searchVolume": after.get("searchVolume"),
                "estimatedTrafficDelta": as_float(after.get("estimatedTraffic")) or 0,
                "status": "new",
            }
            new_keywords.append(row)
            keyword_rows.append(row)
            continue
        if not before or not after:
            continue
        base_pos = as_float(before.get("position"))
        compare_pos = as_float(after.get("position"))
        position_delta = compare_pos - base_pos if base_pos is not None and compare_pos is not None else None
        traffic_delta = (as_float(after.get("estimatedTraffic")) or 0) - (as_float(before.get("estimatedTraffic")) or 0)
        status = "unchanged"
        if position_delta is not None and position_delta < 0:
            status = "improved"
        elif position_delta is not None and position_delta > 0:
            status = "declined"
        row = {
            "keyword": after.get("keyword") or before.get("keyword"),
            "rankingUrl": after.get("rankingUrl") or before.get("rankingUrl"),
            "basePosition": base_pos,
            "comparePosition": compare_pos,
            "positionDelta": position_delta,
            "searchVolume": after.get("searchVolume") or before.get("searchVolume"),
            "estimatedTrafficDelta": traffic_delta,
            "status": status,
        }
        keyword_rows.append(row)
        if status == "improved":
            improved.append(row)
        elif status == "declined":
            declined.append(row)

    base_pages = {page_compare_key(row): row for row in base.get("pages") or [] if page_compare_key(row)}
    compare_pages = {page_compare_key(row): row for row in compare.get("pages") or [] if page_compare_key(row)}
    page_rows: list[dict[str, Any]] = []
    for key in sorted(set(base_pages) | set(compare_pages)):
        before = base_pages.get(key)
        after = compare_pages.get(key)
        source = after or before or {}
        base_traffic = as_float(before.get("organicTraffic")) if before else None
        compare_traffic = as_float(after.get("organicTraffic")) if after else None
        base_count = as_float(before.get("organicKeywords")) if before else None
        compare_count = as_float(after.get("organicKeywords")) if after else None
        traffic_delta = (compare_traffic or 0) - (base_traffic or 0)
        keyword_delta = (compare_count or 0) - (base_count or 0)
        if before and not after:
            status = "lost"
        elif after and not before:
            status = "new"
        elif traffic_delta > 0:
            status = "gained"
        elif traffic_delta < 0:
            status = "lost_traffic"
        else:
            status = "unchanged"
        page_rows.append(
            {
                "url": source.get("url"),
                "baseOrganicTraffic": base_traffic,
                "compareOrganicTraffic": compare_traffic,
                "organicTrafficDelta": traffic_delta,
                "baseOrganicKeywords": base_count,
                "compareOrganicKeywords": compare_count,
                "organicKeywordDelta": keyword_delta,
                "status": status,
            }
        )
    page_rows.sort(key=lambda row: abs(as_float(row.get("organicTrafficDelta")) or 0), reverse=True)
    keyword_rows.sort(key=lambda row: (row["status"] not in {"improved", "declined", "new", "lost"}, abs(as_float(row.get("positionDelta")) or 0)), reverse=False)
    return {
        "base": base_snapshot,
        "compare": compare_snapshot,
        "summary": {
            "newKeywords": len(new_keywords),
            "lostKeywords": len(lost_keywords),
            "improvedKeywords": len(improved),
            "declinedKeywords": len(declined),
            "pageGains": len([row for row in page_rows if row["status"] in {"new", "gained"}]),
            "pageLosses": len([row for row in page_rows if row["status"] in {"lost", "lost_traffic"}]),
        },
        "keywords": keyword_rows,
        "newKeywords": new_keywords,
        "lostKeywords": lost_keywords,
        "improvedKeywords": improved,
        "declinedKeywords": declined,
        "pages": page_rows,
    }


def save_ranking_snapshot(
    project_id: int | None,
    target: str,
    location_code: int,
    language_code: str,
    limit: int,
    include_subdomains: bool,
    overview: dict[str, Any],
    keywords: list[dict[str, Any]],
    pages: list[dict[str, Any]],
    errors: dict[str, str],
) -> int:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO ranking_snapshots
            (project_id, target, location_code, language_code, limit_value, include_subdomains, overview_json, errors_json, source, freshness, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (project_id, target, location_code, language_code, limit, 1 if include_subdomains else 0, json.dumps(overview), json.dumps(errors), "DataForSEO Labs", "weekly", now),
        )
        snapshot_id = int(cur.lastrowid)
        for row in keywords:
            con.execute(
                """
                INSERT INTO ranking_snapshot_keywords
                (snapshot_id, keyword, ranking_url, position, previous_position, search_volume, cpc, competition, competition_level,
                 keyword_difficulty, estimated_traffic, traffic_cost, serp_features_json, ai_overview_present, ai_overview_reference, intent, last_updated, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    row.get("keyword") or "",
                    row.get("rankingUrl"),
                    row.get("position"),
                    row.get("previousPosition"),
                    row.get("searchVolume"),
                    row.get("cpc"),
                    row.get("competition"),
                    row.get("competitionLevel"),
                    row.get("keywordDifficulty"),
                    row.get("estimatedTraffic"),
                    row.get("trafficCost"),
                    json.dumps(row.get("serpFeatures") or []),
                    1 if row.get("aiOverviewPresent") else 0,
                    1 if row.get("aiOverviewReference") else 0,
                    row.get("intent"),
                    row.get("lastUpdated"),
                    now,
                ),
            )
        for row in pages:
            con.execute(
                """
                INSERT INTO ranking_snapshot_pages
                (snapshot_id, url, organic_keywords, organic_traffic, organic_traffic_cost, top1, top3, top10, top20, top100, paid_keywords, paid_traffic, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    row.get("url") or "",
                    row.get("organicKeywords"),
                    row.get("organicTraffic"),
                    row.get("organicTrafficCost"),
                    row.get("top1"),
                    row.get("top3"),
                    row.get("top10"),
                    row.get("top20"),
                    row.get("top100"),
                    row.get("paidKeywords"),
                    row.get("paidTraffic"),
                    now,
                ),
            )
    return snapshot_id


def create_or_get_ranking_snapshot(payload: dict[str, Any]) -> dict[str, Any]:
    target = normalize_ranking_snapshot_target(payload.get("target"))
    project_id = int(payload["project_id"]) if payload.get("project_id") else None
    location_code = int(payload.get("location_code") or 2840)
    language_code = clean_text(payload.get("language_code")) or "en"
    limit = max(1, min(int(payload.get("limit") or 1000), 1000))
    include_subdomains = bool(payload.get("include_subdomains"))
    force_refresh = bool(payload.get("force_refresh"))
    cutoff = (datetime.now() - timedelta(days=RANKING_SNAPSHOT_TTL_DAYS)).isoformat(timespec="seconds")
    if not force_refresh:
        with connect() as con:
            cached = con.execute(
                """
                SELECT id FROM ranking_snapshots
                WHERE COALESCE(project_id, 0) = COALESCE(?, 0)
                  AND target = ? AND location_code = ? AND language_code = ? AND include_subdomains = ?
                  AND created_at >= ?
                ORDER BY created_at DESC, id DESC
                LIMIT 1
                """,
                (project_id, target, location_code, language_code, 1 if include_subdomains else 0, cutoff),
            ).fetchone()
        if cached:
            data = get_ranking_snapshot(int(cached["id"]))
            data["meta"]["cached"] = True
            return data

    errors: dict[str, str] = {}
    overview: dict[str, Any] = normalize_domain_overview_item({}, target, location_code, language_code)
    keywords: list[dict[str, Any]] = []
    pages: list[dict[str, Any]] = []
    endpoints = {
        "overview": "/v3/dataforseo_labs/google/domain_rank_overview/live",
        "keywords": "/v3/dataforseo_labs/google/ranked_keywords/live",
        "pages": "/v3/dataforseo_labs/google/relevant_pages/live",
    }
    try:
        data = dataforseo_post(endpoints["overview"], ranking_snapshot_payload(target, location_code, language_code, include_subdomains=include_subdomains))
        items = dataforseo_result_items(data)
        if items:
            overview = normalize_domain_overview_item(items[0], target, location_code, language_code)
    except Exception as exc:
        errors["overview"] = str(exc)
    try:
        data = dataforseo_post(
            endpoints["keywords"],
            ranking_snapshot_payload(target, location_code, language_code, limit, include_subdomains, ["keyword_data.keyword_info.search_volume,desc"]),
        )
        keywords = [normalize_ranked_keyword_item(item) for item in dataforseo_result_items(data)]
        keywords = [row for row in keywords if row.get("keyword")]
    except Exception as exc:
        errors["keywords"] = str(exc)
    try:
        data = dataforseo_post(
            endpoints["pages"],
            ranking_snapshot_payload(target, location_code, language_code, limit, include_subdomains, ["metrics.organic.etv,desc"]),
        )
        pages = [normalize_relevant_page_item(item) for item in dataforseo_result_items(data)]
        pages = [row for row in pages if row.get("url")]
    except Exception as exc:
        errors["pages"] = str(exc)
    if errors and not keywords and not pages and not overview.get("organicKeywords"):
        raise ValueError("Snapshot failed. Please check the domain and try again.")
    snapshot_id = save_ranking_snapshot(project_id, target, location_code, language_code, limit, include_subdomains, overview, keywords, pages, errors)
    data = get_ranking_snapshot(snapshot_id)
    data["meta"]["cached"] = False
    return data


ENTITY_LSI_PROMPT_VERSION = "entity-lsi-v1"
CORA_ENTITY_IMPORT_PROMPT_VERSION = "cora-xlsx-import-v1"
ENTITY_LSI_DEPTH_LIMITS = {
    1: {"entities": 10, "lsi_terms": 10, "related_keywords": 10, "questions": 5, "topic_clusters": 3},
    2: {"entities": 20, "lsi_terms": 20, "related_keywords": 20, "questions": 8, "topic_clusters": 5},
    3: {"entities": 35, "lsi_terms": 35, "related_keywords": 35, "questions": 12, "topic_clusters": 7},
    4: {"entities": 50, "lsi_terms": 50, "related_keywords": 50, "questions": 18, "topic_clusters": 10},
    5: {"entities": 75, "lsi_terms": 75, "related_keywords": 75, "questions": 25, "topic_clusters": 14},
}
LLM_PROVIDER_KEYS = {"openai", "anthropic", "google", "xai", "perplexity"}
ENTITY_LSI_OUTPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "summary": {"type": "string"},
        "entities": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "type": {"type": "string"},
                    "relevance_score": {"type": "number"},
                    "suggested_usage": {"type": "string"},
                },
                "required": ["name"],
            },
        },
        "lsi_terms": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "term": {"type": "string"},
                    "relevance_score": {"type": "number"},
                    "intent": {"type": "string"},
                },
                "required": ["term"],
            },
        },
        "related_keywords": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "keyword": {"type": "string"},
                    "intent": {"type": "string"},
                    "funnel_stage": {"type": "string"},
                },
                "required": ["keyword"],
            },
        },
        "questions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "question": {"type": "string"},
                    "intent": {"type": "string"},
                    "content_opportunity": {"type": "string"},
                },
                "required": ["question"],
            },
        },
        "topic_clusters": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "cluster": {"type": "string"},
                    "terms": {"type": "array", "items": {"type": "string"}},
                    "content_angle": {"type": "string"},
                },
                "required": ["cluster"],
            },
        },
        "warnings": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["summary", "entities", "lsi_terms", "related_keywords", "questions", "topic_clusters", "warnings"],
}


def clamp_entity_depth(value: Any) -> int:
    try:
        depth = int(value)
    except (TypeError, ValueError):
        depth = 3
    return max(1, min(5, depth))


def entity_lsi_run_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    data = row_to_dict(row)
    if not data:
        return None
    result_json = data.pop("result_json", None)
    try:
        data["result"] = json.loads(result_json) if result_json else {}
    except json.JSONDecodeError:
        data["result"] = {"warnings": ["Saved result JSON could not be parsed."]}
    data.pop("prompt", None)
    data.pop("raw_response", None)
    return data


def parse_llm_json_text(text: str) -> dict[str, Any]:
    cleaned = (text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if not match:
            raise ValueError("The LLM response did not include valid JSON")
        parsed = json.loads(match.group(0))
    if not isinstance(parsed, dict):
        raise ValueError("The LLM response JSON must be an object")
    for key in ("entities", "lsi_terms", "related_keywords", "questions", "topic_clusters", "warnings"):
        value = parsed.get(key)
        if value is None:
            parsed[key] = []
        elif not isinstance(value, list):
            parsed[key] = [value]
    if not isinstance(parsed.get("summary"), str):
        parsed["summary"] = ""
    return parsed


def normalize_entity_lsi_result(parsed: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(parsed, dict):
        raise ValueError("The LLM response JSON must be an object")
    for key in ("entities", "lsi_terms", "related_keywords", "questions", "topic_clusters", "warnings"):
        value = parsed.get(key)
        if value is None:
            parsed[key] = []
        elif not isinstance(value, list):
            parsed[key] = [value]
    if not isinstance(parsed.get("summary"), str):
        parsed["summary"] = ""
    return parsed


def build_entity_lsi_prompt(
    project: sqlite3.Row,
    seed_keyword: str,
    depth: int,
    main_url: str,
    client_keywords: list[str],
) -> str:
    limits = ENTITY_LSI_DEPTH_LIMITS[depth]
    keyword_context = ", ".join(client_keywords[:40]) if client_keywords else "No other client keywords provided."
    return f"""You are building an SEO Entity and LSI exploration artifact.

Return only valid JSON with these exact keys:
summary: string
entities: array of objects with name, type, relevance_score, suggested_usage
lsi_terms: array of objects with term, relevance_score, intent
related_keywords: array of objects with keyword, intent, funnel_stage
questions: array of objects with question, intent, content_opportunity
topic_clusters: array of objects with cluster, terms, content_angle
warnings: array of strings

Client: {project['name']}
Client main URL: {main_url or 'Not provided'}
Seed keyword: {seed_keyword}
Other client keywords: {keyword_context}

Depth: {depth}
Target approximate counts:
entities: {limits['entities']}
lsi_terms: {limits['lsi_terms']}
related_keywords: {limits['related_keywords']}
questions: {limits['questions']}
topic_clusters: {limits['topic_clusters']}

Favor terms that are useful for on-page optimization, content briefs, headings, schema/entity coverage, and topical completeness. Do not include invented metrics. Use relevance_score from 1 to 100."""


TRANSIENT_PROVIDER_STATUS_CODES = {408, 409, 429, 500, 502, 503, 504}
PROVIDER_RETRY_DELAYS = [2, 6]


def post_json_request(
    url: str,
    headers: dict[str, str],
    payload: dict[str, Any],
    timeout: int = 90,
    retries: int = 2,
) -> dict[str, Any]:
    body = json.dumps(payload).encode("utf-8")
    attempts = max(1, retries + 1)
    last_error = ""
    for attempt in range(attempts):
        request = urllib.request.Request(url, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                raw = response.read().decode("utf-8", errors="replace")
            return json.loads(raw) if raw.strip() else {}
        except urllib.error.HTTPError as exc:
            raw_error = exc.read(2048).decode("utf-8", errors="replace")
            last_error = raw_error or f"HTTP {exc.code}: {exc.reason}"
            if exc.code not in TRANSIENT_PROVIDER_STATUS_CODES or attempt >= attempts - 1:
                suffix = f" after {attempt + 1} attempts" if attempt else ""
                raise RuntimeError(f"{last_error}{suffix}") from exc
        except (TimeoutError, socket.timeout) as exc:
            last_error = f"Provider request timed out after {timeout} seconds"
            if attempt >= attempts - 1:
                raise RuntimeError(f"{last_error} after {attempt + 1} attempts") from exc
        time.sleep(PROVIDER_RETRY_DELAYS[min(attempt, len(PROVIDER_RETRY_DELAYS) - 1)])
    raise RuntimeError(last_error or "Provider request failed")


def llm_request_timeout(provider_key: str, requested_timeout: int) -> int:
    minimums = {
        "openai": 240,
        "anthropic": 240,
        "google": 180,
        "xai": 180,
        "perplexity": 240,
    }
    return max(int(requested_timeout or 90), minimums.get(provider_key, 90))


def extract_perplexity_agent_text(data: dict[str, Any]) -> str:
    error = data.get("error")
    if isinstance(error, dict) and error.get("message"):
        raise RuntimeError(str(error["message"]))
    output_text = data.get("output_text")
    if isinstance(output_text, str):
        return output_text
    parts: list[str] = []
    for item in data.get("output") or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content") or []:
            if isinstance(content, dict) and isinstance(content.get("text"), str):
                parts.append(content["text"])
    return "\n".join(parts)


def perplexity_sonar_model_name(model: str) -> str:
    clean_model = clean_text(model) or "perplexity/sonar"
    if clean_model.startswith("perplexity/"):
        clean_model = clean_model.split("/", 1)[1]
    return clean_model


def is_perplexity_sonar_model(model: str) -> bool:
    return perplexity_sonar_model_name(model).startswith("sonar")


def call_llm_provider(
    provider_key: str,
    key_value: str,
    base_url: str,
    model: str,
    prompt: str,
    timeout: int = 90,
) -> str:
    root = base_url.rstrip("/")
    timeout = llm_request_timeout(provider_key, timeout)
    if provider_key == "anthropic":
        data = post_json_request(
            f"{root}/v1/messages",
            {
                "Accept": "application/json",
                "Content-Type": "application/json",
                "x-api-key": key_value,
                "anthropic-version": "2023-06-01",
            },
            {
                "model": model,
                "max_tokens": 5000,
                "tools": [
                    {
                        "name": "save_entity_lsi_exploration",
                        "description": "Save the structured Entity and LSI exploration result.",
                        "input_schema": ENTITY_LSI_OUTPUT_SCHEMA,
                    }
                ],
                "tool_choice": {"type": "tool", "name": "save_entity_lsi_exploration"},
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout,
        )
        for part in data.get("content", []):
            if isinstance(part, dict) and part.get("type") == "tool_use" and isinstance(part.get("input"), dict):
                return json.dumps(part["input"])
        return "\n".join(part.get("text", "") for part in data.get("content", []) if isinstance(part, dict))
    if provider_key == "google":
        data = post_json_request(
            f"{root}/v1beta/models/{quote(model, safe='')}:generateContent?key={quote(key_value)}",
            {"Accept": "application/json", "Content-Type": "application/json"},
            {
                "contents": [{"role": "user", "parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.2, "responseMimeType": "application/json"},
            },
            timeout,
        )
        candidates = data.get("candidates") or []
        if not candidates:
            return ""
        parts = candidates[0].get("content", {}).get("parts", [])
        return "\n".join(part.get("text", "") for part in parts if isinstance(part, dict))
    if provider_key == "perplexity":
        if is_perplexity_sonar_model(model):
            data = post_json_request(
                f"{root}/v1/sonar",
                {
                    "Accept": "application/json",
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {key_value}",
                },
                {
                    "model": perplexity_sonar_model_name(model),
                    "messages": [
                        {"role": "system", "content": "Return only valid JSON. No markdown."},
                        {"role": "user", "content": prompt},
                    ],
                },
                timeout,
            )
            choices = data.get("choices") or []
            if not choices:
                return ""
            return choices[0].get("message", {}).get("content", "")
        data = post_json_request(
            f"{root}/v1/agent",
            {
                "Accept": "application/json",
                "Content-Type": "application/json",
                "Authorization": f"Bearer {key_value}",
            },
            {
                "model": model,
                "instructions": "Return only valid JSON. No markdown.",
                "input": prompt,
                "max_output_tokens": 5000,
            },
            timeout,
        )
        return extract_perplexity_agent_text(data)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {key_value}",
    }
    data = post_json_request(
        f"{root}/v1/chat/completions",
        headers,
        {
            "model": model,
            "messages": [
                {"role": "system", "content": "Return only valid JSON. No markdown."},
                {"role": "user", "content": prompt},
            ],
        },
        timeout,
    )
    choices = data.get("choices") or []
    if not choices:
        return ""
    return choices[0].get("message", {}).get("content", "")


def create_entity_lsi_run_shell(
    project_id: int,
    seed_keyword: str,
    depth: Any,
    api_key_id: int,
    model: str | None = None,
    batch_id: int | None = None,
    status: str = "queued",
) -> dict[str, Any]:
    seed_keyword = clean_text(seed_keyword) or ""
    if not seed_keyword:
        raise ValueError("Seed keyword is required")
    depth_value = clamp_entity_depth(depth)
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Client not found")
        api_key = con.execute("SELECT * FROM api_keys WHERE id = ?", (api_key_id,)).fetchone()
        if not api_key:
            raise ValueError("Choose a saved LLM API key")
        provider_key = normalize_ai_provider(api_key["provider"])
        if provider_key not in LLM_PROVIDER_KEYS:
            raise ValueError("Choose an OpenAI, Anthropic, Google, xAI / Grok, or Perplexity key")
        main_url = con.execute(
            """
            SELECT COALESCE(
                (SELECT pg.url FROM pages pg JOIN sites s ON s.id = pg.site_id WHERE s.project_id = p.id ORDER BY pg.id LIMIT 1),
                (SELECT s.domain FROM sites s WHERE s.project_id = p.id ORDER BY s.id LIMIT 1),
                ''
            ) AS main_url
            FROM projects p
            WHERE p.id = ?
            """,
            (project_id,),
        ).fetchone()["main_url"]
        keywords = [
            str(row["keyword"])
            for row in con.execute("SELECT keyword FROM keywords WHERE project_id = ? ORDER BY id", (project_id,)).fetchall()
        ]
        resolved_model = clean_text(model) or clean_text(api_key["default_model"]) or AI_PROVIDERS[provider_key].get("default_model") or ""
        if not resolved_model:
            raise ValueError("Choose a model for this provider")
        prompt = build_entity_lsi_prompt(project, seed_keyword, depth_value, main_url, keywords)
        cur = con.execute(
            """
            INSERT INTO entity_lsi_runs
            (project_id, batch_id, seed_keyword, depth, provider, provider_key, api_key_id, model, main_url, prompt_version, prompt, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                batch_id,
                seed_keyword,
                depth_value,
                api_key["provider"],
                provider_key,
                api_key_id,
                resolved_model,
                main_url,
                ENTITY_LSI_PROMPT_VERSION,
                prompt,
                status,
                now,
                now,
            ),
        )
        run_id = cur.lastrowid
        row = con.execute(
            """
            SELECT er.*, p.name AS project_name
            FROM entity_lsi_runs er
            JOIN projects p ON p.id = er.project_id
            WHERE er.id = ?
            """,
            (run_id,),
        ).fetchone()
    return entity_lsi_run_public(row) or {}


def execute_entity_lsi_run(run_id: int, failure_warning: str = "The LLM run failed before structured results were created.") -> dict[str, Any]:
    with connect() as con:
        row = con.execute("SELECT * FROM entity_lsi_runs WHERE id = ?", (run_id,)).fetchone()
        if not row:
            raise ValueError("Entity & LSI run not found")
        if row["status"] == "complete" or row["provider_key"] == "cora":
            return entity_lsi_run_public(row) or {}
        api_key = con.execute("SELECT * FROM api_keys WHERE id = ?", (row["api_key_id"],)).fetchone()
        if not api_key:
            raise ValueError(f"API key missing for run {run_id}")
        provider_key = normalize_ai_provider(api_key["provider"])
        key_value = api_key["key_value"]
        base_url = clean_text(api_key["base_url"]) or AI_PROVIDERS[provider_key]["base_url"]
        model = row["model"] or clean_text(api_key["default_model"]) or AI_PROVIDERS[provider_key].get("default_model") or ""
        prompt = row["prompt"] or ""
        con.execute(
            """
            UPDATE entity_lsi_runs
            SET status = 'running', error = NULL, updated_at = ?
            WHERE id = ?
            """,
            (datetime.now().isoformat(timespec="seconds"), run_id),
        )
    raw_response = ""
    try:
        raw_response = call_llm_provider(provider_key, key_value, base_url, model, prompt)
        parsed = normalize_entity_lsi_result(parse_llm_json_text(raw_response))
        status = "complete"
        error = None
    except Exception as exc:
        parsed = {"warnings": [failure_warning]}
        status = "failed"
        error = sanitize_provider_message(str(exc), key_value)
    updated_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        con.execute(
            """
            UPDATE entity_lsi_runs
            SET raw_response = ?, result_json = ?, status = ?, error = ?, updated_at = ?
            WHERE id = ?
            """,
            (raw_response, json.dumps(parsed), status, error, updated_at, run_id),
        )
        row = con.execute(
            """
            SELECT er.*, p.name AS project_name
            FROM entity_lsi_runs er
            JOIN projects p ON p.id = er.project_id
            WHERE er.id = ?
            """,
            (run_id,),
        ).fetchone()
        batch_id = int(row["batch_id"]) if row["batch_id"] else None
    if batch_id:
        refresh_entity_lsi_batch_status(batch_id)
    return entity_lsi_run_public(row) or {}


def create_entity_lsi_run(
    project_id: int,
    seed_keyword: str,
    depth: Any,
    api_key_id: int,
    model: str | None = None,
    batch_id: int | None = None,
) -> dict[str, Any]:
    run = create_entity_lsi_run_shell(project_id, seed_keyword, depth, api_key_id, model, batch_id, "running")
    return execute_entity_lsi_run(int(run["id"]))


def refresh_entity_lsi_batch_status(batch_id: int) -> dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        counts = con.execute(
            """
            SELECT
                SUM(CASE WHEN status = 'complete' THEN 1 ELSE 0 END) AS complete_count,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) AS failed_count,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) AS cancelled_count,
                SUM(CASE WHEN status IN ('queued', 'running') THEN 1 ELSE 0 END) AS active_count,
                COUNT(*) AS run_count
            FROM entity_lsi_runs
            WHERE batch_id = ?
              AND COALESCE(provider_key, '') != 'cora'
            """,
            (batch_id,),
        ).fetchone()
        complete_count = int(counts["complete_count"] or 0)
        failed_count = int(counts["failed_count"] or 0)
        cancelled_count = int(counts["cancelled_count"] or 0)
        active_count = int(counts["active_count"] or 0)
        run_count = int(counts["run_count"] or 0)
        if active_count:
            status = "running"
        elif failed_count == 0 and run_count:
            status = "complete" if cancelled_count == 0 else ("partial" if complete_count else "cancelled")
        elif complete_count == 0:
            status = "cancelled" if cancelled_count and not failed_count else "failed"
        else:
            status = "partial"
        con.execute(
            """
            UPDATE entity_lsi_batches
            SET status = ?, complete_count = ?, failed_count = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, complete_count, failed_count, now, batch_id),
        )
        row = con.execute("SELECT * FROM entity_lsi_batches WHERE id = ?", (batch_id,)).fetchone()
    if not row:
        raise ValueError("Entity & LSI batch not found")
    return entity_lsi_batch_public(row) or {}


def retry_entity_lsi_run(run_id: int) -> dict[str, Any]:
    with connect() as con:
        row = con.execute("SELECT * FROM entity_lsi_runs WHERE id = ?", (run_id,)).fetchone()
        if not row:
            raise ValueError("Entity & LSI run not found")
        if row["status"] != "failed":
            return entity_lsi_run_public(row) or {}
    return execute_entity_lsi_run(run_id, "The LLM retry failed before structured results were created.")


def retry_failed_entity_lsi_batch(batch_id: int) -> dict[str, Any]:
    with connect() as con:
        batch = con.execute("SELECT * FROM entity_lsi_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            raise ValueError("Entity & LSI batch not found")
        rows = con.execute(
            """
            SELECT id
            FROM entity_lsi_runs
            WHERE batch_id = ? AND status = 'failed' AND COALESCE(provider_key, '') != 'cora'
            ORDER BY id
            """,
            (batch_id,),
        ).fetchall()
    retried = [retry_entity_lsi_run(int(row["id"])) for row in rows]
    refresh_entity_lsi_batch_status(batch_id)
    data = get_entity_lsi_batch(batch_id)
    data["retried"] = retried
    return data


def cancel_remaining_entity_lsi_batch(batch_id: int) -> dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        batch = con.execute("SELECT * FROM entity_lsi_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            raise ValueError("Entity & LSI batch not found")
        cur = con.execute(
            """
            UPDATE entity_lsi_runs
            SET status = 'cancelled', error = 'Cancelled before run started', updated_at = ?
            WHERE batch_id = ?
              AND status = 'queued'
              AND COALESCE(provider_key, '') != 'cora'
            """,
            (now, batch_id),
        )
    refresh_entity_lsi_batch_status(batch_id)
    data = get_entity_lsi_batch(batch_id)
    data["cancelled_count"] = cur.rowcount
    return data


def cora_term_candidate(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"\s+", " ", text).strip()
    lower = text.lower()
    if lower in {
        "entity", "entities", "keyword", "keywords", "question", "questions",
        "term", "terms", "lsi", "lsi keyword", "lsi keywords", "url", "rank",
        "average", "max", "min", "total", "value", "score", "page",
    }:
        return None
    if len(text) < 2 or len(text) > 90:
        return None
    if re.fullmatch(r"[-+]?\d+(\.\d+)?%?", text):
        return None
    if "://" in lower or lower.startswith("www.") or "@" in text:
        return None
    if len(text.split()) > 9:
        return None
    return text


def add_unique_entity_item(items: list[dict[str, Any]], key_name: str, value: str, extra: dict[str, Any] | None = None) -> None:
    normalized = normalize_crossover_term(value)
    if not normalized:
        return
    existing = {normalize_crossover_term(item.get(key_name)) for item in items}
    if normalized in existing:
        return
    item = {key_name: value}
    if extra:
        item.update(extra)
    items.append(item)


def cora_report_terms_for_run(run_id: int) -> dict[str, Any]:
    result = normalize_entity_lsi_result({})
    with connect() as con:
        lsi_rows = con.execute(
            "SELECT keyword FROM lsi_keywords WHERE run_id = ? ORDER BY best_of_both IS NULL, best_of_both DESC, id LIMIT 300",
            (run_id,),
        ).fetchall()
        for row in lsi_rows:
            term = cora_term_candidate(row["keyword"])
            if term:
                add_unique_entity_item(result["lsi_terms"], "term", term, {"intent": "Cora LSI keyword"})

        sheet_map = {
            "Entities": ("entities", "name", {"type": "Cora entity"}),
            "Keywords": ("related_keywords", "keyword", {"intent": "Cora keyword"}),
            "LSI Keywords": ("lsi_terms", "term", {"intent": "Cora LSI keyword"}),
            "Questions": ("questions", "question", {"intent": "Cora question"}),
        }
        rows = con.execute(
            """
            SELECT sheet, row_json
            FROM workbook_rows
            WHERE run_id = ? AND sheet IN ('Entities', 'Keywords', 'LSI Keywords', 'Questions')
            ORDER BY sheet, row_index
            """,
            (run_id,),
        ).fetchall()
    for row in rows:
        target = sheet_map.get(row["sheet"])
        if not target:
            continue
        result_key, label_key, extra = target
        try:
            values = json.loads(row["row_json"])
        except Exception:
            values = []
        for value in values:
            term = cora_term_candidate(value)
            if term:
                add_unique_entity_item(result[result_key], label_key, term, extra)
                break
    result["summary"] = "Imported from a Cora XLSX report."
    result["warnings"] = []
    return result


def import_cora_report_to_entity_batch(batch_id: int, run_id: int) -> dict[str, Any]:
    with connect() as con:
        batch = con.execute("SELECT * FROM entity_lsi_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            raise ValueError("Entity & LSI batch not found")
        run = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        if not run:
            raise ValueError("Cora report run not found")
        if run["project_id"] and int(run["project_id"]) != int(batch["project_id"]):
            raise ValueError("Choose a Cora report attached to the same client")
        existing = con.execute(
            """
            SELECT id FROM entity_lsi_runs
            WHERE batch_id = ? AND provider_key = 'cora' AND model = ?
            """,
            (batch_id, run["file_name"]),
        ).fetchone()
        if existing:
            return get_entity_lsi_batch(batch_id)
    result = cora_report_terms_for_run(run_id)
    term_count = (
        len(result.get("entities") or [])
        + len(result.get("lsi_terms") or [])
        + len(result.get("related_keywords") or [])
        + len(result.get("questions") or [])
    )
    if term_count == 0:
        raise ValueError("No usable Entity, LSI, Keyword, or Question terms were found in that Cora report")
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        batch = con.execute("SELECT * FROM entity_lsi_batches WHERE id = ?", (batch_id,)).fetchone()
        run = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        con.execute(
            """
            INSERT INTO entity_lsi_runs
            (project_id, batch_id, seed_keyword, depth, provider, provider_key, api_key_id, model,
             main_url, prompt_version, prompt, raw_response, result_json, status, error, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'Cora Report', 'cora', NULL, ?, ?, ?, NULL, NULL, ?, 'complete', NULL, ?, ?)
            """,
            (
                int(batch["project_id"]),
                batch_id,
                batch["seed_keyword"],
                int(batch["depth"]),
                run["file_name"],
                run["target_url"],
                CORA_ENTITY_IMPORT_PROMPT_VERSION,
                json.dumps(result),
                now,
                now,
            ),
        )
    return get_entity_lsi_batch(batch_id)


def process_entity_lsi_batch(batch_id: int) -> dict[str, Any]:
    while True:
        with connect() as con:
            row = con.execute(
                """
                SELECT id
                FROM entity_lsi_runs
                WHERE batch_id = ? AND status = 'queued' AND COALESCE(provider_key, '') != 'cora'
                ORDER BY id
                LIMIT 1
                """,
                (batch_id,),
        ).fetchone()
        if not row:
            break
        execute_entity_lsi_run(int(row["id"]))
        with connect() as con:
            cancelled = con.execute(
                """
                SELECT COUNT(*) AS count
                FROM entity_lsi_runs
                WHERE batch_id = ? AND status = 'cancelled' AND COALESCE(provider_key, '') != 'cora'
                """,
                (batch_id,),
            ).fetchone()["count"]
            queued = con.execute(
                """
                SELECT COUNT(*) AS count
                FROM entity_lsi_runs
                WHERE batch_id = ? AND status = 'queued' AND COALESCE(provider_key, '') != 'cora'
                """,
                (batch_id,),
            ).fetchone()["count"]
        if cancelled and not queued:
            break
    return get_entity_lsi_batch(batch_id)


def start_entity_lsi_batch_worker(batch_id: int) -> None:
    with ENTITY_LSI_WORKER_LOCK:
        worker = ENTITY_LSI_WORKERS.get(batch_id)
        if worker and worker.is_alive():
            return

        def run_worker() -> None:
            try:
                process_entity_lsi_batch(batch_id)
            finally:
                with ENTITY_LSI_WORKER_LOCK:
                    ENTITY_LSI_WORKERS.pop(batch_id, None)

        thread = threading.Thread(target=run_worker, name=f"entity-lsi-batch-{batch_id}", daemon=True)
        ENTITY_LSI_WORKERS[batch_id] = thread
        thread.start()


def create_entity_lsi_runs(
    project_id: int,
    seed_keyword: str,
    depth: Any,
    targets: list[dict[str, Any]],
    run_async: bool = False,
) -> list[dict[str, Any]]:
    if not targets:
        raise ValueError("Select at least one LLM model")
    seed_keyword = clean_text(seed_keyword) or ""
    if not seed_keyword:
        raise ValueError("Seed keyword is required")
    depth_value = clamp_entity_depth(depth)
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Client not found")
        cur = con.execute(
            """
            INSERT INTO entity_lsi_batches
            (project_id, seed_keyword, depth, status, target_count, complete_count, failed_count, created_at, updated_at)
            VALUES (?, ?, ?, 'running', ?, 0, 0, ?, ?)
            """,
            (project_id, seed_keyword, depth_value, len(targets), now, now),
        )
        batch_id = cur.lastrowid
    runs: list[dict[str, Any]] = []
    for target in targets:
        api_key_id = int(target.get("api_key_id") or 0)
        if not api_key_id:
            raise ValueError("Each selected model needs a saved LLM key")
        if run_async:
            runs.append(
                create_entity_lsi_run_shell(
                    project_id,
                    seed_keyword,
                    depth_value,
                    api_key_id,
                    target.get("model"),
                    batch_id,
                    "queued",
                )
            )
        else:
            runs.append(
                create_entity_lsi_run(
                    project_id,
                    seed_keyword,
                    depth_value,
                    api_key_id,
                    target.get("model"),
                    batch_id,
                )
            )
    refresh_entity_lsi_batch_status(batch_id)
    if run_async:
        start_entity_lsi_batch_worker(batch_id)
    return runs


def list_entity_lsi_runs(project_id: int | None = None) -> list[dict[str, Any]]:
    where = "WHERE er.project_id = ?" if project_id else ""
    params: list[Any] = [project_id] if project_id else []
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT er.*, p.name AS project_name
            FROM entity_lsi_runs er
            JOIN projects p ON p.id = er.project_id
            {where}
            ORDER BY er.created_at DESC, er.id DESC
            """,
            params,
        ).fetchall()
    return [entity_lsi_run_public(row) or {} for row in rows]


def entity_lsi_batch_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return row_to_dict(row)


def list_entity_lsi_batches(project_id: int | None = None) -> list[dict[str, Any]]:
    where = "WHERE b.project_id = ?" if project_id else ""
    params: list[Any] = [project_id] if project_id else []
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT b.*, p.name AS project_name
            FROM entity_lsi_batches b
            JOIN projects p ON p.id = b.project_id
            {where}
            ORDER BY b.created_at DESC, b.id DESC
            """,
            params,
        ).fetchall()
    return [entity_lsi_batch_public(row) or {} for row in rows]


def normalize_crossover_term(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def entity_result_items(run: dict[str, Any]) -> list[dict[str, Any]]:
    result = run.get("result") or {}
    specs = [
        ("entity", "entities", "name", ["type", "relevance_score", "suggested_usage"]),
        ("lsi", "lsi_terms", "term", ["relevance_score", "intent"]),
        ("related_keyword", "related_keywords", "keyword", ["intent", "funnel_stage"]),
    ]
    items: list[dict[str, Any]] = []
    for item_type, result_key, label_key, meta_keys in specs:
        for item in result.get(result_key) or []:
            if isinstance(item, str):
                label = item
                meta: dict[str, Any] = {}
            elif isinstance(item, dict):
                label = item.get(label_key) or item.get("name") or item.get("term") or item.get("keyword") or ""
                meta = {key: item.get(key) for key in meta_keys if item.get(key) not in (None, "")}
            else:
                continue
            normalized = normalize_crossover_term(label)
            if not normalized:
                continue
            items.append(
                {
                    "type": item_type,
                    "label": str(label).strip(),
                    "normalized": normalized,
                    "meta": meta,
                    "source": {
                        "run_id": run.get("id"),
                        "provider": run.get("provider"),
                        "provider_key": run.get("provider_key"),
                        "model": run.get("model"),
                        "status": run.get("status"),
                    },
                }
            )
    return items


def build_entity_lsi_crossover(runs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: dict[tuple[str, str], dict[str, Any]] = {}
    for run in runs:
        if run.get("status") != "complete":
            continue
        for item in entity_result_items(run):
            key = (item["type"], item["normalized"])
            row = rows.setdefault(
                key,
                {
                    "type": item["type"],
                    "term": item["label"],
                    "normalized": item["normalized"],
                    "sources": [],
                    "source_count": 0,
                    "providers": [],
                    "models": [],
                    "examples": [],
                },
            )
            source_key = f"{item['source'].get('provider')}::{item['source'].get('model')}"
            if source_key not in {src.get("source_key") for src in row["sources"]}:
                source = {**item["source"], "source_key": source_key}
                row["sources"].append(source)
                row["providers"].append(item["source"].get("provider"))
                row["models"].append(item["source"].get("model"))
            if item["label"] not in row["examples"]:
                row["examples"].append(item["label"])
            row["source_count"] = len(row["sources"])
    return sorted(rows.values(), key=lambda row: (-row["source_count"], row["type"], row["term"].lower()))


def entity_lsi_batch_progress(batch: dict[str, Any], runs: list[dict[str, Any]]) -> dict[str, Any]:
    model_runs = [run for run in runs if run.get("provider_key") != "cora"]
    total = int(batch.get("target_count") or len(model_runs) or 0)
    complete = sum(1 for run in model_runs if run.get("status") == "complete")
    failed = sum(1 for run in model_runs if run.get("status") == "failed")
    cancelled = sum(1 for run in model_runs if run.get("status") == "cancelled")
    queued = sum(1 for run in model_runs if run.get("status") == "queued")
    running = [run for run in model_runs if run.get("status") == "running"]
    queued_runs = [run for run in model_runs if run.get("status") == "queued"]
    current = running[0] if running else None
    up_next = queued_runs[0] if queued_runs else None
    finished = complete + failed + cancelled
    percent = round((finished / total) * 100, 1) if total else 0
    events: list[dict[str, Any]] = []
    for run in sorted(model_runs, key=lambda item: int(item.get("id") or 0)):
        provider_model = f"{run.get('provider') or ''} / {run.get('model') or ''}".strip(" /")
        status = run.get("status") or "unknown"
        if status == "queued":
            continue
        elif status == "running":
            message = f"Running {provider_model}"
        elif status == "complete":
            message = f"Completed {provider_model}"
        elif status == "failed":
            message = f"Failed {provider_model}: {run.get('error') or 'provider error'}"
        elif status == "cancelled":
            message = f"Cancelled {provider_model}"
        else:
            message = f"{status.title()} {provider_model}"
        events.append(
            {
                "run_id": run.get("id"),
                "status": status,
                "message": message,
                "updated_at": run.get("updated_at"),
            }
        )
    return {
        "total": total,
        "complete": complete,
        "failed": failed,
        "cancelled": cancelled,
        "queued": queued,
        "running": len(running),
        "finished": finished,
        "percent": percent,
        "current_run": current,
        "up_next": up_next,
        "queued_runs": queued_runs,
        "events": events,
    }


def get_entity_lsi_batch(batch_id: int) -> dict[str, Any]:
    with connect() as con:
        batch = con.execute(
            """
            SELECT b.*, p.name AS project_name
            FROM entity_lsi_batches b
            JOIN projects p ON p.id = b.project_id
            WHERE b.id = ?
            """,
            (batch_id,),
        ).fetchone()
    if not batch:
        raise ValueError("Entity & LSI batch not found")
    runs = [run for run in list_entity_lsi_runs(int(batch["project_id"])) if int(run.get("batch_id") or 0) == batch_id]
    batch_data = entity_lsi_batch_public(batch) or {}
    return {
        "batch": batch_data,
        "runs": runs,
        "crossover": build_entity_lsi_crossover(runs),
        "progress": entity_lsi_batch_progress(batch_data, runs),
    }


def entity_set_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return row_to_dict(row)


def entity_set_term_public(row: sqlite3.Row | None) -> dict[str, Any] | None:
    item = row_to_dict(row)
    if not item:
        return None
    try:
        item["sources"] = json.loads(item.pop("sources_json") or "[]")
    except json.JSONDecodeError:
        item["sources"] = []
    return item


def get_entity_set(set_id: int) -> dict[str, Any]:
    with connect() as con:
        row = con.execute(
            """
            SELECT es.*, p.name AS project_name, b.seed_keyword AS source_seed_keyword
            FROM entity_sets es
            JOIN projects p ON p.id = es.project_id
            LEFT JOIN entity_lsi_batches b ON b.id = es.source_batch_id
            WHERE es.id = ?
            """,
            (set_id,),
        ).fetchone()
        if not row:
            raise ValueError("Entity set not found")
        terms = con.execute(
            """
            SELECT *
            FROM entity_set_terms
            WHERE set_id = ?
            ORDER BY type, term COLLATE NOCASE
            """,
            (set_id,),
        ).fetchall()
    return {"set": entity_set_public(row) or {}, "terms": [entity_set_term_public(term) or {} for term in terms]}


def list_entity_sets(project_id: int | None = None) -> list[dict[str, Any]]:
    where = "WHERE es.project_id = ?" if project_id else ""
    params: list[Any] = [project_id] if project_id else []
    with connect() as con:
        rows = con.execute(
            f"""
            SELECT es.*, p.name AS project_name, b.seed_keyword AS source_seed_keyword,
                   COUNT(est.id) AS term_count
            FROM entity_sets es
            JOIN projects p ON p.id = es.project_id
            LEFT JOIN entity_lsi_batches b ON b.id = es.source_batch_id
            LEFT JOIN entity_set_terms est ON est.set_id = es.id
            {where}
            GROUP BY es.id
            ORDER BY es.updated_at DESC, es.id DESC
            """,
            params,
        ).fetchall()
    return [entity_set_public(row) or {} for row in rows]


def create_entity_set(
    project_id: int,
    name: str,
    terms: list[dict[str, Any]],
    source_batch_id: int | None = None,
    notes: str | None = None,
) -> dict[str, Any]:
    set_name = clean_text(name) or ""
    if not set_name:
        raise ValueError("Entity set name is required")
    if not terms:
        raise ValueError("Select at least one term to save")
    now = datetime.now().isoformat(timespec="seconds")
    allowed_types = {"entity", "lsi", "related_keyword", "question", "topic_cluster"}
    normalized_terms: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for item in terms:
        term = clean_text(item.get("term"))
        item_type = clean_text(item.get("type")) or "entity"
        normalized = clean_text(item.get("normalized")) or normalize_crossover_term(term)
        if not term or not normalized or item_type not in allowed_types:
            continue
        key = (item_type, normalized)
        if key in seen:
            continue
        seen.add(key)
        normalized_terms.append(
            {
                "term": term,
                "type": item_type,
                "normalized": normalized,
                "source_count": int(item.get("source_count") or 0),
                "sources": item.get("sources") if isinstance(item.get("sources"), list) else [],
                "notes": clean_text(item.get("notes")),
            }
        )
    if not normalized_terms:
        raise ValueError("No valid entity terms were selected")
    with connect() as con:
        project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Client not found")
        if source_batch_id:
            batch = con.execute(
                "SELECT id FROM entity_lsi_batches WHERE id = ? AND project_id = ?",
                (source_batch_id, project_id),
            ).fetchone()
            if not batch:
                raise ValueError("Source batch not found for this client")
        cur = con.execute(
            """
            INSERT INTO entity_sets (project_id, source_batch_id, name, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (project_id, source_batch_id, set_name, clean_text(notes), now, now),
        )
        set_id = int(cur.lastrowid)
        for item in normalized_terms:
            con.execute(
                """
                INSERT OR IGNORE INTO entity_set_terms
                (set_id, term, normalized, type, source_count, sources_json, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    set_id,
                    item["term"],
                    item["normalized"],
                    item["type"],
                    item["source_count"],
                    json.dumps(item["sources"]),
                    item["notes"],
                    now,
                ),
            )
    return get_entity_set(set_id)


def delete_entity_set(set_id: int) -> None:
    with connect() as con:
        cur = con.execute("DELETE FROM entity_sets WHERE id = ?", (set_id,))
        if cur.rowcount == 0:
            raise ValueError("Entity set not found")


def get_entity_lsi_run(run_id: int) -> dict[str, Any]:
    with connect() as con:
        row = con.execute(
            """
            SELECT er.*, p.name AS project_name
            FROM entity_lsi_runs er
            JOIN projects p ON p.id = er.project_id
            WHERE er.id = ?
            """,
            (run_id,),
        ).fetchone()
    if not row:
        raise ValueError("Entity & LSI run not found")
    return entity_lsi_run_public(row) or {}


def delete_entity_lsi_run(run_id: int) -> None:
    with connect() as con:
        cur = con.execute("DELETE FROM entity_lsi_runs WHERE id = ?", (run_id,))
    if cur.rowcount == 0:
        raise ValueError("Entity & LSI run not found")


def create_content_plan(
    project_id: int,
    title: str,
    site_id: int | None = None,
    page_id: int | None = None,
    keyword_id: int | None = None,
    content_type: str | None = None,
    intent: str | None = None,
    priority: str | None = None,
    status: str | None = None,
    due_date: str | None = None,
    notes: str | None = None,
) -> dict[str, Any]:
    title = title.strip()
    if not title:
        raise ValueError("Content plan title is required")
    now = datetime.now().isoformat(timespec="seconds")
    plan_status = clean_text(status) or "planned"
    with connect() as con:
        project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Project not found")
        cur = con.execute(
            """
            INSERT INTO content_plans
            (project_id, site_id, page_id, keyword_id, title, content_type, intent, priority, status, due_date, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                site_id,
                page_id,
                keyword_id,
                title,
                clean_text(content_type),
                clean_text(intent),
                clean_text(priority),
                plan_status,
                clean_text(due_date),
                clean_text(notes),
                now,
                now,
            ),
        )
        row = con.execute("SELECT * FROM content_plans WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row) or {}


def numeric_delta(base: Any, compare: Any) -> float | None:
    left = as_float(base)
    right = as_float(compare)
    if left is None or right is None:
        return None
    return right - left


def target_rank_for_run(con: sqlite3.Connection, run: sqlite3.Row) -> float | None:
    domain = clean_text(run["target_domain"])
    url = clean_text(run["target_url"])
    if not domain and url:
        domain = domain_from_url(url)
    if not domain:
        return None
    like_domain = f"%{domain}%"
    row = con.execute(
        """
        SELECT COALESCE(avg_rank, rank) AS rank_value
        FROM serp_results
        WHERE run_id = ?
          AND (host LIKE ? OR url LIKE ?)
        ORDER BY COALESCE(avg_rank, rank)
        LIMIT 1
        """,
        (run["id"], like_domain, like_domain),
    ).fetchone()
    return as_float(row["rank_value"]) if row else None


def compare_runs(base_id: int, compare_id: int) -> dict[str, Any]:
    if base_id == compare_id:
        raise ValueError("Choose two different runs to compare")
    with connect() as con:
        base = con.execute("SELECT * FROM runs WHERE id = ?", (base_id,)).fetchone()
        compare = con.execute("SELECT * FROM runs WHERE id = ?", (compare_id,)).fetchone()
        if not base or not compare:
            missing = []
            if not base:
                missing.append(str(base_id))
            if not compare:
                missing.append(str(compare_id))
            raise ValueError(
                "One of the selected comparison runs is not in the database. "
                f"Missing run id(s): {', '.join(missing)}. Refresh the run list and choose two imported reports."
            )

        base_summary = row_to_dict(base) or {}
        compare_summary = row_to_dict(compare) or {}
        base_target_rank = target_rank_for_run(con, base)
        compare_target_rank = target_rank_for_run(con, compare)
        counts = {}
        for name, table in {
            "serp_results": "serp_results",
            "recommendations": "recommendations",
            "lsi_keywords": "lsi_keywords",
            "workbook_rows": "workbook_rows",
        }.items():
            base_count = con.execute(f"SELECT COUNT(*) FROM {table} WHERE run_id = ?", (base_id,)).fetchone()[0]
            compare_count = con.execute(f"SELECT COUNT(*) FROM {table} WHERE run_id = ?", (compare_id,)).fetchone()[0]
            counts[name] = {"base": base_count, "compare": compare_count, "delta": compare_count - base_count}

        base_serp = {
            (r["url"] or r["host"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM serp_results WHERE run_id = ? AND (url IS NOT NULL OR host IS NOT NULL)",
                (base_id,),
            ).fetchall()
        }
        compare_serp = {
            (r["url"] or r["host"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM serp_results WHERE run_id = ? AND (url IS NOT NULL OR host IS NOT NULL)",
                (compare_id,),
            ).fetchall()
        }
        serp_changes = []
        for key in sorted(set(base_serp) | set(compare_serp)):
            before = base_serp.get(key)
            after = compare_serp.get(key)
            base_rank = as_float((before or {}).get("avg_rank") or (before or {}).get("rank"))
            compare_rank = as_float((after or {}).get("avg_rank") or (after or {}).get("rank"))
            if before and after:
                status = "changed" if base_rank != compare_rank else "same"
            else:
                status = "new" if after else "removed"
            if status == "same":
                continue
            serp_changes.append(
                {
                    "status": status,
                    "host": (after or before or {}).get("host"),
                    "title": (after or before or {}).get("title"),
                    "url": (after or before or {}).get("url"),
                    "base_rank": base_rank,
                    "compare_rank": compare_rank,
                    "rank_delta": numeric_delta(base_rank, compare_rank),
                }
            )
        serp_changes.sort(key=lambda r: (abs(r["rank_delta"] or 999), r["status"]), reverse=True)

        base_recs = {
            (r["sheet"], r["factor_id"] or "", r["factor"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM recommendations WHERE run_id = ? AND recommendation IS NOT NULL",
                (base_id,),
            ).fetchall()
        }
        compare_recs = {
            (r["sheet"], r["factor_id"] or "", r["factor"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM recommendations WHERE run_id = ? AND recommendation IS NOT NULL",
                (compare_id,),
            ).fetchall()
        }
        recommendation_changes = []
        for key in sorted(set(base_recs) | set(compare_recs)):
            before = base_recs.get(key)
            after = compare_recs.get(key)
            if before and after:
                changed = (
                    before.get("current_value") != after.get("current_value")
                    or before.get("goal") != after.get("goal")
                    or before.get("recommendation") != after.get("recommendation")
                    or before.get("percent") != after.get("percent")
                )
                if not changed:
                    continue
                status = "changed"
            else:
                status = "new" if after else "resolved"
            source = after or before or {}
            recommendation_changes.append(
                {
                    "status": status,
                    "sheet": source.get("sheet"),
                    "factor_id": source.get("factor_id"),
                    "factor": source.get("factor"),
                    "base_current": (before or {}).get("current_value"),
                    "compare_current": (after or {}).get("current_value"),
                    "base_goal": (before or {}).get("goal"),
                    "compare_goal": (after or {}).get("goal"),
                    "base_percent": (before or {}).get("percent"),
                    "compare_percent": (after or {}).get("percent"),
                    "recommendation": source.get("recommendation"),
                }
            )

        base_lsi = {
            r["keyword"]: row_to_dict(r)
            for r in con.execute("SELECT * FROM lsi_keywords WHERE run_id = ?", (base_id,)).fetchall()
        }
        compare_lsi = {
            r["keyword"]: row_to_dict(r)
            for r in con.execute("SELECT * FROM lsi_keywords WHERE run_id = ?", (compare_id,)).fetchall()
        }
        lsi_changes = []
        for keyword in sorted(set(base_lsi) | set(compare_lsi)):
            before = base_lsi.get(keyword)
            after = compare_lsi.get(keyword)
            if before and after:
                deficit_delta = numeric_delta(before.get("deficit"), after.get("deficit"))
                tracked_delta = numeric_delta(before.get("tracked_value"), after.get("tracked_value"))
                if deficit_delta == 0 and tracked_delta == 0:
                    continue
                status = "changed"
            else:
                deficit_delta = None
                tracked_delta = None
                status = "new" if after else "removed"
            lsi_changes.append(
                {
                    "status": status,
                    "keyword": keyword,
                    "base_tracked": (before or {}).get("tracked_value"),
                    "compare_tracked": (after or {}).get("tracked_value"),
                    "tracked_delta": tracked_delta,
                    "base_deficit": (before or {}).get("deficit"),
                    "compare_deficit": (after or {}).get("deficit"),
                    "deficit_delta": deficit_delta,
                }
            )
        lsi_changes.sort(key=lambda r: abs(r["deficit_delta"] or r["tracked_delta"] or 0), reverse=True)

    return {
        "base_run": base_summary,
        "compare_run": compare_summary,
        "summary": {
            "target_rank": {
                "base": base_target_rank,
                "compare": compare_target_rank,
                "delta": numeric_delta(base_target_rank, compare_target_rank),
            },
            "counts": counts,
            "serp_change_count": len(serp_changes),
            "recommendation_change_count": len(recommendation_changes),
            "lsi_change_count": len(lsi_changes),
        },
        "serp_changes": serp_changes[:200],
        "recommendation_changes": recommendation_changes[:300],
        "lsi_changes": lsi_changes[:300],
    }


def report_level_limits(level: str) -> dict[str, int]:
    if level == "basic":
        return {"results": 10, "recommendations": 5, "lsi": 8}
    if level == "comprehensive":
        return {"results": 50, "recommendations": 50, "lsi": 50}
    return {"results": 20, "recommendations": 15, "lsi": 20}


COMPREHENSIVE_SHEET_ORDER = [
    "Overview",
    "Roadmap",
    "Results",
    "Basic Tunings",
    "Intermediate Tunings",
    "Off Page Tunings",
    "LSI Keywords",
    "Entities",
    "Keywords",
    "Questions",
    "Sentences",
    "Schema",
    "Outbound Links",
    "Outbound Links 2",
    "WP Themes",
    "Variations",
    "Witcher View",
    "Shared Data",
]


def workbook_sheet_sections(con: sqlite3.Connection, run_id: int, row_limit: int = 150, col_limit: int = 20) -> dict[str, Any]:
    counts = [
        row_to_dict(r)
        for r in con.execute(
            """
            SELECT sheet, COUNT(*) AS row_count
            FROM workbook_rows
            WHERE run_id = ?
            GROUP BY sheet
            ORDER BY sheet COLLATE NOCASE
            """,
            (run_id,),
        ).fetchall()
    ]
    available = {r["sheet"]: int(r["row_count"]) for r in counts if r}
    ordered = [sheet for sheet in COMPREHENSIVE_SHEET_ORDER if sheet in available]
    sections: list[dict[str, Any]] = []
    for sheet in ordered:
        rows = []
        truncated_columns = False
        for row in con.execute(
            """
            SELECT row_index, row_json
            FROM workbook_rows
            WHERE run_id = ? AND sheet = ?
            ORDER BY row_index
            LIMIT ?
            """,
            (run_id, sheet, row_limit),
        ).fetchall():
            try:
                values = json.loads(row["row_json"])
            except Exception:
                values = []
            if len(values) > col_limit:
                truncated_columns = True
            values = values[:col_limit]
            rows.append({"row_index": row["row_index"], "values": values})
        sections.append(
            {
                "sheet": sheet,
                "row_count": available[sheet],
                "rows": rows,
                "truncated_rows": max(0, available[sheet] - len(rows)),
                "truncated_columns": truncated_columns,
            }
        )
    return {"counts": counts, "sections": sections}


def report_payload(
    run_id: int,
    level: str = "medium",
    ranking_snapshot_id: int | None = None,
    optimization_target_ids: list[int] | None = None,
    entity_set_id: int | None = None,
) -> dict[str, Any]:
    level = (level or "medium").strip().lower()
    if level not in {"basic", "medium", "comprehensive"}:
        level = "medium"
    limits = report_level_limits(level)
    with connect() as con:
        run = row_to_dict(
            con.execute(
                """
                SELECT r.*,
                       p.name AS project_name,
                       k.keyword AS assigned_keyword
                FROM runs r
                LEFT JOIN projects p ON p.id = r.project_id
                LEFT JOIN keywords k ON k.id = r.keyword_id
                WHERE r.id = ?
                """,
                (run_id,),
            ).fetchone()
        )
        if not run:
            raise ValueError("Run not found")
        results = [
            row_to_dict(r)
            for r in con.execute(
                """
                SELECT * FROM serp_results
                WHERE run_id = ?
                ORDER BY CASE WHEN rank IS NULL THEN 2 ELSE 1 END, rank
                LIMIT ?
                """,
                (run_id, limits["results"]),
            ).fetchall()
        ]
        recommendations = [
            row_to_dict(r)
            for r in con.execute(
                """
                SELECT * FROM recommendations
                WHERE run_id = ? AND recommendation IS NOT NULL
                ORDER BY CASE WHEN percent IS NULL THEN 2 ELSE 1 END, percent ASC
                LIMIT ?
                """,
                (run_id, limits["recommendations"]),
            ).fetchall()
        ]
        lsi = [
            row_to_dict(r)
            for r in con.execute(
                """
                SELECT * FROM lsi_keywords
                WHERE run_id = ?
                ORDER BY ABS(COALESCE(deficit, 0)) DESC, ABS(COALESCE(best_of_both, 0)) DESC
                LIMIT ?
                """,
                (run_id, limits["lsi"]),
            ).fetchall()
        ]
        counts = {
            "results": con.execute("SELECT COUNT(*) FROM serp_results WHERE run_id = ?", (run_id,)).fetchone()[0],
            "recommendations": con.execute("SELECT COUNT(*) FROM recommendations WHERE run_id = ?", (run_id,)).fetchone()[0],
            "lsi": con.execute("SELECT COUNT(*) FROM lsi_keywords WHERE run_id = ?", (run_id,)).fetchone()[0],
        }
        workbook = workbook_sheet_sections(con, run_id) if level == "comprehensive" else {"counts": [], "sections": []}

    with connect() as con:
        all_results = [
            row_to_dict(r)
            for r in con.execute(
                """
                SELECT * FROM serp_results
                WHERE run_id = ?
                ORDER BY CASE WHEN rank IS NULL THEN 2 ELSE 1 END, rank
                """,
                (run_id,),
            ).fetchall()
        ]
    target_matches = target_url_matches(run, all_results)
    target_result = target_matches[0] if target_matches else None
    run_project_id = int(run["project_id"]) if run.get("project_id") else None
    ranking_snapshot = get_ranking_snapshot(ranking_snapshot_id)["snapshot"] if ranking_snapshot_id else None
    if ranking_snapshot and run_project_id and ranking_snapshot.get("project_id") and int(ranking_snapshot["project_id"]) != run_project_id:
        raise ValueError("Ranking Snapshot must belong to the same client as the Cora run")
    optimization_targets = list_ranking_optimization_targets(target_ids=optimization_target_ids or []) if optimization_target_ids else []
    for target_item in optimization_targets:
        if run_project_id and target_item.get("projectId") and int(target_item["projectId"]) != run_project_id:
            raise ValueError("Optimization Targets must belong to the same client as the Cora run")
        if ranking_snapshot_id and target_item.get("snapshotId") and int(target_item["snapshotId"]) != int(ranking_snapshot_id):
            raise ValueError("Optimization Targets must belong to the attached Ranking Snapshot")
    if optimization_target_ids and len(optimization_targets) != len(set(int(value) for value in optimization_target_ids if value)):
        raise ValueError("One or more Optimization Targets were not found")
    entity_set = get_entity_set(entity_set_id) if entity_set_id else None
    entity_set_project_id = ((entity_set or {}).get("set") or {}).get("project_id") if entity_set else None
    if entity_set and run_project_id and entity_set_project_id and int(entity_set_project_id) != run_project_id:
        raise ValueError("Entity Set must belong to the same client as the Cora run")
    return {
        "level": level,
        "run": run,
        "results": results,
        "recommendations": recommendations,
        "lsi": lsi,
        "counts": counts,
        "target_result": target_result,
        "target_matches": target_matches,
        "ranking_snapshot": ranking_snapshot,
        "optimization_targets": optimization_targets,
        "entity_set": entity_set,
        "workbook": workbook,
    }


def create_share_report(
    run_id: int,
    level: str = "medium",
    title: str | None = None,
    notes: str | None = None,
    ranking_snapshot_id: int | None = None,
    optimization_target_ids: list[int] | None = None,
    entity_set_id: int | None = None,
) -> dict[str, Any]:
    payload = report_payload(run_id, level, ranking_snapshot_id, optimization_target_ids, entity_set_id)
    token = secrets.token_urlsafe(24)
    now = datetime.now().isoformat(timespec="seconds")
    title = clean_text(title) or f"{payload['run']['keyword']} On Page SEO Report"
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO share_reports
            (token, run_id, level, title, notes, ranking_snapshot_id, entity_set_id, optimization_target_ids_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                token,
                run_id,
                payload["level"],
                title,
                clean_text(notes),
                ranking_snapshot_id,
                entity_set_id,
                json.dumps(optimization_target_ids or []),
                now,
            ),
        )
        report = row_to_dict(con.execute("SELECT * FROM share_reports WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}
    report["url"] = f"/share/report/{token}"
    return report


def shared_report_by_token(token: str) -> dict[str, Any]:
    with connect() as con:
        report = row_to_dict(
            con.execute("SELECT * FROM share_reports WHERE token = ? AND revoked_at IS NULL", (token,)).fetchone()
        )
    if not report:
        raise ValueError("Shared report not found")
    try:
        optimization_target_ids = json.loads(report.get("optimization_target_ids_json") or "[]")
    except json.JSONDecodeError:
        optimization_target_ids = []
    payload = report_payload(
        int(report["run_id"]),
        report["level"],
        int(report["ranking_snapshot_id"]) if report.get("ranking_snapshot_id") else None,
        [int(value) for value in optimization_target_ids if value],
        int(report["entity_set_id"]) if report.get("entity_set_id") else None,
    )
    payload["share"] = report
    return payload


def fmt_report_num(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if num.is_integer():
        return str(int(num))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def render_workbook_sheet_table(section: dict[str, Any]) -> str:
    rows = section.get("rows") or []
    if not rows:
        return "<p>No rows imported for this sheet.</p>"
    header = rows[0]["values"]
    body_rows = rows[1:] if len(rows) > 1 else []
    if not header or all(value in (None, "") for value in header):
        header = [f"Column {i + 1}" for i in range(max((len(r['values']) for r in rows), default=0))]
        body_rows = rows
    header_html = "<tr><th>#</th>" + "".join(f"<th>{html_escape(str(value or ''))}</th>" for value in header) + "</tr>"
    row_html = []
    for row in body_rows:
        values = row["values"]
        cells = "".join(f"<td>{html_escape(str(value or ''))}</td>" for value in values)
        row_html.append(f"<tr><td class=\"row-num\">{html_escape(str(row['row_index']))}</td>{cells}</tr>")
    trunc = []
    if section.get("truncated_rows"):
        trunc.append(f"{section['truncated_rows']} additional rows")
    if section.get("truncated_columns"):
        trunc.append("additional columns")
    note = f"<p class=\"sheet-note\">Showing first {len(rows)} rows; {html_escape(', '.join(trunc))} available in the source XLSX.</p>" if trunc else ""
    return f"""
      {note}
      <div class="sheet-scroll">
        <table class="sheet-table"><thead>{header_html}</thead><tbody>{''.join(row_html)}</tbody></table>
      </div>
    """


def render_comprehensive_workbook_html(data: dict[str, Any]) -> str:
    if data.get("level") != "comprehensive":
        return ""
    workbook = data.get("workbook") or {}
    counts = workbook.get("counts") or []
    sections = workbook.get("sections") or []
    if not counts:
        return """
        <section class="card">
          <h2>Workbook Appendix</h2>
          <p>No workbook sheet rows were imported for this run. Download the source XLSX for the full Cora report.</p>
        </section>
        """
    rendered_names = {str(section.get("sheet") or "") for section in sections}
    sheet_index = "".join(
        f"<a href=\"#sheet-{re.sub(r'[^a-z0-9]+', '-', str(row.get('sheet', '')).lower()).strip('-')}\">{html_escape(row.get('sheet') or '')} <span>{html_escape(str(row.get('row_count') or 0))}</span></a>"
        for row in counts
        if str(row.get("sheet") or "") in rendered_names
    )
    other_sheets = [
        f"{row.get('sheet')} ({row.get('row_count') or 0})"
        for row in counts
        if str(row.get("sheet") or "") not in rendered_names
    ]
    other_sheet_html = (
        f"<p class=\"sheet-note\">Additional workbook tabs in the source XLSX: {html_escape(', '.join(other_sheets))}.</p>"
        if other_sheets else ""
    )
    section_html = []
    for index, section in enumerate(sections):
        anchor = re.sub(r"[^a-z0-9]+", "-", str(section["sheet"]).lower()).strip("-")
        open_attr = " open" if index < 2 else ""
        section_html.append(
            f"""
            <details class="sheet-card" id="sheet-{html_escape(anchor)}"{open_attr}>
              <summary>
                <strong>{html_escape(section['sheet'])}</strong>
                <span>{html_escape(str(section['row_count']))} rows</span>
              </summary>
              {render_workbook_sheet_table(section)}
            </details>
            """
        )
    return f"""
      <section class="card">
        <h2>Workbook Appendix</h2>
        <p>This comprehensive report includes key imported Cora workbook sheets in a browser-friendly format. Wide or very large sheets are previewed here; the source XLSX contains the full report.</p>
        <div class="sheet-index">{sheet_index}</div>
        {other_sheet_html}
      </section>
      <section class="workbook-sections">
        {''.join(section_html)}
      </section>
    """


def render_shared_report_html(data: dict[str, Any], base_url: str = "") -> bytes:
    run = data["run"]
    share = data["share"]
    level = data["level"].title()
    target = data.get("target_result")
    target_matches = data.get("target_matches") or []
    title = share.get("title") or f"{run['keyword']} On Page SEO Report"
    download_url = f"{base_url}/share/report/{share['token']}/download"
    rec_rows = "\n".join(
        f"<tr><td>{html_escape(r.get('factor') or '')}</td><td>{html_escape(r.get('recommendation') or '')}</td><td>{html_escape(fmt_report_num(r.get('percent')))}</td></tr>"
        for r in data["recommendations"]
    )
    serp_rows = "\n".join(
        f"<tr><td>{html_escape(fmt_report_num(r.get('rank')))}</td><td>{html_escape(r.get('host') or '')}</td><td>{html_escape(r.get('title') or '')}</td><td>{html_escape(r.get('url') or '')}</td></tr>"
        for r in data["results"]
    )
    target_match_rows = "\n".join(
        f"<tr><td>{html_escape(fmt_report_num(r.get('rank')))}</td><td>{html_escape((r.get('match_type') or '').replace('_', ' ').title())}</td><td>{html_escape(r.get('host') or '')}</td><td>{html_escape(r.get('url') or '')}</td></tr>"
        for r in target_matches[:12]
    )
    lsi_rows = "\n".join(
        f"<tr><td>{html_escape(r.get('keyword') or '')}</td><td>{html_escape(fmt_report_num(r.get('tracked_value')))}</td><td>{html_escape(fmt_report_num(r.get('deficit')))}</td></tr>"
        for r in data["lsi"]
    )
    optimization_target_rows = "\n".join(
        f"<tr><td>{html_escape(r.get('url') or '')}</td><td>{html_escape(r.get('keyword') or '')}</td><td>{html_escape(fmt_report_num(r.get('bestPosition')))}</td><td>{html_escape(fmt_report_num(r.get('opportunityScore')))}</td><td>{html_escape(r.get('status') or '')}</td><td>{html_escape(r.get('recommendedAction') or '')}</td></tr>"
        for r in data.get("optimization_targets", [])
    )
    entity_terms = (data.get("entity_set") or {}).get("terms") or []
    entity_rows = "\n".join(
        f"<tr><td>{html_escape(r.get('term') or '')}</td><td>{html_escape((r.get('type') or '').replace('_', ' ').title())}</td><td>{html_escape(fmt_report_num(r.get('source_count')))}</td></tr>"
        for r in entity_terms[:80]
    )
    target_rank = fmt_report_num(target.get("rank")) if target else "Not found in top imported results"
    notes = html_escape(share.get("notes") or "")
    comprehensive_html = render_comprehensive_workbook_html(data)
    optimization_html = f"""
    <section class="card">
      <h2>Optimization Targets</h2>
      <p>Saved ranking pages selected from the Ranking Snapshot workflow.</p>
      <table><thead><tr><th>URL</th><th>Focus Keyword</th><th>Best Position</th><th>Score</th><th>Status</th><th>Recommended Action</th></tr></thead><tbody>{optimization_target_rows or '<tr><td colspan="6">No saved optimization targets were attached to this report.</td></tr>'}</tbody></table>
    </section>
    """ if data.get("optimization_targets") else ""
    entity_html = f"""
    <section class="card">
      <h2>Entity Set</h2>
      <p>{html_escape(((data.get('entity_set') or {}).get('set') or {}).get('name') or 'Attached entity set')}</p>
      <table><thead><tr><th>Term</th><th>Type</th><th>Source Count</th></tr></thead><tbody>{entity_rows or '<tr><td colspan="3">No entity terms were attached.</td></tr>'}</tbody></table>
    </section>
    """ if data.get("entity_set") else ""
    body = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(title)}</title>
  <style>
    :root {{ --ink:#17202a; --muted:#657080; --line:#dfe5ec; --accent:#2563eb; --panel:#fff; --soft:#f6f8fb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Inter,Segoe UI,Arial,sans-serif; color:var(--ink); background:var(--soft); }}
    header {{ background:#0f172a; color:#fff; padding:34px 28px; }}
    main {{ max-width:1120px; margin:0 auto; padding:24px; }}
    h1,h2,h3 {{ margin:0; }}
    h1 {{ font-size:30px; }}
    header p {{ color:#cbd5e1; margin:8px 0 0; }}
    .meta {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:18px; }}
    .meta span,.pill {{ border:1px solid rgba(255,255,255,.24); border-radius:999px; padding:6px 10px; font-size:13px; }}
    .grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; margin-bottom:18px; }}
    .card {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:16px; }}
    .metric strong {{ display:block; font-size:24px; }}
    .metric label {{ color:var(--muted); font-size:12px; }}
    section {{ margin-bottom:18px; }}
    p {{ color:var(--muted); line-height:1.55; }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:9px; text-align:left; vertical-align:top; }}
    th {{ color:var(--muted); font-size:12px; text-transform:uppercase; letter-spacing:.04em; }}
    .actions {{ margin-top:18px; }}
    .button {{ display:inline-block; background:var(--accent); color:#fff; text-decoration:none; border-radius:6px; padding:10px 12px; }}
    .note {{ white-space:pre-wrap; }}
    .sheet-index {{ display:flex; flex-wrap:wrap; gap:8px; margin-top:12px; }}
    .sheet-index a {{ border:1px solid var(--line); border-radius:999px; color:var(--accent); padding:7px 10px; text-decoration:none; }}
    .sheet-index span {{ color:var(--muted); margin-left:4px; }}
    .workbook-sections {{ display:grid; gap:14px; }}
    .sheet-card {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; overflow:hidden; }}
    .sheet-card summary {{ cursor:pointer; display:flex; justify-content:space-between; gap:12px; padding:14px 16px; }}
    .sheet-card summary span {{ color:var(--muted); font-size:13px; }}
    .sheet-note {{ border-top:1px solid var(--line); margin:0; padding:10px 16px; }}
    .sheet-scroll {{ overflow:auto; max-height:620px; border-top:1px solid var(--line); }}
    .sheet-table {{ min-width:980px; font-size:12px; }}
    .sheet-table th {{ background:#f8fafc; position:sticky; top:0; z-index:1; }}
    .sheet-table td,.sheet-table th {{ max-width:320px; min-width:90px; overflow-wrap:anywhere; }}
    .sheet-table .row-num {{ color:var(--muted); font-size:11px; min-width:48px; width:48px; }}
    @media (max-width:800px) {{ .grid {{ grid-template-columns:1fr 1fr; }} main {{ padding:14px; }} }}
    @media print {{ body {{ background:#fff; }} .button {{ display:none; }} }}
  </style>
</head>
<body>
  <header>
    <h1>{html_escape(title)}</h1>
    <p>{html_escape(level)} customer report for {html_escape(run.get('keyword') or '')}</p>
    <div class="meta">
      <span>Client: {html_escape(run.get('project_name') or 'Unassigned')}</span>
      <span>Target: {html_escape(run.get('target_url') or run.get('target_domain') or '')}</span>
      <span>Imported: {html_escape(run.get('imported_at') or '')}</span>
    </div>
    <div class="actions"><a class="button" href="{html_escape(download_url)}">Download Source XLSX</a></div>
  </header>
  <main>
    <section class="grid">
      <div class="card metric"><strong>{html_escape(target_rank)}</strong><label>Target Rank</label></div>
      <div class="card metric"><strong>{data['counts']['results']}</strong><label>SERP Results</label></div>
      <div class="card metric"><strong>{data['counts']['recommendations']}</strong><label>Recommendations</label></div>
      <div class="card metric"><strong>{data['counts']['lsi']}</strong><label>LSI Terms</label></div>
    </section>
    <section class="card">
      <h2>Executive Summary</h2>
      <p>This report summarizes the strongest on-page opportunities found in the Cora analysis for the selected keyword. The recommendations are prioritized from imported Cora tuning data and grouped for customer review.</p>
      {f'<p class="note">{notes}</p>' if notes else ''}
    </section>
    <section class="card">
      <h2>Target URL Visibility</h2>
      <p>Requested target: {html_escape(run.get('target_url') or run.get('target_domain') or 'Not stored')}</p>
      <table><thead><tr><th>Rank</th><th>Match</th><th>Host</th><th>Found URL</th></tr></thead><tbody>{target_match_rows or '<tr><td colspan="4">No matching target URL or domain was found in the imported SERP results.</td></tr>'}</tbody></table>
    </section>
    <section class="card">
      <h2>Priority Action Plan</h2>
      <table><thead><tr><th>Factor</th><th>Recommendation</th><th>Gap %</th></tr></thead><tbody>{rec_rows or '<tr><td colspan="3">No recommendations imported.</td></tr>'}</tbody></table>
    </section>
    <section class="card">
      <h2>Competitor Snapshot</h2>
      <table><thead><tr><th>Rank</th><th>Host</th><th>Title</th><th>URL</th></tr></thead><tbody>{serp_rows or '<tr><td colspan="4">No SERP rows imported.</td></tr>'}</tbody></table>
    </section>
    <section class="card">
      <h2>Entity & LSI Opportunities</h2>
      <table><thead><tr><th>Term</th><th>Current</th><th>Deficit</th></tr></thead><tbody>{lsi_rows or '<tr><td colspan="3">No LSI rows imported.</td></tr>'}</tbody></table>
    </section>
    {optimization_html}
    {entity_html}
    {comprehensive_html}
  </main>
</body>
</html>"""
    return body.encode("utf-8")


class AppHandler(BaseHTTPRequestHandler):
    server_version = "CoraDashboard/0.1"

    def log_message(self, fmt: str, *args: Any) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)
        try:
            if path == "/":
                self.serve_file(STATIC_DIR / "index.html")
            elif path.startswith("/static/"):
                self.serve_file(STATIC_DIR / unquote(path.removeprefix("/static/")))
            elif re.match(r"^/share/report/[^/]+$", path):
                self.serve_shared_report(path.rsplit("/", 1)[1])
            elif re.match(r"^/share/report/[^/]+/download$", path):
                self.download_shared_report(path.split("/")[3])
            elif path == "/api/overview":
                self.api_overview(query)
            elif path == "/api/cloudflare/status":
                json_response(self, cloudflare_sync_state())
            elif path == "/api/cloudflare/artifacts":
                json_response(self, cloudflare_artifact_state())
            elif path == "/api/cloudflare/bridge":
                json_response(self, bridge_status())
            elif path == "/api/content-plans":
                self.api_content_plans(query)
            elif path == "/api/entity-lsi/runs":
                self.api_entity_lsi_runs(query)
            elif path == "/api/entity-lsi/batches":
                self.api_entity_lsi_batches(query)
            elif path == "/api/entity-sets":
                project_id = int(query["project_id"][0]) if query.get("project_id") else None
                json_response(self, {"sets": list_entity_sets(project_id)})
            elif path == "/api/seo/ranking-snapshots":
                project_id = int(query["project_id"][0]) if query.get("project_id") else None
                json_response(self, {"snapshots": list_ranking_snapshots(project_id)})
            elif path == "/api/seo/optimization-targets":
                project_id = int(query["project_id"][0]) if query.get("project_id") else None
                snapshot_id = int(query["snapshot_id"][0]) if query.get("snapshot_id") else None
                json_response(self, {"targets": list_ranking_optimization_targets(project_id=project_id, snapshot_id=snapshot_id)})
            elif path == "/api/seo/ranking-snapshots/compare":
                base_id = int((query.get("base_id") or ["0"])[0] or 0)
                compare_id = int((query.get("compare_id") or ["0"])[0] or 0)
                if not base_id or not compare_id:
                    error_response(self, "base_id and compare_id are required", 400)
                    return
                json_response(self, compare_ranking_snapshots(base_id, compare_id))
            elif re.match(r"^/api/seo/ranking-snapshots/\d+$", path):
                json_response(self, get_ranking_snapshot(int(path.rsplit("/", 1)[1])))
            elif path == "/api/runs":
                self.api_runs(query)
            elif path == "/api/share-reports":
                self.api_share_reports(query)
            elif path == "/api/compare":
                self.api_compare(query)
            elif re.match(r"^/api/runs/\d+$", path):
                self.api_run_detail(int(path.rsplit("/", 1)[1]))
            elif re.match(r"^/api/runs/\d+/download$", path):
                self.download_run(int(path.split("/")[3]))
            elif path == "/api/latest-report":
                latest = latest_xlsx()
                json_response(self, {"path": str(latest) if latest else None})
            elif path == "/api/cora/status":
                json_response(self, query_cora("/api/status"))
            elif path == "/api/cora/settings":
                json_response(self, query_cora("/api/settings"))
            elif path == "/api/cora/log":
                limit = int((query.get("lines") or ["80"])[0] or "80")
                json_response(self, cora_log_tail(limit))
            elif path == "/api/activity":
                limit = int((query.get("limit") or ["120"])[0] or "120")
                kind = (query.get("kind") or [""])[0]
                json_response(self, activity_log_tail(limit, kind))
            elif path == "/api/cora/domains":
                json_response(self, query_cora("/api/domains"))
            elif path == "/api/jobs":
                self.api_jobs(query)
            elif path == "/api/jobs/queue":
                json_response(self, queue_state())
            elif re.match(r"^/api/jobs/\d+$", path):
                self.api_job_detail(int(path.rsplit("/", 1)[1]))
            elif re.match(r"^/api/runs/\d+/workbook$", path):
                self.api_run_workbook(int(path.split("/")[3]), query)
            elif path == "/api/profiles":
                self.api_profiles()
            elif path == "/api/projects":
                self.api_projects(query)
            elif re.match(r"^/api/projects/\d+$", path):
                self.api_project_detail(int(path.rsplit("/", 1)[1]))
            elif path == "/api/api-keys":
                self.api_keys()
            elif path == "/api/ai-providers":
                json_response(self, {"providers": ai_provider_catalog()})
            elif re.match(r"^/api/entity-lsi/batches/\d+$", path):
                json_response(self, get_entity_lsi_batch(int(path.rsplit("/", 1)[1])))
            elif re.match(r"^/api/entity-lsi/runs/\d+$", path):
                json_response(self, {"run": get_entity_lsi_run(int(path.rsplit("/", 1)[1]))})
            elif re.match(r"^/api/entity-sets/\d+$", path):
                json_response(self, get_entity_set(int(path.rsplit("/", 1)[1])))
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            body = read_json_body(self)
            if parsed.path == "/api/ingest":
                report_path = body.get("path") or (str(latest_xlsx()) if latest_xlsx() else None)
                if not report_path:
                    error_response(self, "No report path provided and no .xlsx found in the user folder", 400)
                    return
                result = ingest_report(
                    Path(report_path),
                    target_url=body.get("target_url"),
                    keyword=body.get("keyword"),
                    notes=body.get("notes"),
                )
                json_response(self, result)
            elif parsed.path == "/api/jobs":
                profile_name = clean_text(body.get("new_profile_name")) or clean_text(body.get("cora_profile"))
                if clean_text(body.get("new_profile_name")):
                    create_cora_profile(clean_text(body.get("new_profile_name")) or "")
                    create_profile(clean_text(body.get("new_profile_name")) or "")
                job = create_managed_job(body.get("keyword", ""), body.get("target_url", ""), profile_name)
                json_response(self, {"job": job}, 201)
            elif parsed.path == "/api/jobs/queue":
                paused = bool(body.get("paused"))
                auto_resume = bool(body.get("auto_resume"))
                stop_after_current = bool(body.get("stop_after_current"))
                reason = clean_text(body.get("reason")) or ("Stop after current run" if stop_after_current else "")
                set_queue_paused(paused, auto_resume, stop_after_current, reason)
                if not paused:
                    resume_pending_jobs()
                json_response(self, queue_state())
            elif parsed.path == "/api/tools/run":
                result = run_client_tool(
                    int(body.get("project_id")),
                    [int(value) for value in body.get("keyword_ids", [])],
                    body.get("tool", ""),
                    body.get("cora_profile"),
                )
                json_response(self, result, 201 if result.get("jobs") else 200)
            elif parsed.path == "/api/cora/stop":
                result = force_stop_cora(body.get("reason") or "Manual dashboard force stop")
                json_response(self, result)
            elif parsed.path == "/api/cora/restart":
                result = restart_cora(body.get("reason") or "Manual dashboard restart")
                json_response(self, result, 200 if result.get("ok") else 500)
            elif parsed.path == "/api/cora/domains":
                result = post_cora("/api/domains", body)
                json_response(self, result)
            elif parsed.path == "/api/cora/settings":
                result = post_cora("/api/settings", body)
                json_response(self, result)
            elif parsed.path == "/api/backfill-workbook":
                result = backfill_workbook_rows(body.get("run_id"))
                json_response(self, result)
            elif parsed.path == "/api/cloudflare/sync":
                tables = body.get("tables") if isinstance(body.get("tables"), list) else None
                result = push_cloudflare_sync(tables=tables, dry_run=bool(body.get("dry_run")))
                json_response(self, result)
            elif parsed.path == "/api/cloudflare/artifacts/sync":
                report_ids = [int(value) for value in body.get("report_ids", []) if value] if isinstance(body.get("report_ids"), list) else None
                result = sync_cloudflare_report_artifacts(
                    report_ids=report_ids,
                    dry_run=bool(body.get("dry_run")),
                    force=bool(body.get("force")),
                )
                json_response(self, result, 200 if result.get("ok") else 502)
            elif parsed.path == "/api/cloudflare/commands/pull":
                result = pull_cloudflare_commands(int(body.get("limit") or 25))
                json_response(self, result, 200 if result.get("ok") else 502)
            elif parsed.path == "/api/cloudflare/bridge":
                result = set_bridge_settings(
                    enabled=bool(body.get("enabled")) if "enabled" in body else None,
                    allow_cora=bool(body.get("allow_cora")) if "allow_cora" in body else None,
                    allow_paid_tools=bool(body.get("allow_paid_tools")) if "allow_paid_tools" in body else None,
                    poll_interval=int(body.get("poll_interval")) if body.get("poll_interval") else None,
                )
                ensure_bridge_worker()
                if cloudflare_sync_configured():
                    try:
                        send_bridge_heartbeat(result={"settings_updated": True})
                    except Exception:
                        pass
                json_response(self, result)
            elif parsed.path == "/api/share-reports":
                report = create_share_report(
                    int(body.get("run_id")),
                    body.get("level") or "medium",
                    body.get("title"),
                    body.get("notes"),
                    int(body["ranking_snapshot_id"]) if body.get("ranking_snapshot_id") else None,
                    [int(value) for value in body.get("optimization_target_ids", []) if value] if isinstance(body.get("optimization_target_ids"), list) else [],
                    int(body["entity_set_id"]) if body.get("entity_set_id") else None,
                )
                host = self.headers.get("Host") or f"127.0.0.1:{DEFAULT_PORT}"
                report["absolute_url"] = f"http://{host}{report['url']}"
                json_response(self, {"report": report}, 201)
            elif parsed.path == "/api/projects":
                if clean_text(body.get("profile_name")):
                    create_cora_profile(clean_text(body.get("profile_name")) or "")
                project = create_project(
                    body.get("name", ""),
                    body.get("client"),
                    body.get("site_domain"),
                    body.get("notes"),
                    int(body["profile_id"]) if body.get("profile_id") else None,
                    body.get("profile_name"),
                )
                json_response(self, {"project": project}, 201)
            elif re.match(r"^/api/projects/\d+/profile$", parsed.path):
                project_id = int(parsed.path.split("/")[3])
                profile_name = clean_text(body.get("profile_name"))
                if profile_name:
                    create_cora_profile(profile_name)
                project = attach_project_profile(
                    project_id,
                    int(body["profile_id"]) if body.get("profile_id") else None,
                    profile_name,
                    bool(body.get("detach")),
                )
                json_response(self, {"project": project})
            elif parsed.path == "/api/profiles":
                create_cora_profile(body.get("name", ""))
                profile = create_profile(body.get("name", ""), body.get("client"), body.get("notes"))
                json_response(self, {"profile": profile}, 201)
            elif re.match(r"^/api/profiles/\d+$", parsed.path):
                profile = update_profile(
                    int(parsed.path.rsplit("/", 1)[1]),
                    body.get("name", ""),
                    body.get("client"),
                    body.get("notes"),
                )
                json_response(self, {"profile": profile})
            elif re.match(r"^/api/profiles/\d+/apply-cora$", parsed.path):
                profile_id = int(parsed.path.split("/")[3])
                with connect() as con:
                    profile = con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
                if not profile:
                    raise ValueError("Profile not found")
                result = post_cora("/api/settings", {"profile": profile["name"]})
                json_response(self, {"profile": row_to_dict(profile), "cora": result})
            elif re.match(r"^/api/profiles/\d+/push-cora$", parsed.path):
                profile_id = int(parsed.path.split("/")[3])
                with connect() as con:
                    profile = con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
                if not profile:
                    raise ValueError("Profile not found")
                result = create_cora_profile(profile["name"])
                json_response(self, {"profile": row_to_dict(profile), "cora": result})
            elif parsed.path == "/api/content-plans":
                plan = create_content_plan(
                    int(body.get("project_id")),
                    body.get("title", ""),
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    int(body["keyword_id"]) if body.get("keyword_id") else None,
                    body.get("content_type"),
                    body.get("intent"),
                    body.get("priority"),
                    body.get("status"),
                    body.get("due_date"),
                    body.get("notes"),
                )
                json_response(self, {"content_plan": plan}, 201)
            elif parsed.path == "/api/entity-lsi/runs":
                if isinstance(body.get("targets"), list):
                    runs = create_entity_lsi_runs(
                        int(body.get("project_id")),
                        body.get("seed_keyword", ""),
                        body.get("depth", 3),
                        body.get("targets") or [],
                        True,
                    )
                else:
                    runs = [
                        create_entity_lsi_run(
                            int(body.get("project_id")),
                            body.get("seed_keyword", ""),
                            body.get("depth", 3),
                            int(body.get("api_key_id")),
                            body.get("model"),
                        )
                    ]
                json_response(self, {"run": runs[0], "runs": runs}, 201)
            elif re.match(r"^/api/entity-lsi/batches/\d+/retry-failed$", parsed.path):
                batch_id = int(parsed.path.split("/")[4])
                json_response(self, retry_failed_entity_lsi_batch(batch_id))
            elif re.match(r"^/api/entity-lsi/batches/\d+/cancel-remaining$", parsed.path):
                batch_id = int(parsed.path.split("/")[4])
                json_response(self, cancel_remaining_entity_lsi_batch(batch_id))
            elif re.match(r"^/api/entity-lsi/batches/\d+/import-cora-report$", parsed.path):
                batch_id = int(parsed.path.split("/")[4])
                json_response(self, import_cora_report_to_entity_batch(batch_id, int(body.get("run_id"))), 201)
            elif parsed.path == "/api/entity-sets":
                data = create_entity_set(
                    int(body.get("project_id")),
                    body.get("name", ""),
                    body.get("terms") if isinstance(body.get("terms"), list) else [],
                    int(body["source_batch_id"]) if body.get("source_batch_id") else None,
                    body.get("notes"),
                )
                json_response(self, data, 201)
            elif parsed.path == "/api/seo/ranking-snapshot":
                json_response(self, create_or_get_ranking_snapshot(body), 201)
            elif parsed.path == "/api/seo/optimization-targets":
                json_response(self, save_ranking_optimization_targets(body), 201)
            elif parsed.path == "/api/seo/optimization-targets/status":
                json_response(
                    self,
                    update_ranking_optimization_target_status(
                        body.get("target_ids") or [],
                        body.get("status") or "",
                        int(body["project_id"]) if body.get("project_id") else None,
                    ),
                )
            elif parsed.path == "/api/seo/ranking-snapshot/queue-cora":
                json_response(
                    self,
                    queue_ranking_snapshot_cora_job(
                        int(body.get("project_id")),
                        body.get("keyword", ""),
                        body.get("ranking_url", ""),
                        body.get("cora_profile"),
                        bool(body.get("create_keyword", True)),
                    ),
                    201,
                )
            elif parsed.path == "/api/sites":
                site = create_site(int(body.get("project_id")), body.get("domain", ""), body.get("name"))
                json_response(self, {"site": site}, 201)
            elif parsed.path == "/api/pages":
                page = create_page(int(body.get("site_id")), body.get("url", ""), body.get("title"))
                json_response(self, {"page": page}, 201)
            elif parsed.path == "/api/keywords":
                keyword = create_keyword(
                    int(body.get("project_id")),
                    body.get("keyword", ""),
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    body.get("intent"),
                    body.get("priority"),
                )
                json_response(self, {"keyword": keyword}, 201)
            elif re.match(r"^/api/runs/\d+/assign$", parsed.path):
                run = assign_run(
                    int(parsed.path.split("/")[3]),
                    int(body["project_id"]) if body.get("project_id") else None,
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    int(body["keyword_id"]) if body.get("keyword_id") else None,
                )
                json_response(self, {"run": run})
            elif parsed.path == "/api/api-keys":
                provider = body.get("provider", "")
                key = create_api_key(
                    provider,
                    body.get("label", ""),
                    api_key_value_from_payload(provider, body),
                    body.get("notes"),
                    body.get("base_url"),
                    body.get("default_model"),
                )
                json_response(self, {"api_key": key}, 201)
            elif parsed.path == "/api/api-keys/test":
                json_response(self, {"test": test_api_key_payload(body)})
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        try:
            if re.match(r"^/api/api-keys/\d+$", parsed.path):
                delete_api_key(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"deleted": True})
            elif re.match(r"^/api/entity-lsi/runs/\d+$", parsed.path):
                delete_entity_lsi_run(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"deleted": True})
            elif re.match(r"^/api/entity-sets/\d+$", parsed.path):
                delete_entity_set(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"deleted": True})
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def serve_file(self, path: Path) -> None:
        path = path.resolve()
        if not str(path).startswith(str(STATIC_DIR.resolve())) or not path.exists() or not path.is_file():
            error_response(self, "Not found", 404)
            return
        ctype = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        if path.suffix.lower() in {".html", ".js", ".css"}:
            self.send_header("Cache-Control", "no-store, max-age=0")
            self.send_header("Pragma", "no-cache")
        self.end_headers()
        self.wfile.write(body)

    def serve_shared_report(self, token: str) -> None:
        data = shared_report_by_token(token)
        host = self.headers.get("Host") or f"127.0.0.1:{DEFAULT_PORT}"
        body = render_shared_report_html(data, f"http://{host}")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store, max-age=0")
        self.end_headers()
        self.wfile.write(body)

    def download_shared_report(self, token: str) -> None:
        data = shared_report_by_token(token)
        run = data["run"]
        path = Path(run["archive_path"])
        if not path.exists():
            error_response(self, "Archived file missing", 404)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{run["file_name"]}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_overview(self, query: dict[str, list[str]]) -> None:
        profile_id = query_profile_id(query)
        with connect() as con:
            run_filter = "WHERE p.profile_id = ?" if profile_id else ""
            run_params: list[Any] = [profile_id] if profile_id else []
            project_filter = "WHERE p.profile_id = ?" if profile_id else ""
            project_params: list[Any] = [profile_id] if profile_id else []
            profile = row_to_dict(con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()) if profile_id else None
            job_filter = ""
            job_params: list[Any] = []
            if profile_id:
                job_filter = "WHERE p.profile_id = ? OR jp.profile_id = ? OR j.cora_profile = ?"
                job_params = [profile_id, profile_id, profile.get("name", "") if profile else ""]
            counts = {
                "profiles": 1 if profile_id else con.execute("SELECT COUNT(*) FROM profiles").fetchone()[0],
                "projects": con.execute(f"SELECT COUNT(*) FROM projects p {project_filter}", project_params).fetchone()[0],
                "sites": con.execute(f"SELECT COUNT(*) FROM sites s JOIN projects p ON p.id = s.project_id {project_filter}", project_params).fetchone()[0],
                "pages": con.execute(f"SELECT COUNT(*) FROM pages pg JOIN sites s ON s.id = pg.site_id JOIN projects p ON p.id = s.project_id {project_filter}", project_params).fetchone()[0],
                "keywords": con.execute(f"SELECT COUNT(*) FROM keywords k JOIN projects p ON p.id = k.project_id {project_filter}", project_params).fetchone()[0],
                "runs": con.execute(f"SELECT COUNT(*) FROM runs r LEFT JOIN projects p ON p.id = r.project_id {run_filter}", run_params).fetchone()[0],
                "api_keys": con.execute("SELECT COUNT(*) FROM api_keys").fetchone()[0],
                "content_plans": con.execute(f"SELECT COUNT(*) FROM content_plans cp JOIN projects p ON p.id = cp.project_id {project_filter}", project_params).fetchone()[0],
                "workbook_rows": con.execute(f"SELECT COUNT(*) FROM workbook_rows wr JOIN runs r ON r.id = wr.run_id LEFT JOIN projects p ON p.id = r.project_id {run_filter}", run_params).fetchone()[0],
            }
            job_counts = [
                row_to_dict(r)
                for r in con.execute(
                    f"""
                    SELECT j.status, COUNT(*) AS count
                    FROM managed_jobs j
                    LEFT JOIN runs r ON r.id = j.imported_run_id
                    LEFT JOIN projects p ON p.id = r.project_id
                    LEFT JOIN projects jp ON jp.id = j.project_id
                    {job_filter}
                    GROUP BY j.status
                    ORDER BY j.status
                    """,
                    job_params,
                ).fetchall()
            ]
            recent_runs = [
                row_to_dict(r)
                for r in con.execute(
                    f"""
                    SELECT r.id, r.keyword, r.target_url, r.target_domain, r.imported_at, r.file_name,
                           p.name AS project_name,
                           (SELECT COUNT(*) FROM recommendations rec WHERE rec.run_id = r.id) AS recommendation_count,
                           (SELECT COUNT(*) FROM lsi_keywords lsi WHERE lsi.run_id = r.id) AS lsi_count
                    FROM runs r
                    LEFT JOIN projects p ON p.id = r.project_id
                    {run_filter}
                    ORDER BY r.imported_at DESC, r.id DESC
                    LIMIT 8
                    """,
                    run_params,
                ).fetchall()
            ]
            recent_jobs = [
                row_to_dict(r)
                for r in con.execute(
                    f"""
                    SELECT j.id, j.keyword, j.target_domain, j.status, j.status_message, j.started_at, j.completed_at, j.imported_run_id
                    FROM managed_jobs j
                    LEFT JOIN runs r ON r.id = j.imported_run_id
                    LEFT JOIN projects p ON p.id = r.project_id
                    LEFT JOIN projects jp ON jp.id = j.project_id
                    {job_filter}
                    ORDER BY j.started_at DESC, j.id DESC
                    LIMIT 8
                    """,
                    job_params,
                ).fetchall()
            ]
            api_key_providers = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT provider, COUNT(*) AS count
                    FROM api_keys
                    GROUP BY provider
                    ORDER BY provider COLLATE NOCASE
                    """
                ).fetchall()
            ]
            recent_content_plans = [
                row_to_dict(r)
                for r in con.execute(
                    f"""
                    SELECT cp.id, cp.title, cp.content_type, cp.intent, cp.priority, cp.status, cp.due_date,
                           p.name AS project_name, k.keyword
                    FROM content_plans cp
                    JOIN projects p ON p.id = cp.project_id
                    LEFT JOIN keywords k ON k.id = cp.keyword_id
                    {project_filter}
                    ORDER BY cp.updated_at DESC, cp.id DESC
                    LIMIT 8
                    """,
                    project_params,
                ).fetchall()
            ]
        json_response(
            self,
            {
                "counts": counts,
                "job_counts": job_counts,
                "recent_runs": recent_runs,
                "recent_jobs": recent_jobs,
                "api_key_providers": api_key_providers,
                "recent_content_plans": recent_content_plans,
            },
        )

    def api_runs(self, query: dict[str, list[str]]) -> None:
        search = (query.get("q") or [""])[0].strip()
        profile_id = query_profile_id(query)
        raw_project_id = (query.get("project_id") or [""])[0].strip()
        project_id = int(raw_project_id) if raw_project_id else None
        params: list[Any] = []
        clauses: list[str] = []
        if profile_id:
            clauses.append("p.profile_id = ?")
            params.append(profile_id)
        if project_id:
            clauses.append("r.project_id = ?")
            params.append(project_id)
        if search:
            clauses.append("(r.keyword LIKE ? OR r.target_domain LIKE ? OR r.file_name LIKE ? OR p.name LIKE ?)")
            like = f"%{search}%"
            params.extend([like, like, like, like])
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT r.*,
                       p.name AS project_name,
                       s.domain AS site_domain_assigned,
                       pg.url AS page_url,
                       k.keyword AS assigned_keyword,
                       (SELECT COUNT(*) FROM serp_results s WHERE s.run_id = r.id) AS result_count,
                       (SELECT COUNT(*) FROM recommendations rec WHERE rec.run_id = r.id) AS recommendation_count
                FROM runs r
                LEFT JOIN projects p ON p.id = r.project_id
                LEFT JOIN sites s ON s.id = r.site_id
                LEFT JOIN pages pg ON pg.id = r.page_id
                LEFT JOIN keywords k ON k.id = r.keyword_id
                {where}
                ORDER BY imported_at DESC, id DESC
                LIMIT 200
                """,
                params,
            ).fetchall()
        json_response(self, {"runs": [row_to_dict(r) for r in rows]})

    def api_share_reports(self, query: dict[str, list[str]]) -> None:
        project_id = int((query.get("project_id") or ["0"])[0] or 0)
        params: list[Any] = []
        where = "WHERE sr.revoked_at IS NULL"
        if project_id:
            where += " AND r.project_id = ?"
            params.append(project_id)
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT sr.*,
                       r.keyword,
                       r.target_url,
                       r.target_domain,
                       r.imported_at,
                       p.name AS project_name,
                       (SELECT COUNT(*) FROM cloud_report_artifacts cra WHERE cra.share_report_id = sr.id AND cra.status = 'synced') AS cloud_synced_artifacts,
                       (SELECT COUNT(*) FROM cloud_report_artifacts cra WHERE cra.share_report_id = sr.id) AS cloud_total_artifacts,
                       (SELECT MAX(cra.uploaded_at) FROM cloud_report_artifacts cra WHERE cra.share_report_id = sr.id AND cra.status = 'synced') AS cloud_uploaded_at,
                       (SELECT cra.cloud_url FROM cloud_report_artifacts cra WHERE cra.share_report_id = sr.id AND cra.artifact_type = 'report_html' AND cra.status = 'synced' ORDER BY cra.uploaded_at DESC LIMIT 1) AS cloud_url
                FROM share_reports sr
                JOIN runs r ON r.id = sr.run_id
                LEFT JOIN projects p ON p.id = r.project_id
                {where}
                ORDER BY sr.created_at DESC, sr.id DESC
                LIMIT 200
                """,
                params,
            ).fetchall()
        reports = []
        for row in rows:
            report = row_to_dict(row) or {}
            report["url"] = f"/share/report/{report['token']}"
            reports.append(report)
        json_response(self, {"reports": reports})

    def api_jobs(self, query: dict[str, list[str]]) -> None:
        profile_id = query_profile_id(query)
        with connect() as con:
            profile = row_to_dict(con.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()) if profile_id else None
            where = ""
            params: list[Any] = []
            if profile_id:
                where = "WHERE p.profile_id = ? OR jp.profile_id = ? OR j.cora_profile = ?"
                params = [profile_id, profile_id, profile.get("name", "") if profile else ""]
            rows = con.execute(
                f"""
                SELECT j.*, r.file_name AS imported_file_name
                FROM managed_jobs j
                LEFT JOIN runs r ON r.id = j.imported_run_id
                LEFT JOIN projects p ON p.id = r.project_id
                LEFT JOIN projects jp ON jp.id = j.project_id
                {where}
                ORDER BY j.started_at DESC, j.id DESC
                LIMIT 100
                """,
                params,
            ).fetchall()
        state = queue_state()
        jobs = [decorate_job(r) for r in rows]
        json_response(
            self,
            {
                "jobs": jobs,
                "queue_paused": state.get("paused", False),
                "queue": state,
                "summary": queue_summary(jobs, state),
            },
        )

    def api_job_detail(self, job_id: int) -> None:
        with connect() as con:
            row = con.execute(
                """
                SELECT j.*, r.file_name AS imported_file_name
                FROM managed_jobs j
                LEFT JOIN runs r ON r.id = j.imported_run_id
                WHERE j.id = ?
                """,
                (job_id,),
            ).fetchone()
        if not row:
            error_response(self, "Job not found", 404)
            return
        json_response(self, {"job": decorate_job(row)})

    def api_run_workbook(self, run_id: int, query: dict[str, list[str]]) -> None:
        sheet = (query.get("sheet") or [""])[0].strip()
        with connect() as con:
            if sheet:
                rows = con.execute(
                    """
                    SELECT sheet, row_index, column_count, row_json
                    FROM workbook_rows
                    WHERE run_id = ? AND sheet = ?
                    ORDER BY row_index
                    LIMIT 1000
                    """,
                    (run_id, sheet),
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT sheet, COUNT(*) AS row_count, MAX(column_count) AS max_columns
                    FROM workbook_rows
                    WHERE run_id = ?
                    GROUP BY sheet
                    ORDER BY sheet
                    """,
                    (run_id,),
                ).fetchall()
        json_response(self, {"rows": [row_to_dict(r) for r in rows]})

    def api_keys(self) -> None:
        with connect() as con:
            rows = con.execute(
                """
                SELECT * FROM api_keys
                ORDER BY provider COLLATE NOCASE, label COLLATE NOCASE
                """
            ).fetchall()
        json_response(self, {"api_keys": [api_key_public(r) for r in rows], "providers": ai_provider_catalog()})

    def api_entity_lsi_runs(self, query: dict[str, list[str]]) -> None:
        raw_project_id = (query.get("project_id") or [""])[0]
        project_id = int(raw_project_id) if raw_project_id else None
        json_response(self, {"runs": list_entity_lsi_runs(project_id)})

    def api_entity_lsi_batches(self, query: dict[str, list[str]]) -> None:
        raw_project_id = (query.get("project_id") or [""])[0]
        project_id = int(raw_project_id) if raw_project_id else None
        json_response(self, {"batches": list_entity_lsi_batches(project_id)})

    def api_content_plans(self, query: dict[str, list[str]]) -> None:
        profile_id = query_profile_id(query)
        where = "WHERE p.profile_id = ?" if profile_id else ""
        params: list[Any] = [profile_id] if profile_id else []
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT cp.*,
                       p.name AS project_name,
                       s.domain AS site_domain,
                       pg.url AS page_url,
                       k.keyword
                FROM content_plans cp
                JOIN projects p ON p.id = cp.project_id
                LEFT JOIN sites s ON s.id = cp.site_id
                LEFT JOIN pages pg ON pg.id = cp.page_id
                LEFT JOIN keywords k ON k.id = cp.keyword_id
                {where}
                ORDER BY
                    CASE cp.status
                        WHEN 'planned' THEN 1
                        WHEN 'drafting' THEN 2
                        WHEN 'review' THEN 3
                        WHEN 'published' THEN 4
                        ELSE 5
                    END,
                    cp.due_date IS NULL,
                    cp.due_date,
                    cp.updated_at DESC
                LIMIT 300
                """,
                params,
            ).fetchall()
        json_response(self, {"content_plans": [row_to_dict(r) for r in rows]})

    def api_compare(self, query: dict[str, list[str]]) -> None:
        base_id = int((query.get("base_id") or ["0"])[0])
        compare_id = int((query.get("compare_id") or ["0"])[0])
        if not base_id or not compare_id:
            error_response(self, "base_id and compare_id are required", 400)
            return
        json_response(self, compare_runs(base_id, compare_id))

    def api_profiles(self) -> None:
        cora = cora_profiles()
        profiles = sync_cora_profiles()
        json_response(
            self,
            {
                "profiles": profiles,
                "cora_profiles": cora.get("profiles") or [],
                "selected_cora_profile": cora.get("selected") or "",
                "cora_error": cora.get("error"),
            },
        )

    def api_projects(self, query: dict[str, list[str]]) -> None:
        profile_id = query_profile_id(query)
        where = "WHERE p.profile_id = ?" if profile_id else ""
        params: list[Any] = [profile_id] if profile_id else []
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT p.*,
                       pr.name AS profile_name,
                       pr.client AS profile_client,
                       (SELECT COUNT(*) FROM sites s WHERE s.project_id = p.id) AS site_count,
                       (SELECT COUNT(*) FROM pages pg JOIN sites s ON s.id = pg.site_id WHERE s.project_id = p.id) AS page_count,
                       (SELECT COUNT(*) FROM keywords k WHERE k.project_id = p.id) AS keyword_count,
                       (SELECT COUNT(*) FROM runs r WHERE r.project_id = p.id) AS run_count,
                       (SELECT MAX(r.imported_at) FROM runs r WHERE r.project_id = p.id) AS last_run_at
                FROM projects p
                LEFT JOIN profiles pr ON pr.id = p.profile_id
                {where}
                ORDER BY p.name COLLATE NOCASE
                """,
                params,
            ).fetchall()
        json_response(self, {"projects": [row_to_dict(r) for r in rows]})

    def api_project_detail(self, project_id: int) -> None:
        with connect() as con:
            project = row_to_dict(
                con.execute(
                    """
                    SELECT p.*, pr.name AS profile_name, pr.client AS profile_client
                    FROM projects p
                    LEFT JOIN profiles pr ON pr.id = p.profile_id
                    WHERE p.id = ?
                    """,
                    (project_id,),
                ).fetchone()
            )
            if not project:
                error_response(self, "Project not found", 404)
                return
            sites = [
                row_to_dict(r)
                for r in con.execute(
                    "SELECT * FROM sites WHERE project_id = ? ORDER BY domain COLLATE NOCASE",
                    (project_id,),
                ).fetchall()
            ]
            pages = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT pg.*, s.domain AS site_domain
                    FROM pages pg
                    JOIN sites s ON s.id = pg.site_id
                    WHERE s.project_id = ?
                    ORDER BY s.domain COLLATE NOCASE, pg.url COLLATE NOCASE
                    """,
                    (project_id,),
                ).fetchall()
            ]
            keywords = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT k.*, s.domain AS site_domain, pg.url AS page_url
                    FROM keywords k
                    LEFT JOIN sites s ON s.id = k.site_id
                    LEFT JOIN pages pg ON pg.id = k.page_id
                    WHERE k.project_id = ?
                    ORDER BY k.keyword COLLATE NOCASE
                    """,
                    (project_id,),
                ).fetchall()
            ]
            runs = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT r.*, s.domain AS site_domain_assigned, pg.url AS page_url, k.keyword AS assigned_keyword
                    FROM runs r
                    LEFT JOIN sites s ON s.id = r.site_id
                    LEFT JOIN pages pg ON pg.id = r.page_id
                    LEFT JOIN keywords k ON k.id = r.keyword_id
                    WHERE r.project_id = ?
                    ORDER BY r.imported_at DESC, r.id DESC
                    LIMIT 100
                    """,
                    (project_id,),
                ).fetchall()
            ]
        json_response(
            self,
            {"project": project, "sites": sites, "pages": pages, "keywords": keywords, "runs": runs},
        )

    def api_run_detail(self, run_id: int) -> None:
        with connect() as con:
            run = row_to_dict(
                con.execute(
                    """
                    SELECT r.*,
                           p.name AS project_name,
                           s.domain AS site_domain_assigned,
                           pg.url AS page_url,
                           k.keyword AS assigned_keyword
                    FROM runs r
                    LEFT JOIN projects p ON p.id = r.project_id
                    LEFT JOIN sites s ON s.id = r.site_id
                    LEFT JOIN pages pg ON pg.id = r.page_id
                    LEFT JOIN keywords k ON k.id = r.keyword_id
                    WHERE r.id = ?
                    """,
                    (run_id,),
                ).fetchone()
            )
            if not run:
                error_response(self, "Run not found", 404)
                return
            results = [
                row_to_dict(r)
                for r in con.execute(
                    "SELECT * FROM serp_results WHERE run_id = ? ORDER BY rank LIMIT 100", (run_id,)
                ).fetchall()
            ]
            recommendations = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT * FROM recommendations
                    WHERE run_id = ? AND recommendation IS NOT NULL
                    ORDER BY CASE WHEN percent IS NULL THEN 2 ELSE 1 END, percent ASC
                    LIMIT 200
                    """,
                    (run_id,),
                ).fetchall()
            ]
            lsi = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT * FROM lsi_keywords
                    WHERE run_id = ?
                    ORDER BY ABS(COALESCE(deficit, 0)) DESC, ABS(COALESCE(best_of_both, 0)) DESC
                    LIMIT 100
                    """,
                    (run_id,),
                ).fetchall()
            ]
        target_matches = target_url_matches(run, results)
        json_response(
            self,
            {
                "run": run,
                "results": results,
                "recommendations": recommendations,
                "lsi": lsi,
                "target_matches": target_matches,
                "target_result": target_matches[0] if target_matches else None,
            },
        )

    def download_run(self, run_id: int) -> None:
        with connect() as con:
            run = con.execute("SELECT archive_path, file_name FROM runs WHERE id = ?", (run_id,)).fetchone()
        if not run:
            error_response(self, "Run not found", 404)
            return
        path = Path(run["archive_path"])
        if not path.exists():
            error_response(self, "Archived file missing", 404)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{run["file_name"]}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def start_server(port: int = DEFAULT_PORT) -> ThreadingHTTPServer:
    init_db()
    log_activity("system", f"Dashboard server started on port {port}", "info")
    resume_pending_jobs()
    ensure_bridge_worker()
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


def main(argv: list[str]) -> int:
    init_db()
    if len(argv) > 1 and argv[1] == "ingest":
        path = Path(argv[2]) if len(argv) > 2 else latest_xlsx()
        if not path:
            print("No .xlsx file found to import.", file=sys.stderr)
            return 1
        target_url = argv[3] if len(argv) > 3 else None
        result = ingest_report(path, target_url=target_url)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "backfill-workbook":
        run_id = int(argv[2]) if len(argv) > 2 else None
        result = backfill_workbook_rows(run_id)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-status":
        print(json.dumps(cloudflare_sync_state(), indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-sync":
        dry_run = "--dry-run" in argv
        table_arg = next((arg for arg in argv[2:] if arg.startswith("--tables=")), "")
        tables = [part.strip() for part in table_arg.removeprefix("--tables=").split(",") if part.strip()] if table_arg else None
        result = push_cloudflare_sync(tables=tables, dry_run=dry_run)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-artifacts-sync":
        dry_run = "--dry-run" in argv
        force = "--force" in argv
        report_arg = next((arg for arg in argv[2:] if arg.startswith("--report-ids=")), "")
        report_ids = [int(part.strip()) for part in report_arg.removeprefix("--report-ids=").split(",") if part.strip()] if report_arg else None
        result = sync_cloudflare_report_artifacts(report_ids=report_ids, dry_run=dry_run, force=force)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-commands-pull":
        limit_arg = next((arg for arg in argv[2:] if arg.startswith("--limit=")), "")
        limit = int(limit_arg.removeprefix("--limit=")) if limit_arg else 25
        result = pull_cloudflare_commands(limit=limit)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-sync-pull":
        tables_arg = next((arg for arg in argv[2:] if arg.startswith("--tables=")), "")
        tables = [value.strip() for value in tables_arg.removeprefix("--tables=").split(",") if value.strip()] if tables_arg else None
        limit_arg = next((arg for arg in argv[2:] if arg.startswith("--limit=")), "")
        limit = int(limit_arg.removeprefix("--limit=")) if limit_arg else 5000
        result = pull_cloudflare_sync(tables=tables, limit=limit)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-bridge-enable":
        allow_cora = "--allow-cora" in argv
        allow_paid_tools = "--allow-paid-tools" in argv
        interval_arg = next((arg for arg in argv[2:] if arg.startswith("--interval=")), "")
        interval = int(interval_arg.removeprefix("--interval=")) if interval_arg else None
        result = set_bridge_settings(
            enabled=True,
            allow_cora=allow_cora,
            allow_paid_tools=allow_paid_tools,
            poll_interval=interval,
        )
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "cloudflare-bridge-disable":
        result = set_bridge_settings(enabled=False)
        print(json.dumps(result, indent=2, default=str))
        return 0

    port = int(argv[1]) if len(argv) > 1 else DEFAULT_PORT
    server = start_server(port)
    threading.Thread(target=auto_resume_loop, daemon=True).start()
    print(f"Cora dashboard running at http://127.0.0.1:{port}/")
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        server.shutdown()
        return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
