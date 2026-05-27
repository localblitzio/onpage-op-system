from __future__ import annotations

import hashlib
import json
import mimetypes
import os
import re
import shutil
import sqlite3
import sys
import threading
import time
import urllib.error
import urllib.request
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
ARCHIVE_DIR = DATA_DIR / "archive"
DB_PATH = DATA_DIR / "cora_runs.sqlite3"
STATIC_DIR = APP_DIR / "static"
DEFAULT_REPORT_DIR = Path.home()
DEFAULT_PORT = int(os.environ.get("CORA_DASHBOARD_PORT", "9191"))
CORA_API_BASE = os.environ.get("CORA_API_BASE", "http://127.0.0.1:9090")
JOB_POLL_SECONDS = 10
JOB_TIMEOUT_SECONDS = 60 * 60 * 2
CORA_FREEZE_SECONDS = int(os.environ.get("CORA_FREEZE_SECONDS", str(60 * 10)))


def ensure_dirs() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    ARCHIVE_DIR.mkdir(exist_ok=True)
    STATIC_DIR.mkdir(exist_ok=True)


class ClosingConnection(sqlite3.Connection):
    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> bool:
        result = super().__exit__(exc_type, exc_value, traceback)
        self.close()
        return result


def connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, factory=ClosingConnection)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con


def ensure_column(con: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    columns = {row["name"] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db() -> None:
    ensure_dirs()
    with connect() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                target_url TEXT,
                target_domain TEXT,
                search_engine TEXT,
                country TEXT,
                language TEXT,
                cora_version TEXT,
                report_date TEXT,
                imported_at TEXT NOT NULL,
                source_path TEXT NOT NULL,
                archive_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                sha256 TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'imported',
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS serp_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                rank REAL,
                avg_rank REAL,
                weighted_rank REAL,
                data_rank REAL,
                host TEXT,
                title TEXT,
                url TEXT,
                summary TEXT
            );

            CREATE TABLE IF NOT EXISTS recommendations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                section TEXT,
                factor_id TEXT,
                factor TEXT,
                current_value TEXT,
                goal TEXT,
                percent REAL,
                recommendation TEXT,
                shared_correlation TEXT,
                page_one_max TEXT,
                page_one_avg TEXT
            );

            CREATE TABLE IF NOT EXISTS lsi_keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                keyword TEXT NOT NULL,
                spearman REAL,
                pearson REAL,
                best_of_both REAL,
                pages REAL,
                max_value REAL,
                avg_value REAL,
                total REAL,
                tracked_value REAL,
                deficit REAL
            );

            CREATE TABLE IF NOT EXISTS sheet_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                row_json TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS workbook_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL REFERENCES runs(id) ON DELETE CASCADE,
                sheet TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                column_count INTEGER NOT NULL,
                row_json TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS managed_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                target_url TEXT NOT NULL,
                target_domain TEXT,
                cora_script TEXT NOT NULL,
                status TEXT NOT NULL,
                status_message TEXT,
                cora_running INTEGER NOT NULL DEFAULT 0,
                cora_action TEXT,
                progress REAL,
                started_at TEXT NOT NULL,
                completed_at TEXT,
                report_path TEXT,
                imported_run_id INTEGER REFERENCES runs(id) ON DELETE SET NULL,
                error TEXT
            );

            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                client TEXT,
                notes TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                domain TEXT NOT NULL,
                name TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL REFERENCES sites(id) ON DELETE CASCADE,
                url TEXT NOT NULL,
                title TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,
                page_id INTEGER REFERENCES pages(id) ON DELETE SET NULL,
                keyword TEXT NOT NULL,
                intent TEXT,
                priority TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS api_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                label TEXT NOT NULL,
                key_value TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS content_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,
                page_id INTEGER REFERENCES pages(id) ON DELETE SET NULL,
                keyword_id INTEGER REFERENCES keywords(id) ON DELETE SET NULL,
                title TEXT NOT NULL,
                content_type TEXT,
                intent TEXT,
                priority TEXT,
                status TEXT NOT NULL DEFAULT 'planned',
                due_date TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_runs_keyword ON runs(keyword);
            CREATE INDEX IF NOT EXISTS idx_runs_target_domain ON runs(target_domain);
            CREATE INDEX IF NOT EXISTS idx_serp_run_rank ON serp_results(run_id, rank);
            CREATE INDEX IF NOT EXISTS idx_recommendations_run ON recommendations(run_id);
            CREATE INDEX IF NOT EXISTS idx_lsi_run ON lsi_keywords(run_id);
            CREATE INDEX IF NOT EXISTS idx_workbook_rows_run_sheet ON workbook_rows(run_id, sheet, row_index);
            CREATE INDEX IF NOT EXISTS idx_managed_jobs_status ON managed_jobs(status);
            CREATE INDEX IF NOT EXISTS idx_sites_project ON sites(project_id);
            CREATE INDEX IF NOT EXISTS idx_pages_site ON pages(site_id);
            CREATE INDEX IF NOT EXISTS idx_keywords_project ON keywords(project_id);
            CREATE INDEX IF NOT EXISTS idx_api_keys_provider ON api_keys(provider);
            CREATE INDEX IF NOT EXISTS idx_content_plans_project ON content_plans(project_id);
            CREATE INDEX IF NOT EXISTS idx_content_plans_status ON content_plans(status);
            """
        )
        ensure_column(con, "runs", "project_id", "INTEGER")
        ensure_column(con, "runs", "site_id", "INTEGER")
        ensure_column(con, "runs", "page_id", "INTEGER")
        ensure_column(con, "runs", "keyword_id", "INTEGER")


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    if row is None:
        return None
    return {k: row[k] for k in row.keys()}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def domain_from_url(url: str | None) -> str | None:
    if not url:
        return None
    parsed = urlparse(url if "://" in url else "https://" + url)
    host = parsed.netloc or parsed.path.split("/")[0]
    return host.lower().removeprefix("www.") if host else None


def latest_xlsx(directory: Path = DEFAULT_REPORT_DIR) -> Path | None:
    files = sorted(directory.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def normalize_slug(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def find_report_for_job(keyword: str, started_at: datetime, directory: Path = DEFAULT_REPORT_DIR) -> Path | None:
    keyword_slug = normalize_slug(keyword)
    candidates: list[Path] = []
    for path in directory.glob("*.xlsx"):
        try:
            if datetime.fromtimestamp(path.stat().st_mtime) < started_at:
                continue
        except OSError:
            continue
        if keyword_slug and keyword_slug in normalize_slug(path.stem):
            candidates.append(path)
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def wait_for_stable_file(path: Path, checks: int = 3, delay: float = 2.0) -> bool:
    last_size = -1
    stable = 0
    for _ in range(checks + 8):
        try:
            size = path.stat().st_size
        except OSError:
            stable = 0
            time.sleep(delay)
            continue
        if size > 0 and size == last_size:
            stable += 1
            if stable >= checks:
                return True
        else:
            stable = 0
            last_size = size
        time.sleep(delay)
    return False


def find_label_value(ws, label: str) -> Any:
    label_lower = label.lower()
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip().lower() == label_lower:
                for next_value in row[idx + 1 :]:
                    if next_value not in (None, ""):
                        return next_value
    return None


def workbook_metadata(wb) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        meta["keyword"] = clean_text(find_label_value(ws, "Search Terms"))
        meta["cora_version"] = clean_text(find_label_value(ws, "Version"))
        report_date = find_label_value(ws, "Date")
        meta["report_date"] = str(report_date) if report_date else None
        user_agent = clean_text(find_label_value(ws, "User Agent"))
        meta["search_engine"] = "Google" if user_agent and "google" in user_agent.lower() else user_agent
    if "Roadmap" in wb.sheetnames:
        ws = wb["Roadmap"]
        meta["country"] = clean_text(find_label_value(ws, "Google Country"))
        meta["language"] = clean_text(find_label_value(ws, "Google Language"))
    return meta


def archive_report(source: Path, sha: str) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d")
    dest_dir = ARCHIVE_DIR / stamp / sha[:12]
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / source.name
    if not dest.exists():
        shutil.copy2(source, dest)
    return dest


def parse_results(wb, run_id: int, con: sqlite3.Connection) -> None:
    if "Results" not in wb.sheetnames:
        return
    ws = wb["Results"]
    headers = [clean_text(c.value) or "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h.lower(): i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        con.execute(
            """
            INSERT INTO serp_results
            (run_id, rank, avg_rank, weighted_rank, data_rank, host, title, url, summary)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                as_float(row[idx.get("rank", -1)]) if "rank" in idx else None,
                as_float(row[idx.get("avg rank", -1)]) if "avg rank" in idx else None,
                as_float(row[idx.get("weighted rank", -1)]) if "weighted rank" in idx else None,
                as_float(row[idx.get("data", -1)]) if "data" in idx else None,
                clean_text(row[idx.get("host", -1)]) if "host" in idx else None,
                clean_text(row[idx.get("link text", -1)]) if "link text" in idx else None,
                clean_text(row[idx.get("url", -1)]) if "url" in idx else None,
                clean_text(row[idx.get("summary", -1)]) if "summary" in idx else None,
            ),
        )


def parse_tunings(wb, run_id: int, con: sqlite3.Connection) -> None:
    sheets = ["Basic Tunings", "Intermediate Tunings", "Off Page Tunings"]
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        current_section = None
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            first = clean_text(values[0]) if len(values) > 0 else None
            factor_id = clean_text(values[1]) if len(values) > 1 else None
            factor = clean_text(values[2]) if len(values) > 2 else None
            if first and not factor_id and not factor:
                current_section = first
                continue
            if not factor_id or not factor or factor_id.lower() == "factor id":
                continue
            recommendation = clean_text(values[6]) if len(values) > 6 else None
            con.execute(
                """
                INSERT INTO recommendations
                (run_id, sheet, section, factor_id, factor, current_value, goal, percent,
                 recommendation, shared_correlation, page_one_max, page_one_avg)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    sheet,
                    current_section,
                    factor_id,
                    factor,
                    clean_text(values[3]) if len(values) > 3 else None,
                    clean_text(values[4]) if len(values) > 4 else None,
                    as_float(values[5]) if len(values) > 5 else None,
                    recommendation,
                    clean_text(values[8]) if len(values) > 8 else None,
                    clean_text(values[9]) if len(values) > 9 else None,
                    clean_text(values[10]) if len(values) > 10 else None,
                ),
            )


def parse_lsi(wb, run_id: int, con: sqlite3.Connection) -> None:
    if "LSI Keywords" not in wb.sheetnames:
        return
    ws = wb["LSI Keywords"]
    header_row = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_headers = [clean_text(v) or "" for v in row]
        if row_headers and row_headers[0] == "LSI Keyword":
            header_row = i
            headers = row_headers
            break
    if not header_row:
        return
    tracked_col = 8
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        kw = clean_text(row[0] if len(row) > 0 else None)
        if not kw:
            continue
        con.execute(
            """
            INSERT INTO lsi_keywords
            (run_id, keyword, spearman, pearson, best_of_both, pages, max_value,
             avg_value, total, tracked_value, deficit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                kw,
                as_float(row[1] if len(row) > 1 else None),
                as_float(row[2] if len(row) > 2 else None),
                as_float(row[3] if len(row) > 3 else None),
                as_float(row[4] if len(row) > 4 else None),
                as_float(row[5] if len(row) > 5 else None),
                as_float(row[6] if len(row) > 6 else None),
                as_float(row[7] if len(row) > 7 else None),
                as_float(row[tracked_col] if len(row) > tracked_col else None),
                as_float(row[9] if len(row) > 9 else None),
            ),
        )


def parse_sheet_rows(wb, run_id: int, con: sqlite3.Connection) -> None:
    keep_sheets = ["Roadmap", "Overview", "Witcher View", "Keywords", "Entities", "Questions", "Sentences"]
    for sheet in keep_sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not any(v is not None for v in row):
                continue
            con.execute(
                "INSERT INTO sheet_rows (run_id, sheet, row_index, row_json) VALUES (?, ?, ?, ?)",
                (run_id, sheet, i, json.dumps([str(v) if v is not None else None for v in row])),
            )


def parse_workbook_rows(wb, run_id: int, con: sqlite3.Connection) -> None:
    con.execute("DELETE FROM workbook_rows WHERE run_id = ?", (run_id,))
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [str(v) if v is not None else None for v in row]
            if not any(v is not None and v != "" for v in values):
                continue
            con.execute(
                """
                INSERT INTO workbook_rows (run_id, sheet, row_index, column_count, row_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                (run_id, ws.title, i, len(values), json.dumps(values)),
            )


def infer_target_url(con: sqlite3.Connection, run_id: int, target_url: str | None) -> str | None:
    if target_url:
        return target_url
    row = con.execute(
        "SELECT url FROM serp_results WHERE run_id = ? AND url IS NOT NULL ORDER BY rank LIMIT 1",
        (run_id,),
    ).fetchone()
    return row["url"] if row else None


def ingest_report(path: Path, target_url: str | None = None, keyword: str | None = None, notes: str | None = None) -> dict[str, Any]:
    path = path.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx Cora reports can be imported")

    sha = sha256_file(path)
    archive_path = archive_report(path, sha)
    wb = load_workbook(path, read_only=True, data_only=True)
    meta = workbook_metadata(wb)
    keyword = keyword or meta.get("keyword") or path.stem.replace("_", " ")
    imported_at = datetime.now().isoformat(timespec="seconds")

    with connect() as con:
        existing = con.execute("SELECT * FROM runs WHERE sha256 = ?", (sha,)).fetchone()
        if existing:
            return {"run": row_to_dict(existing), "created": False}
        cur = con.execute(
            """
            INSERT INTO runs
            (keyword, target_url, target_domain, search_engine, country, language, cora_version,
             report_date, imported_at, source_path, archive_path, file_name, file_size, sha256, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                keyword,
                target_url,
                domain_from_url(target_url),
                meta.get("search_engine"),
                meta.get("country"),
                meta.get("language"),
                meta.get("cora_version"),
                meta.get("report_date"),
                imported_at,
                str(path),
                str(archive_path),
                path.name,
                path.stat().st_size,
                sha,
                notes,
            ),
        )
        run_id = int(cur.lastrowid)
        parse_results(wb, run_id, con)
        if not target_url:
            target_url = infer_target_url(con, run_id, target_url)
            con.execute(
                "UPDATE runs SET target_url = ?, target_domain = ? WHERE id = ?",
                (target_url, domain_from_url(target_url), run_id),
            )
        parse_tunings(wb, run_id, con)
        parse_lsi(wb, run_id, con)
        parse_sheet_rows(wb, run_id, con)
        parse_workbook_rows(wb, run_id, con)
        run = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        return {"run": row_to_dict(run), "created": True}


def backfill_workbook_rows(run_id: int | None = None) -> dict[str, Any]:
    init_db()
    with connect() as con:
        if run_id is None:
            runs = con.execute("SELECT id, archive_path, source_path FROM runs ORDER BY id").fetchall()
        else:
            runs = con.execute("SELECT id, archive_path, source_path FROM runs WHERE id = ?", (run_id,)).fetchall()
        updated: list[dict[str, Any]] = []
        for run in runs:
            path = Path(run["archive_path"])
            if not path.exists():
                path = Path(run["source_path"])
            if not path.exists():
                updated.append({"run_id": run["id"], "status": "missing file"})
                continue
            wb = load_workbook(path, read_only=True, data_only=True)
            parse_workbook_rows(wb, int(run["id"]), con)
            count = con.execute("SELECT COUNT(*) FROM workbook_rows WHERE run_id = ?", (run["id"],)).fetchone()[0]
            updated.append({"run_id": run["id"], "status": "backfilled", "rows": count})
    return {"updated": updated}


def json_response(handler: BaseHTTPRequestHandler, data: Any, status: int = 200) -> None:
    body = json.dumps(data, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def error_response(handler: BaseHTTPRequestHandler, message: str, status: int = 400) -> None:
    json_response(handler, {"error": message}, status)


def read_json_body(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length") or 0)
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw) if raw.strip() else {}


def query_cora(path: str) -> Any:
    try:
        with urllib.request.urlopen(CORA_API_BASE + path, timeout=4) as resp:
            raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
    except urllib.error.URLError as exc:
        return {"error": str(exc)}


def post_cora(path: str, payload: dict[str, Any]) -> Any:
    try:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            CORA_API_BASE + path,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return raw
    except urllib.error.URLError as exc:
        return {"error": str(exc)}


def cora_status_signature(status: dict[str, Any]) -> tuple[Any, ...]:
    progress = as_float(status.get("progress"))
    return (
        bool(status.get("running")),
        bool(status.get("searchRunning")),
        clean_text(status.get("searchTerm")) or "",
        clean_text(status.get("action")) or "",
        round(progress, 4) if progress is not None else None,
    )


def force_stop_cora(reason: str = "Stop requested") -> dict[str, Any]:
    before = query_cora("/api/status")
    response = post_cora("/api/stop", {"reason": reason})
    ok = not (isinstance(response, dict) and response.get("error"))
    return {"ok": ok, "reason": reason, "status_before": before, "response": response}


def job_update(job_id: int, **fields: Any) -> None:
    if not fields:
        return
    cols = ", ".join(f"{key} = ?" for key in fields)
    values = list(fields.values()) + [job_id]
    with connect() as con:
        con.execute(f"UPDATE managed_jobs SET {cols} WHERE id = ?", values)


def create_managed_job(keyword: str, target_url: str) -> dict[str, Any]:
    keyword = keyword.strip()
    target_url = target_url.strip()
    if not keyword:
        raise ValueError("Keyword is required")
    if not target_url:
        raise ValueError("Target URL/domain is required")
    target_domain = domain_from_url(target_url)
    if not target_domain:
        raise ValueError("Could not determine target domain")

    status = query_cora("/api/status")
    if isinstance(status, dict) and status.get("error"):
        raise RuntimeError(f"Cora API is not reachable: {status['error']}")
    if isinstance(status, dict) and status.get("running"):
        action = clean_text(status.get("action"))
        progress = as_float(status.get("progress")) or 0.0
        if action or progress > 0:
            raise RuntimeError("Cora is already running a job")

    script = f"search {keyword}; track domain {target_domain}; click get data"
    started_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO managed_jobs
            (keyword, target_url, target_domain, cora_script, status, status_message, started_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (keyword, target_url, target_domain, script, "queued", "Queued", started_at),
        )
        job_id = int(cur.lastrowid)
        job = row_to_dict(con.execute("SELECT * FROM managed_jobs WHERE id = ?", (job_id,)).fetchone())

    thread = threading.Thread(target=run_managed_job, args=(job_id,), daemon=True)
    thread.start()
    return job or {}


def run_managed_job(job_id: int) -> None:
    with connect() as con:
        job = con.execute("SELECT * FROM managed_jobs WHERE id = ?", (job_id,)).fetchone()
    if not job:
        return

    keyword = job["keyword"]
    target_url = job["target_url"]
    script = job["cora_script"]
    started_at = datetime.fromisoformat(job["started_at"])
    deadline = time.time() + JOB_TIMEOUT_SECONDS
    job_update(job_id, status="submitting", status_message="Submitting Cora workflow")

    response = post_cora("/api/script", {"script": script})
    if isinstance(response, dict) and response.get("error"):
        job_update(
            job_id,
            status="error",
            status_message="Cora workflow submission failed",
            error=response["error"],
            completed_at=datetime.now().isoformat(timespec="seconds"),
        )
        return

    job_update(job_id, status="running", status_message="Cora workflow submitted")
    seen_report: Path | None = None
    last_signature: tuple[Any, ...] | None = None
    last_activity = time.time()
    last_candidate_state: tuple[str, int, float] | None = None

    while time.time() < deadline:
        status = query_cora("/api/status")
        if isinstance(status, dict) and not status.get("error"):
            signature = cora_status_signature(status)
            if signature != last_signature:
                last_signature = signature
                last_activity = time.time()
            job_update(
                job_id,
                cora_running=1 if status.get("running") else 0,
                cora_action=status.get("action"),
                progress=as_float(status.get("progress")),
                status_message=status.get("action") or "Waiting for Cora report",
            )
            if (status.get("running") or status.get("searchRunning")) and time.time() - last_activity >= CORA_FREEZE_SECONDS:
                minutes = max(1, round(CORA_FREEZE_SECONDS / 60))
                stop = force_stop_cora(f"Frozen dashboard managed job {job_id}")
                job_update(
                    job_id,
                    status="stopped",
                    status_message=f"Cora appeared frozen for {minutes} minutes; stop sent",
                    error=json.dumps(stop, default=str),
                    completed_at=datetime.now().isoformat(timespec="seconds"),
                )
                return

        candidate = find_report_for_job(keyword, started_at)
        if candidate:
            seen_report = candidate
            candidate_state = (str(candidate), candidate.stat().st_size, candidate.stat().st_mtime)
            if candidate_state != last_candidate_state:
                last_candidate_state = candidate_state
                last_activity = time.time()
            if wait_for_stable_file(candidate):
                try:
                    result = ingest_report(candidate, target_url=target_url, keyword=keyword, notes=f"Managed job {job_id}")
                    run = result.get("run") or {}
                    job_update(
                        job_id,
                        status="imported",
                        status_message="Report imported",
                        completed_at=datetime.now().isoformat(timespec="seconds"),
                        report_path=str(candidate),
                        imported_run_id=run.get("id"),
                        progress=1.0,
                    )
                    return
                except Exception as exc:
                    job_update(
                        job_id,
                        status="error",
                        status_message="Report import failed",
                        report_path=str(candidate),
                        error=str(exc),
                        completed_at=datetime.now().isoformat(timespec="seconds"),
                    )
                    return

        time.sleep(JOB_POLL_SECONDS)

    job_update(
        job_id,
        status="timeout",
        status_message="Timed out waiting for Cora report",
        report_path=str(seen_report) if seen_report else None,
        error="Timed out waiting for a matching .xlsx report",
        completed_at=datetime.now().isoformat(timespec="seconds"),
    )


def create_project(
    name: str,
    client: str | None = None,
    site_domain: str | None = None,
    notes: str | None = None,
) -> dict[str, Any]:
    name = name.strip()
    if not name:
        raise ValueError("Project name is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            "INSERT INTO projects (name, client, notes, created_at) VALUES (?, ?, ?, ?)",
            (name, clean_text(client), clean_text(notes), created_at),
        )
        project_id = int(cur.lastrowid)
        if clean_text(site_domain):
            domain = domain_from_url(site_domain) or site_domain.strip().lower()
            con.execute(
                "INSERT INTO sites (project_id, domain, name, created_at) VALUES (?, ?, ?, ?)",
                (project_id, domain, domain, created_at),
            )
        return row_to_dict(con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()) or {}


def create_site(project_id: int, domain: str, name: str | None = None) -> dict[str, Any]:
    domain = domain_from_url(domain) or domain.strip().lower()
    if not domain:
        raise ValueError("Domain is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            "INSERT INTO sites (project_id, domain, name, created_at) VALUES (?, ?, ?, ?)",
            (project_id, domain, clean_text(name) or domain, created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM sites WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def create_page(site_id: int, url: str, title: str | None = None) -> dict[str, Any]:
    url = url.strip()
    if not url:
        raise ValueError("Page URL is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            "INSERT INTO pages (site_id, url, title, created_at) VALUES (?, ?, ?, ?)",
            (site_id, url, clean_text(title), created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM pages WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def create_keyword(
    project_id: int,
    keyword: str,
    site_id: int | None = None,
    page_id: int | None = None,
    intent: str | None = None,
    priority: str | None = None,
) -> dict[str, Any]:
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("Keyword is required")
    created_at = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO keywords (project_id, site_id, page_id, keyword, intent, priority, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (project_id, site_id, page_id, keyword, clean_text(intent), clean_text(priority), created_at),
        )
        return row_to_dict(con.execute("SELECT * FROM keywords WHERE id = ?", (cur.lastrowid,)).fetchone()) or {}


def assign_run(run_id: int, project_id: int | None, site_id: int | None, page_id: int | None, keyword_id: int | None) -> dict[str, Any]:
    with connect() as con:
        con.execute(
            "UPDATE runs SET project_id = ?, site_id = ?, page_id = ?, keyword_id = ? WHERE id = ?",
            (project_id, site_id, page_id, keyword_id, run_id),
        )
        row = con.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
    if not row:
        raise ValueError("Run not found")
    return row_to_dict(row) or {}


def mask_secret(value: str) -> str:
    if len(value) <= 8:
        return "*" * len(value)
    return f"{value[:4]}...{value[-4:]}"


def api_key_public(row: sqlite3.Row) -> dict[str, Any]:
    data = row_to_dict(row) or {}
    key_value = data.pop("key_value", "") or ""
    data["key_preview"] = mask_secret(key_value)
    data["key_length"] = len(key_value)
    return data


def create_api_key(provider: str, label: str, key_value: str, notes: str | None = None) -> dict[str, Any]:
    provider = provider.strip()
    label = label.strip()
    key_value = key_value.strip()
    if not provider:
        raise ValueError("Provider is required")
    if not label:
        raise ValueError("Label is required")
    if not key_value:
        raise ValueError("API key is required")
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        cur = con.execute(
            """
            INSERT INTO api_keys (provider, label, key_value, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (provider, label, key_value, clean_text(notes), now, now),
        )
        row = con.execute("SELECT * FROM api_keys WHERE id = ?", (cur.lastrowid,)).fetchone()
    return api_key_public(row)


def delete_api_key(key_id: int) -> None:
    with connect() as con:
        cur = con.execute("DELETE FROM api_keys WHERE id = ?", (key_id,))
    if cur.rowcount == 0:
        raise ValueError("API key not found")


def create_content_plan(
    project_id: int,
    title: str,
    site_id: int | None = None,
    page_id: int | None = None,
    keyword_id: int | None = None,
    content_type: str | None = None,
    intent: str | None = None,
    priority: str | None = None,
    status: str | None = None,
    due_date: str | None = None,
    notes: str | None = None,
) -> dict[str, Any]:
    title = title.strip()
    if not title:
        raise ValueError("Content plan title is required")
    now = datetime.now().isoformat(timespec="seconds")
    plan_status = clean_text(status) or "planned"
    with connect() as con:
        project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("Project not found")
        cur = con.execute(
            """
            INSERT INTO content_plans
            (project_id, site_id, page_id, keyword_id, title, content_type, intent, priority, status, due_date, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                site_id,
                page_id,
                keyword_id,
                title,
                clean_text(content_type),
                clean_text(intent),
                clean_text(priority),
                plan_status,
                clean_text(due_date),
                clean_text(notes),
                now,
                now,
            ),
        )
        row = con.execute("SELECT * FROM content_plans WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row) or {}


def numeric_delta(base: Any, compare: Any) -> float | None:
    left = as_float(base)
    right = as_float(compare)
    if left is None or right is None:
        return None
    return right - left


def target_rank_for_run(con: sqlite3.Connection, run: sqlite3.Row) -> float | None:
    domain = clean_text(run["target_domain"])
    url = clean_text(run["target_url"])
    if not domain and url:
        domain = domain_from_url(url)
    if not domain:
        return None
    like_domain = f"%{domain}%"
    row = con.execute(
        """
        SELECT COALESCE(avg_rank, rank) AS rank_value
        FROM serp_results
        WHERE run_id = ?
          AND (host LIKE ? OR url LIKE ?)
        ORDER BY COALESCE(avg_rank, rank)
        LIMIT 1
        """,
        (run["id"], like_domain, like_domain),
    ).fetchone()
    return as_float(row["rank_value"]) if row else None


def compare_runs(base_id: int, compare_id: int) -> dict[str, Any]:
    if base_id == compare_id:
        raise ValueError("Choose two different runs to compare")
    with connect() as con:
        base = con.execute("SELECT * FROM runs WHERE id = ?", (base_id,)).fetchone()
        compare = con.execute("SELECT * FROM runs WHERE id = ?", (compare_id,)).fetchone()
        if not base or not compare:
            missing = []
            if not base:
                missing.append(str(base_id))
            if not compare:
                missing.append(str(compare_id))
            raise ValueError(
                "One of the selected comparison runs is not in the database. "
                f"Missing run id(s): {', '.join(missing)}. Refresh the run list and choose two imported reports."
            )

        base_summary = row_to_dict(base) or {}
        compare_summary = row_to_dict(compare) or {}
        base_target_rank = target_rank_for_run(con, base)
        compare_target_rank = target_rank_for_run(con, compare)
        counts = {}
        for name, table in {
            "serp_results": "serp_results",
            "recommendations": "recommendations",
            "lsi_keywords": "lsi_keywords",
            "workbook_rows": "workbook_rows",
        }.items():
            base_count = con.execute(f"SELECT COUNT(*) FROM {table} WHERE run_id = ?", (base_id,)).fetchone()[0]
            compare_count = con.execute(f"SELECT COUNT(*) FROM {table} WHERE run_id = ?", (compare_id,)).fetchone()[0]
            counts[name] = {"base": base_count, "compare": compare_count, "delta": compare_count - base_count}

        base_serp = {
            (r["url"] or r["host"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM serp_results WHERE run_id = ? AND (url IS NOT NULL OR host IS NOT NULL)",
                (base_id,),
            ).fetchall()
        }
        compare_serp = {
            (r["url"] or r["host"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM serp_results WHERE run_id = ? AND (url IS NOT NULL OR host IS NOT NULL)",
                (compare_id,),
            ).fetchall()
        }
        serp_changes = []
        for key in sorted(set(base_serp) | set(compare_serp)):
            before = base_serp.get(key)
            after = compare_serp.get(key)
            base_rank = as_float((before or {}).get("avg_rank") or (before or {}).get("rank"))
            compare_rank = as_float((after or {}).get("avg_rank") or (after or {}).get("rank"))
            if before and after:
                status = "changed" if base_rank != compare_rank else "same"
            else:
                status = "new" if after else "removed"
            if status == "same":
                continue
            serp_changes.append(
                {
                    "status": status,
                    "host": (after or before or {}).get("host"),
                    "title": (after or before or {}).get("title"),
                    "url": (after or before or {}).get("url"),
                    "base_rank": base_rank,
                    "compare_rank": compare_rank,
                    "rank_delta": numeric_delta(base_rank, compare_rank),
                }
            )
        serp_changes.sort(key=lambda r: (abs(r["rank_delta"] or 999), r["status"]), reverse=True)

        base_recs = {
            (r["sheet"], r["factor_id"] or "", r["factor"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM recommendations WHERE run_id = ? AND recommendation IS NOT NULL",
                (base_id,),
            ).fetchall()
        }
        compare_recs = {
            (r["sheet"], r["factor_id"] or "", r["factor"] or ""): row_to_dict(r)
            for r in con.execute(
                "SELECT * FROM recommendations WHERE run_id = ? AND recommendation IS NOT NULL",
                (compare_id,),
            ).fetchall()
        }
        recommendation_changes = []
        for key in sorted(set(base_recs) | set(compare_recs)):
            before = base_recs.get(key)
            after = compare_recs.get(key)
            if before and after:
                changed = (
                    before.get("current_value") != after.get("current_value")
                    or before.get("goal") != after.get("goal")
                    or before.get("recommendation") != after.get("recommendation")
                    or before.get("percent") != after.get("percent")
                )
                if not changed:
                    continue
                status = "changed"
            else:
                status = "new" if after else "resolved"
            source = after or before or {}
            recommendation_changes.append(
                {
                    "status": status,
                    "sheet": source.get("sheet"),
                    "factor_id": source.get("factor_id"),
                    "factor": source.get("factor"),
                    "base_current": (before or {}).get("current_value"),
                    "compare_current": (after or {}).get("current_value"),
                    "base_goal": (before or {}).get("goal"),
                    "compare_goal": (after or {}).get("goal"),
                    "base_percent": (before or {}).get("percent"),
                    "compare_percent": (after or {}).get("percent"),
                    "recommendation": source.get("recommendation"),
                }
            )

        base_lsi = {
            r["keyword"]: row_to_dict(r)
            for r in con.execute("SELECT * FROM lsi_keywords WHERE run_id = ?", (base_id,)).fetchall()
        }
        compare_lsi = {
            r["keyword"]: row_to_dict(r)
            for r in con.execute("SELECT * FROM lsi_keywords WHERE run_id = ?", (compare_id,)).fetchall()
        }
        lsi_changes = []
        for keyword in sorted(set(base_lsi) | set(compare_lsi)):
            before = base_lsi.get(keyword)
            after = compare_lsi.get(keyword)
            if before and after:
                deficit_delta = numeric_delta(before.get("deficit"), after.get("deficit"))
                tracked_delta = numeric_delta(before.get("tracked_value"), after.get("tracked_value"))
                if deficit_delta == 0 and tracked_delta == 0:
                    continue
                status = "changed"
            else:
                deficit_delta = None
                tracked_delta = None
                status = "new" if after else "removed"
            lsi_changes.append(
                {
                    "status": status,
                    "keyword": keyword,
                    "base_tracked": (before or {}).get("tracked_value"),
                    "compare_tracked": (after or {}).get("tracked_value"),
                    "tracked_delta": tracked_delta,
                    "base_deficit": (before or {}).get("deficit"),
                    "compare_deficit": (after or {}).get("deficit"),
                    "deficit_delta": deficit_delta,
                }
            )
        lsi_changes.sort(key=lambda r: abs(r["deficit_delta"] or r["tracked_delta"] or 0), reverse=True)

    return {
        "base_run": base_summary,
        "compare_run": compare_summary,
        "summary": {
            "target_rank": {
                "base": base_target_rank,
                "compare": compare_target_rank,
                "delta": numeric_delta(base_target_rank, compare_target_rank),
            },
            "counts": counts,
            "serp_change_count": len(serp_changes),
            "recommendation_change_count": len(recommendation_changes),
            "lsi_change_count": len(lsi_changes),
        },
        "serp_changes": serp_changes[:200],
        "recommendation_changes": recommendation_changes[:300],
        "lsi_changes": lsi_changes[:300],
    }


class AppHandler(BaseHTTPRequestHandler):
    server_version = "CoraDashboard/0.1"

    def log_message(self, fmt: str, *args: Any) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)
        try:
            if path == "/":
                self.serve_file(STATIC_DIR / "index.html")
            elif path.startswith("/static/"):
                self.serve_file(STATIC_DIR / unquote(path.removeprefix("/static/")))
            elif path == "/api/overview":
                self.api_overview()
            elif path == "/api/content-plans":
                self.api_content_plans()
            elif path == "/api/runs":
                self.api_runs(query)
            elif path == "/api/compare":
                self.api_compare(query)
            elif re.match(r"^/api/runs/\d+$", path):
                self.api_run_detail(int(path.rsplit("/", 1)[1]))
            elif re.match(r"^/api/runs/\d+/download$", path):
                self.download_run(int(path.split("/")[3]))
            elif path == "/api/latest-report":
                latest = latest_xlsx()
                json_response(self, {"path": str(latest) if latest else None})
            elif path == "/api/cora/status":
                json_response(self, query_cora("/api/status"))
            elif path == "/api/jobs":
                self.api_jobs()
            elif re.match(r"^/api/jobs/\d+$", path):
                self.api_job_detail(int(path.rsplit("/", 1)[1]))
            elif re.match(r"^/api/runs/\d+/workbook$", path):
                self.api_run_workbook(int(path.split("/")[3]), query)
            elif path == "/api/projects":
                self.api_projects()
            elif re.match(r"^/api/projects/\d+$", path):
                self.api_project_detail(int(path.rsplit("/", 1)[1]))
            elif path == "/api/api-keys":
                self.api_keys()
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            body = read_json_body(self)
            if parsed.path == "/api/ingest":
                report_path = body.get("path") or (str(latest_xlsx()) if latest_xlsx() else None)
                if not report_path:
                    error_response(self, "No report path provided and no .xlsx found in the user folder", 400)
                    return
                result = ingest_report(
                    Path(report_path),
                    target_url=body.get("target_url"),
                    keyword=body.get("keyword"),
                    notes=body.get("notes"),
                )
                json_response(self, result)
            elif parsed.path == "/api/jobs":
                job = create_managed_job(body.get("keyword", ""), body.get("target_url", ""))
                json_response(self, {"job": job}, 201)
            elif parsed.path == "/api/cora/stop":
                result = force_stop_cora(body.get("reason") or "Manual dashboard force stop")
                json_response(self, result)
            elif parsed.path == "/api/backfill-workbook":
                result = backfill_workbook_rows(body.get("run_id"))
                json_response(self, result)
            elif parsed.path == "/api/projects":
                project = create_project(
                    body.get("name", ""),
                    body.get("client"),
                    body.get("site_domain"),
                    body.get("notes"),
                )
                json_response(self, {"project": project}, 201)
            elif parsed.path == "/api/content-plans":
                plan = create_content_plan(
                    int(body.get("project_id")),
                    body.get("title", ""),
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    int(body["keyword_id"]) if body.get("keyword_id") else None,
                    body.get("content_type"),
                    body.get("intent"),
                    body.get("priority"),
                    body.get("status"),
                    body.get("due_date"),
                    body.get("notes"),
                )
                json_response(self, {"content_plan": plan}, 201)
            elif parsed.path == "/api/sites":
                site = create_site(int(body.get("project_id")), body.get("domain", ""), body.get("name"))
                json_response(self, {"site": site}, 201)
            elif parsed.path == "/api/pages":
                page = create_page(int(body.get("site_id")), body.get("url", ""), body.get("title"))
                json_response(self, {"page": page}, 201)
            elif parsed.path == "/api/keywords":
                keyword = create_keyword(
                    int(body.get("project_id")),
                    body.get("keyword", ""),
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    body.get("intent"),
                    body.get("priority"),
                )
                json_response(self, {"keyword": keyword}, 201)
            elif re.match(r"^/api/runs/\d+/assign$", parsed.path):
                run = assign_run(
                    int(parsed.path.split("/")[3]),
                    int(body["project_id"]) if body.get("project_id") else None,
                    int(body["site_id"]) if body.get("site_id") else None,
                    int(body["page_id"]) if body.get("page_id") else None,
                    int(body["keyword_id"]) if body.get("keyword_id") else None,
                )
                json_response(self, {"run": run})
            elif parsed.path == "/api/api-keys":
                key = create_api_key(
                    body.get("provider", ""),
                    body.get("label", ""),
                    body.get("key_value", ""),
                    body.get("notes"),
                )
                json_response(self, {"api_key": key}, 201)
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        try:
            if re.match(r"^/api/api-keys/\d+$", parsed.path):
                delete_api_key(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"deleted": True})
            else:
                error_response(self, "Not found", 404)
        except Exception as exc:
            error_response(self, str(exc), 500)

    def serve_file(self, path: Path) -> None:
        path = path.resolve()
        if not str(path).startswith(str(STATIC_DIR.resolve())) or not path.exists() or not path.is_file():
            error_response(self, "Not found", 404)
            return
        ctype = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_overview(self) -> None:
        with connect() as con:
            counts = {
                "projects": con.execute("SELECT COUNT(*) FROM projects").fetchone()[0],
                "sites": con.execute("SELECT COUNT(*) FROM sites").fetchone()[0],
                "pages": con.execute("SELECT COUNT(*) FROM pages").fetchone()[0],
                "keywords": con.execute("SELECT COUNT(*) FROM keywords").fetchone()[0],
                "runs": con.execute("SELECT COUNT(*) FROM runs").fetchone()[0],
                "api_keys": con.execute("SELECT COUNT(*) FROM api_keys").fetchone()[0],
                "content_plans": con.execute("SELECT COUNT(*) FROM content_plans").fetchone()[0],
                "workbook_rows": con.execute("SELECT COUNT(*) FROM workbook_rows").fetchone()[0],
            }
            job_counts = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT status, COUNT(*) AS count
                    FROM managed_jobs
                    GROUP BY status
                    ORDER BY status
                    """
                ).fetchall()
            ]
            recent_runs = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT r.id, r.keyword, r.target_url, r.target_domain, r.imported_at, r.file_name,
                           p.name AS project_name,
                           (SELECT COUNT(*) FROM recommendations rec WHERE rec.run_id = r.id) AS recommendation_count,
                           (SELECT COUNT(*) FROM lsi_keywords lsi WHERE lsi.run_id = r.id) AS lsi_count
                    FROM runs r
                    LEFT JOIN projects p ON p.id = r.project_id
                    ORDER BY r.imported_at DESC, r.id DESC
                    LIMIT 8
                    """
                ).fetchall()
            ]
            recent_jobs = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT id, keyword, target_domain, status, status_message, started_at, completed_at, imported_run_id
                    FROM managed_jobs
                    ORDER BY started_at DESC, id DESC
                    LIMIT 8
                    """
                ).fetchall()
            ]
            api_key_providers = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT provider, COUNT(*) AS count
                    FROM api_keys
                    GROUP BY provider
                    ORDER BY provider COLLATE NOCASE
                    """
                ).fetchall()
            ]
            recent_content_plans = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT cp.id, cp.title, cp.content_type, cp.intent, cp.priority, cp.status, cp.due_date,
                           p.name AS project_name, k.keyword
                    FROM content_plans cp
                    JOIN projects p ON p.id = cp.project_id
                    LEFT JOIN keywords k ON k.id = cp.keyword_id
                    ORDER BY cp.updated_at DESC, cp.id DESC
                    LIMIT 8
                    """
                ).fetchall()
            ]
        json_response(
            self,
            {
                "counts": counts,
                "job_counts": job_counts,
                "recent_runs": recent_runs,
                "recent_jobs": recent_jobs,
                "api_key_providers": api_key_providers,
                "recent_content_plans": recent_content_plans,
            },
        )

    def api_runs(self, query: dict[str, list[str]]) -> None:
        search = (query.get("q") or [""])[0].strip()
        params: list[Any] = []
        where = ""
        if search:
            where = "WHERE r.keyword LIKE ? OR r.target_domain LIKE ? OR r.file_name LIKE ? OR p.name LIKE ?"
            like = f"%{search}%"
            params.extend([like, like, like, like])
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT r.*,
                       p.name AS project_name,
                       s.domain AS site_domain_assigned,
                       pg.url AS page_url,
                       k.keyword AS assigned_keyword,
                       (SELECT COUNT(*) FROM serp_results s WHERE s.run_id = r.id) AS result_count,
                       (SELECT COUNT(*) FROM recommendations rec WHERE rec.run_id = r.id) AS recommendation_count
                FROM runs r
                LEFT JOIN projects p ON p.id = r.project_id
                LEFT JOIN sites s ON s.id = r.site_id
                LEFT JOIN pages pg ON pg.id = r.page_id
                LEFT JOIN keywords k ON k.id = r.keyword_id
                {where}
                ORDER BY imported_at DESC, id DESC
                LIMIT 200
                """,
                params,
            ).fetchall()
        json_response(self, {"runs": [row_to_dict(r) for r in rows]})

    def api_jobs(self) -> None:
        with connect() as con:
            rows = con.execute(
                """
                SELECT j.*, r.file_name AS imported_file_name
                FROM managed_jobs j
                LEFT JOIN runs r ON r.id = j.imported_run_id
                ORDER BY j.started_at DESC, j.id DESC
                LIMIT 100
                """
            ).fetchall()
        json_response(self, {"jobs": [row_to_dict(r) for r in rows]})

    def api_job_detail(self, job_id: int) -> None:
        with connect() as con:
            row = con.execute(
                """
                SELECT j.*, r.file_name AS imported_file_name
                FROM managed_jobs j
                LEFT JOIN runs r ON r.id = j.imported_run_id
                WHERE j.id = ?
                """,
                (job_id,),
            ).fetchone()
        if not row:
            error_response(self, "Job not found", 404)
            return
        json_response(self, {"job": row_to_dict(row)})

    def api_run_workbook(self, run_id: int, query: dict[str, list[str]]) -> None:
        sheet = (query.get("sheet") or [""])[0].strip()
        with connect() as con:
            if sheet:
                rows = con.execute(
                    """
                    SELECT sheet, row_index, column_count, row_json
                    FROM workbook_rows
                    WHERE run_id = ? AND sheet = ?
                    ORDER BY row_index
                    LIMIT 1000
                    """,
                    (run_id, sheet),
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT sheet, COUNT(*) AS row_count, MAX(column_count) AS max_columns
                    FROM workbook_rows
                    WHERE run_id = ?
                    GROUP BY sheet
                    ORDER BY sheet
                    """,
                    (run_id,),
                ).fetchall()
        json_response(self, {"rows": [row_to_dict(r) for r in rows]})

    def api_keys(self) -> None:
        with connect() as con:
            rows = con.execute(
                """
                SELECT * FROM api_keys
                ORDER BY provider COLLATE NOCASE, label COLLATE NOCASE
                """
            ).fetchall()
        json_response(self, {"api_keys": [api_key_public(r) for r in rows]})

    def api_content_plans(self) -> None:
        with connect() as con:
            rows = con.execute(
                """
                SELECT cp.*,
                       p.name AS project_name,
                       s.domain AS site_domain,
                       pg.url AS page_url,
                       k.keyword
                FROM content_plans cp
                JOIN projects p ON p.id = cp.project_id
                LEFT JOIN sites s ON s.id = cp.site_id
                LEFT JOIN pages pg ON pg.id = cp.page_id
                LEFT JOIN keywords k ON k.id = cp.keyword_id
                ORDER BY
                    CASE cp.status
                        WHEN 'planned' THEN 1
                        WHEN 'drafting' THEN 2
                        WHEN 'review' THEN 3
                        WHEN 'published' THEN 4
                        ELSE 5
                    END,
                    cp.due_date IS NULL,
                    cp.due_date,
                    cp.updated_at DESC
                LIMIT 300
                """
            ).fetchall()
        json_response(self, {"content_plans": [row_to_dict(r) for r in rows]})

    def api_compare(self, query: dict[str, list[str]]) -> None:
        base_id = int((query.get("base_id") or ["0"])[0])
        compare_id = int((query.get("compare_id") or ["0"])[0])
        if not base_id or not compare_id:
            error_response(self, "base_id and compare_id are required", 400)
            return
        json_response(self, compare_runs(base_id, compare_id))

    def api_projects(self) -> None:
        with connect() as con:
            rows = con.execute(
                """
                SELECT p.*,
                       (SELECT COUNT(*) FROM sites s WHERE s.project_id = p.id) AS site_count,
                       (SELECT COUNT(*) FROM pages pg JOIN sites s ON s.id = pg.site_id WHERE s.project_id = p.id) AS page_count,
                       (SELECT COUNT(*) FROM keywords k WHERE k.project_id = p.id) AS keyword_count,
                       (SELECT COUNT(*) FROM runs r WHERE r.project_id = p.id) AS run_count,
                       (SELECT MAX(r.imported_at) FROM runs r WHERE r.project_id = p.id) AS last_run_at
                FROM projects p
                ORDER BY p.name COLLATE NOCASE
                """
            ).fetchall()
        json_response(self, {"projects": [row_to_dict(r) for r in rows]})

    def api_project_detail(self, project_id: int) -> None:
        with connect() as con:
            project = row_to_dict(con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone())
            if not project:
                error_response(self, "Project not found", 404)
                return
            sites = [
                row_to_dict(r)
                for r in con.execute(
                    "SELECT * FROM sites WHERE project_id = ? ORDER BY domain COLLATE NOCASE",
                    (project_id,),
                ).fetchall()
            ]
            pages = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT pg.*, s.domain AS site_domain
                    FROM pages pg
                    JOIN sites s ON s.id = pg.site_id
                    WHERE s.project_id = ?
                    ORDER BY s.domain COLLATE NOCASE, pg.url COLLATE NOCASE
                    """,
                    (project_id,),
                ).fetchall()
            ]
            keywords = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT k.*, s.domain AS site_domain, pg.url AS page_url
                    FROM keywords k
                    LEFT JOIN sites s ON s.id = k.site_id
                    LEFT JOIN pages pg ON pg.id = k.page_id
                    WHERE k.project_id = ?
                    ORDER BY k.keyword COLLATE NOCASE
                    """,
                    (project_id,),
                ).fetchall()
            ]
            runs = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT r.*, s.domain AS site_domain_assigned, pg.url AS page_url, k.keyword AS assigned_keyword
                    FROM runs r
                    LEFT JOIN sites s ON s.id = r.site_id
                    LEFT JOIN pages pg ON pg.id = r.page_id
                    LEFT JOIN keywords k ON k.id = r.keyword_id
                    WHERE r.project_id = ?
                    ORDER BY r.imported_at DESC, r.id DESC
                    LIMIT 100
                    """,
                    (project_id,),
                ).fetchall()
            ]
        json_response(
            self,
            {"project": project, "sites": sites, "pages": pages, "keywords": keywords, "runs": runs},
        )

    def api_run_detail(self, run_id: int) -> None:
        with connect() as con:
            run = row_to_dict(
                con.execute(
                    """
                    SELECT r.*,
                           p.name AS project_name,
                           s.domain AS site_domain_assigned,
                           pg.url AS page_url,
                           k.keyword AS assigned_keyword
                    FROM runs r
                    LEFT JOIN projects p ON p.id = r.project_id
                    LEFT JOIN sites s ON s.id = r.site_id
                    LEFT JOIN pages pg ON pg.id = r.page_id
                    LEFT JOIN keywords k ON k.id = r.keyword_id
                    WHERE r.id = ?
                    """,
                    (run_id,),
                ).fetchone()
            )
            if not run:
                error_response(self, "Run not found", 404)
                return
            results = [
                row_to_dict(r)
                for r in con.execute(
                    "SELECT * FROM serp_results WHERE run_id = ? ORDER BY rank LIMIT 100", (run_id,)
                ).fetchall()
            ]
            recommendations = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT * FROM recommendations
                    WHERE run_id = ? AND recommendation IS NOT NULL
                    ORDER BY CASE WHEN percent IS NULL THEN 2 ELSE 1 END, percent ASC
                    LIMIT 200
                    """,
                    (run_id,),
                ).fetchall()
            ]
            lsi = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT * FROM lsi_keywords
                    WHERE run_id = ?
                    ORDER BY ABS(COALESCE(deficit, 0)) DESC, ABS(COALESCE(best_of_both, 0)) DESC
                    LIMIT 100
                    """,
                    (run_id,),
                ).fetchall()
            ]
        json_response(self, {"run": run, "results": results, "recommendations": recommendations, "lsi": lsi})

    def download_run(self, run_id: int) -> None:
        with connect() as con:
            run = con.execute("SELECT archive_path, file_name FROM runs WHERE id = ?", (run_id,)).fetchone()
        if not run:
            error_response(self, "Run not found", 404)
            return
        path = Path(run["archive_path"])
        if not path.exists():
            error_response(self, "Archived file missing", 404)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{run["file_name"]}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def start_server(port: int = DEFAULT_PORT) -> ThreadingHTTPServer:
    init_db()
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


def main(argv: list[str]) -> int:
    init_db()
    if len(argv) > 1 and argv[1] == "ingest":
        path = Path(argv[2]) if len(argv) > 2 else latest_xlsx()
        if not path:
            print("No .xlsx file found to import.", file=sys.stderr)
            return 1
        target_url = argv[3] if len(argv) > 3 else None
        result = ingest_report(path, target_url=target_url)
        print(json.dumps(result, indent=2, default=str))
        return 0

    if len(argv) > 1 and argv[1] == "backfill-workbook":
        run_id = int(argv[2]) if len(argv) > 2 else None
        result = backfill_workbook_rows(run_id)
        print(json.dumps(result, indent=2, default=str))
        return 0

    port = int(argv[1]) if len(argv) > 1 else DEFAULT_PORT
    server = start_server(port)
    print(f"Cora dashboard running at http://127.0.0.1:{port}/")
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        server.shutdown()
        return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
